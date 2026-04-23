# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os, json, shutil, subprocess, sys, tempfile, threading, re
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from spellchecker import SpellChecker
except ImportError:
    SpellChecker = None

# ── Windows OLE drag-and-drop ─────────────────────────────────────────────────
_dnd_targets = {}

if sys.platform == "win32":
    import ctypes, ctypes.wintypes

    ole32  = ctypes.windll.ole32
    shell32= ctypes.windll.shell32

    DROPEFFECT_NONE = 0
    DROPEFFECT_COPY = 1
    CF_HDROP        = 15
    GMEM_MOVEABLE   = 0x0002

    CF_FILEDESCRIPTORW = ctypes.windll.user32.RegisterClipboardFormatW("FileGroupDescriptorW")
    CF_FILECONTENTS    = ctypes.windll.user32.RegisterClipboardFormatW("FileContents")

    class FORMATETC(ctypes.Structure):
        _fields_ = [("cfFormat",ctypes.c_ushort),("ptd",ctypes.c_void_p),
                    ("dwAspect",ctypes.c_ulong),("lindex",ctypes.c_long),
                    ("tymed",ctypes.c_ulong)]

    class STGMEDIUM(ctypes.Structure):
        _fields_ = [("tymed",ctypes.c_ulong),("hGlobal",ctypes.c_void_p),
                    ("pUnkForRelease",ctypes.c_void_p)]

    TYMED_HGLOBAL   = 1
    TYMED_ISTREAM   = 2
    DVASPECT_CONTENT= 1

    class FILEDESCRIPTORW(ctypes.Structure):
        _fields_ = [("dwFlags",ctypes.c_ulong),
                    ("clsid",ctypes.c_byte*16),
                    ("sizel",ctypes.c_long*2),
                    ("pointl",ctypes.c_long*2),
                    ("dwFileAttributes",ctypes.c_ulong),
                    ("ftCreationTime",ctypes.c_ulonglong),
                    ("ftLastAccessTime",ctypes.c_ulonglong),
                    ("ftLastWriteTime",ctypes.c_ulonglong),
                    ("nFileSizeHigh",ctypes.c_ulong),
                    ("nFileSizeLow",ctypes.c_ulong),
                    ("cFileName",ctypes.c_wchar*260)]

    class FILEGROUPDESCRIPTORW(ctypes.Structure):
        _fields_ = [("cItems",ctypes.c_ulong),
                    ("fgd",FILEDESCRIPTORW*1)]

    IDropTarget_vtbl = [
        ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p,
                           ctypes.POINTER(ctypes.c_byte*16), ctypes.c_void_p),
        ctypes.WINFUNCTYPE(ctypes.c_ulong, ctypes.c_void_p),
        ctypes.WINFUNCTYPE(ctypes.c_ulong, ctypes.c_void_p),
        ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p, ctypes.c_void_p,
                           ctypes.c_ulong, ctypes.c_ulonglong,
                           ctypes.POINTER(ctypes.c_ulong)),
        ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p, ctypes.c_ulong,
                           ctypes.c_ulonglong,
                           ctypes.POINTER(ctypes.c_ulong)),
        ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p),
        ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p, ctypes.c_void_p,
                           ctypes.c_ulong, ctypes.c_ulonglong,
                           ctypes.POINTER(ctypes.c_ulong)),
    ]

    class _DropTarget(ctypes.Structure):
        _fields_ = [("lpVtbl", ctypes.c_void_p), ("refcount", ctypes.c_ulong)]

        def __init__(self, hwnd, callback, widget):
            super().__init__()
            self.hwnd     = hwnd
            self.callback = callback
            self.widget   = widget
            self.refcount = 1
            self._vtbl_funcs = []
            self._build_vtbl()

        def _build_vtbl(self):
            def QueryInterface(this, riid, ppv): return 0x80004002
            def AddRef(this):
                self.refcount += 1; return self.refcount
            def Release(this):
                self.refcount -= 1; return self.refcount
            def DragEnter(this, pDataObj, grfKeyState, pt, pdwEffect):
                pdwEffect[0] = DROPEFFECT_COPY; return 0
            def DragOver(this, grfKeyState, pt, pdwEffect):
                pdwEffect[0] = DROPEFFECT_COPY; return 0
            def DragLeave(this): return 0
            def Drop(this, pDataObj, grfKeyState, pt, pdwEffect):
                pdwEffect[0] = DROPEFFECT_COPY
                self._handle_drop(pDataObj)
                return 0

            fns = [IDropTarget_vtbl[i](f) for i,f in enumerate(
                [QueryInterface, AddRef, Release, DragEnter, DragOver, DragLeave, Drop])]
            self._vtbl_funcs = fns
            VtblArray = ctypes.c_void_p * len(fns)
            self._vtbl_array = VtblArray(*[ctypes.cast(f, ctypes.c_void_p) for f in fns])
            self.lpVtbl = ctypes.cast(self._vtbl_array, ctypes.c_void_p)

        def _handle_drop(self, pDataObj):
            paths = []

            fmt = FORMATETC()
            fmt.cfFormat  = CF_HDROP
            fmt.dwAspect  = DVASPECT_CONTENT
            fmt.lindex    = -1
            fmt.tymed     = TYMED_HGLOBAL
            med = STGMEDIUM()
            IDataObject = ctypes.cast(pDataObj, ctypes.POINTER(ctypes.c_void_p))
            vtbl_ptr = ctypes.cast(IDataObject[0], ctypes.POINTER(ctypes.c_void_p))
            GetData = ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_void_p,
                                          ctypes.POINTER(FORMATETC),
                                          ctypes.POINTER(STGMEDIUM))
            get_data = GetData(vtbl_ptr[3])
            hr = get_data(pDataObj, ctypes.byref(fmt), ctypes.byref(med))
            if hr == 0 and med.hGlobal:
                try:
                    hg = ctypes.windll.kernel32.GlobalLock(med.hGlobal)
                    if hg:
                        count = shell32.DragQueryFileW(hg, 0xFFFFFFFF, None, 0)
                        buf   = ctypes.create_unicode_buffer(520)
                        for i in range(count):
                            shell32.DragQueryFileW(hg, i, buf, 520)
                            if os.path.isfile(buf.value):
                                paths.append(buf.value)
                        ctypes.windll.kernel32.GlobalUnlock(med.hGlobal)
                    shell32.DragFinish(med.hGlobal)
                except: pass

            if not paths:
                fmt2 = FORMATETC()
                fmt2.cfFormat  = CF_FILEDESCRIPTORW
                fmt2.dwAspect  = DVASPECT_CONTENT
                fmt2.lindex    = -1
                fmt2.tymed     = TYMED_HGLOBAL
                med2 = STGMEDIUM()
                hr2 = get_data(pDataObj, ctypes.byref(fmt2), ctypes.byref(med2))
                if hr2 == 0 and med2.hGlobal:
                    try:
                        hg2 = ctypes.windll.kernel32.GlobalLock(med2.hGlobal)
                        fgd = ctypes.cast(hg2, ctypes.POINTER(FILEGROUPDESCRIPTORW))
                        n   = fgd[0].cItems
                        class FGD(ctypes.Structure):
                            _fields_=[("cItems",ctypes.c_ulong),
                                      ("fgd",FILEDESCRIPTORW*n)]
                        fgd2 = ctypes.cast(hg2, ctypes.POINTER(FGD))[0]
                        tmp  = tempfile.mkdtemp(prefix="qt_attach_")
                        for i in range(n):
                            fname = fgd2.fgd[i].cFileName
                            if not fname: continue
                            fmt3 = FORMATETC()
                            fmt3.cfFormat = CF_FILECONTENTS
                            fmt3.dwAspect = DVASPECT_CONTENT
                            fmt3.lindex   = i
                            fmt3.tymed    = TYMED_ISTREAM
                            med3 = STGMEDIUM()
                            hr3  = get_data(pDataObj, ctypes.byref(fmt3), ctypes.byref(med3))
                            if hr3 == 0 and med3.hGlobal:
                                try:
                                    dest = os.path.join(tmp, fname)
                                    IStream = ctypes.cast(med3.hGlobal,
                                                          ctypes.POINTER(ctypes.c_void_p))
                                    sv = IStream[0]
                                    sv_ptr = ctypes.cast(sv, ctypes.POINTER(ctypes.c_void_p))
                                    Read = ctypes.WINFUNCTYPE(ctypes.c_long,
                                                              ctypes.c_void_p,
                                                              ctypes.c_void_p,
                                                              ctypes.c_ulong,
                                                              ctypes.POINTER(ctypes.c_ulong))
                                    read_fn = Read(sv_ptr[3])
                                    chunks=[]; read_bytes=ctypes.c_ulong(0)
                                    while True:
                                        buf2=ctypes.create_string_buffer(65536)
                                        hr4=read_fn(med3.hGlobal,buf2,65536,
                                                    ctypes.byref(read_bytes))
                                        if read_bytes.value==0: break
                                        chunks.append(bytes(buf2.raw[:read_bytes.value]))
                                        if hr4!=0: break
                                    if chunks:
                                        with open(dest,"wb") as f: f.write(b"".join(chunks))
                                        paths.append(dest)
                                except: pass
                        ctypes.windll.kernel32.GlobalUnlock(med2.hGlobal)
                    except: pass

            if paths:
                self.widget.after(0, lambda p=paths: self.callback(p))

def _setup_dnd(toplevel, callback):
    if sys.platform != "win32":
        return False
    try:
        ole32.OleInitialize(None)
        hwnd = toplevel.winfo_id()
        target = _DropTarget(hwnd, callback, toplevel)
        _dnd_targets[hwnd] = target
        hr = ole32.RegisterDragDrop(hwnd, ctypes.byref(target))
        return hr == 0
    except Exception as e:
        print(f"DnD setup failed: {e}")
        return False

# Directory the exe (or script) lives in — used for user data/config paths.
APP_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
# Directory bundled resources are loaded from when packaged with PyInstaller.
RESOURCE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

# Store user-specific settings locally so teammates do not overwrite each
# other's name/path settings when they all run the same shared exe.
CONFIG_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "QBOX")
CONFIG_FILE = os.path.join(CONFIG_DIR, "config.json")
LEGACY_CONFIG_FILE = os.path.join(APP_DIR, "config.json")

# Default data paths — created automatically so the team never has to browse
DEFAULT_DATA_DIR    = os.path.join(APP_DIR, "Data")
DEFAULT_EXCEL_FILE  = os.path.join(DEFAULT_DATA_DIR, "query_tracker.xlsx")
DEFAULT_BACKUP_DIR  = os.path.join(DEFAULT_DATA_DIR, "Backups")
DEFAULT_ATTACH_DIR  = os.path.join(APP_DIR, "ATTACHMENTS")

def _ensure_app_folders():
    """Create Data/, Data/Backups/, and ATTACHMENTS/ next to the exe on first run."""
    for folder in (DEFAULT_DATA_DIR, DEFAULT_BACKUP_DIR, DEFAULT_ATTACH_DIR):
        os.makedirs(folder, exist_ok=True)

def _ensure_config_folder():
    os.makedirs(CONFIG_DIR, exist_ok=True)

def resource_path(*parts):
    return os.path.join(RESOURCE_DIR, *parts)

_ensure_app_folders()
_ensure_config_folder()

DEFAULT_QUERY_TYPES = [
    "Sub-Meter reads","Tenancy change","Recharge rework","New instruction",
    "Quote request","Methodology query","Consumption","Missing AMR data",
    "Missing supplier invoice","Other"
]
STATUSES   = ["Open","In Progress","Pending info","Resolved"]
PRIORITIES = ["High","Medium","Low"]
UTILITY_OPTIONS = ["Electricity","Gas","Water","Heat Network","Cooling","Other"]
COLS = ["ID","Reference","Client","Fund","Site","Utility Type","Specific Meter","Type","Status",
        "Priority","Description","Opened","Action Date","Resolved Date","Activity Log",
        "Site Address","Supply Point ID","Meter Serial","Managing Agent Contact","Property Code",
        "Last Updated By","Last Updated Date","Assigned To","Query Raised Date"]

# QUERY_TYPES is populated after load_config is defined — see below
QUERY_TYPES = list(DEFAULT_QUERY_TYPES)

# Default workload warning threshold for calendar scheduling.
HIGH_VOLUME_DAY_THRESHOLD = 5

# Hidden marker used to store machine-readable pushback events in activity logs.
SYS_PUSHBACK_TAG = "[SYS_PUSHBACK]"

SPELLCHECK_WORD_RE = re.compile(r"[A-Za-z][A-Za-z'-]{2,}")
SPELLCHECK_SKIP_WORDS = {
    "amr", "spid", "spids", "qbox", "inteb", "recharge", "recharges",
    "tenancy", "tenancies", "meter", "meters", "submeter", "submeters",
    "query", "queries", "site", "sites", "fund", "funds", "prop",
    "props", "utility", "utilities", "serial", "serials",
}
if SpellChecker:
    try:
        SPELLCHECKER = SpellChecker(language="en")
    except Exception:
        SPELLCHECKER = None
else:
    SPELLCHECKER = None

if SPELLCHECKER is None:
    import atexit
    _spell_warning_shown = [False]
    def _show_spell_install_hint():
        if not _spell_warning_shown[0]:
            _spell_warning_shown[0] = True
            print("\nSpell check is unavailable for this build. Install/rebuild with pyspellchecker data included.\n")
    atexit.register(_show_spell_install_hint)

# ── Theme definitions ────────────────────────────────────────────────────────
# ── Theme definitions ────────────────────────────────────────────────────────
# All dark themes use the same status/priority pill system — only the accent
# and background tones change. Text contrast is kept WCAG AA compliant.
THEMES = {
    "Slate & Teal": {
        "BG":"#0F1419","CARD":"#161D26","CARD2":"#1C2530","NAV":"#0B1017","NAV2":"#111820",
        "BORDER":"#243040","ACCENT":"#2DD4BF","ACCENT2":"#5EEAD4",
        "SUCCESS":"#34D399","WARNING":"#FBBF24","DANGER":"#F87171",
        "MUTED":"#4A6070","TEXT":"#E2EAF0","TEXT2":"#7A92A8","PURPLE":"#A78BFA",
        "NAV_TEXT":"#E7F0F6","NAV_MUTED":"#93A9BB","PRIMARY_FG":"#081310",
        "S_COLORS":{
            "Open":         ("#5EEAD4","#0D2E2A"),
            "In Progress":  ("#FDE68A","#2E2200"),
            "Pending info": ("#A0B4C8","#1A2535"),
            "Resolved":     ("#6EE7B7","#0A2818"),
        },
        "P_COLORS":{"High":("#FCA5A5","#2E1111"),"Medium":("#FDE68A","#2A1E00"),"Low":("#6EE7B7","#0A2818")},
    },
    "Ocean Blue": {
        "BG":"#09111B","CARD":"#101B2A","CARD2":"#152338","NAV":"#07101A","NAV2":"#0D1726",
        "BORDER":"#1D3551","ACCENT":"#38BDF8","ACCENT2":"#7DD3FC",
        "SUCCESS":"#34D399","WARNING":"#FBBF24","DANGER":"#F87171",
        "MUTED":"#35516F","TEXT":"#E2EBF6","TEXT2":"#6F90AE","PURPLE":"#A78BFA",
        "NAV_TEXT":"#EAF2FB","NAV_MUTED":"#98B2CA","PRIMARY_FG":"#07131B",
        "S_COLORS":{
            "Open":         ("#7DD3FC","#091C30"),
            "In Progress":  ("#FDE68A","#2E2200"),
            "Pending info": ("#7090A8","#152034"),
            "Resolved":     ("#6EE7B7","#0A2818"),
        },
        "P_COLORS":{"High":("#FCA5A5","#2E1111"),"Medium":("#FDE68A","#2A1E00"),"Low":("#6EE7B7","#0A2818")},
    },
    "Stone & Sage": {
        "BG":"#121816","CARD":"#18211E","CARD2":"#1F2A26","NAV":"#0E1312","NAV2":"#131A18",
        "BORDER":"#2A3833","ACCENT":"#8FB8A8","ACCENT2":"#B7D5C8",
        "SUCCESS":"#6EE7B7","WARNING":"#F4C56E","DANGER":"#F38B8B",
        "MUTED":"#5C726A","TEXT":"#E7EFEB","TEXT2":"#91A69E","PURPLE":"#BFA7D9",
        "NAV_TEXT":"#EDF4F1","NAV_MUTED":"#9EB4AC","PRIMARY_FG":"#10201B",
        "S_COLORS":{
            "Open":         ("#B7D5C8","#1E312B"),
            "In Progress":  ("#F4C56E","#312407"),
            "Pending info": ("#AABCB5","#1F2A26"),
            "Resolved":     ("#86E0B6","#0D251A"),
        },
        "P_COLORS":{"High":("#F4A4A4","#321212"),"Medium":("#F4C56E","#312407"),"Low":("#86E0B6","#0D251A")},
    },
    "Driftwood": {
        "BG":"#171411","CARD":"#201C18","CARD2":"#29231E","NAV":"#120F0D","NAV2":"#181411",
        "BORDER":"#3A3028","ACCENT":"#C9A87C","ACCENT2":"#DFC4A3",
        "SUCCESS":"#79D7A7","WARNING":"#F1C66B","DANGER":"#F09191",
        "MUTED":"#6E6154","TEXT":"#F1E8DE","TEXT2":"#A79684","PURPLE":"#C9A7B5",
        "NAV_TEXT":"#F5ECE2","NAV_MUTED":"#B8A593","PRIMARY_FG":"#1B1410",
        "S_COLORS":{
            "Open":         ("#DFC4A3","#2F2216"),
            "In Progress":  ("#F1C66B","#302103"),
            "Pending info": ("#B8AA9A","#29231E"),
            "Resolved":     ("#79D7A7","#0C2418"),
        },
        "P_COLORS":{"High":("#F09191","#331212"),"Medium":("#F1C66B","#302103"),"Low":("#79D7A7","#0C2418")},
    },
    "Harbour Green": {
        "BG":"#101716","CARD":"#17201E","CARD2":"#1D2926","NAV":"#0B1110","NAV2":"#111917",
        "BORDER":"#28403A","ACCENT":"#219F78","ACCENT2":"#4FC49A",
        "SUCCESS":"#6EE7B7","WARNING":"#F4C56E","DANGER":"#F38B8B",
        "MUTED":"#557A71","TEXT":"#E4F0EB","TEXT2":"#92ACA3","PURPLE":"#92B7D6",
        "NAV_TEXT":"#EAF6F1","NAV_MUTED":"#A2BBB3","PRIMARY_FG":"#07110D",
        "S_COLORS":{
            "Open":         ("#4FC49A","#113329"),
            "In Progress":  ("#F4C56E","#312407"),
            "Pending info": ("#A3BAB3","#1D2926"),
            "Resolved":     ("#86E0B6","#0D251A"),
        },
        "P_COLORS":{"High":("#F4A4A4","#321212"),"Medium":("#F4C56E","#312407"),"Low":("#86E0B6","#0D251A")},
    },
    "Mist Light": {
        "BG":"#EEF4F6","CARD":"#FFFFFF","CARD2":"#F7FBFC","NAV":"#DDE8EB","NAV2":"#E8F0F2",
        "BORDER":"#C9D8DD","ACCENT":"#23778C","ACCENT2":"#4CA4B5",
        "SUCCESS":"#18794E","WARNING":"#B87116","DANGER":"#C44F4F",
        "MUTED":"#738A93","TEXT":"#18323A","TEXT2":"#516A73","PURPLE":"#6E84B7",
        "NAV_TEXT":"#18323A","NAV_MUTED":"#47616A","PRIMARY_FG":"#F8FCFD",
        "S_COLORS":{
            "Open":         ("#155E75","#D8F0F5"),
            "In Progress":  ("#9A5B10","#FDE7C2"),
            "Pending info": ("#516A73","#E3EDF0"),
            "Resolved":     ("#16603F","#DDF5E8"),
        },
        "P_COLORS":{"High":("#A93D3D","#F9E0E0"),"Medium":("#9A5B10","#FDE7C2"),"Low":("#16603F","#DDF5E8")},
    },
}

THEME_NOTES = {
    "Slate & Teal": ("Cool, crisp default", "Layered panels with teal focus accents"),
    "Ocean Blue": ("Deep navy with brighter contrast", "Best if you want a more polished dashboard feel"),
    "Stone & Sage": ("Low-glare green-grey", "Quiet workspace styling for long sessions"),
    "Driftwood": ("Warm neutral without harsh amber", "A softer, service-desk look with warmer cards"),
    "Harbour Green": ("Company-inspired using #219F78", "Shows your brand colour without making the whole UI green"),
    "Mist Light": ("Soft light theme with readable header", "Airier layout with much stronger toolbar contrast"),
}
ACTIVE_THEME = "Slate & Teal"

def apply_theme(name):
    global BG,CARD,CARD2,NAV,NAV2,BORDER,ACCENT,ACCENT2,SUCCESS,WARNING,DANGER
    global MUTED,TEXT,TEXT2,PURPLE,TEAL,S_COLORS,P_COLORS,ACTIVE_THEME
    global NAV_TEXT,NAV_MUTED,PRIMARY_FG
    t=THEMES.get(name,THEMES["Slate & Teal"])
    BG=t["BG"]; CARD=t["CARD"]; CARD2=t["CARD2"]; NAV=t["NAV"]; NAV2=t["NAV2"]
    BORDER=t["BORDER"]; ACCENT=t["ACCENT"]; ACCENT2=t["ACCENT2"]
    SUCCESS=t["SUCCESS"]; WARNING=t["WARNING"]; DANGER=t["DANGER"]
    MUTED=t["MUTED"]; TEXT=t["TEXT"]; TEXT2=t["TEXT2"]; PURPLE=t["PURPLE"]
    NAV_TEXT=t.get("NAV_TEXT",TEXT); NAV_MUTED=t.get("NAV_MUTED",TEXT2)
    PRIMARY_FG=t.get("PRIMARY_FG","#0B1017")
    TEAL=t["ACCENT"]
    S_COLORS.clear(); S_COLORS.update(t["S_COLORS"])
    P_COLORS.clear(); P_COLORS.update(t["P_COLORS"])
    ACTIVE_THEME=name

BG      = "#0F1419"
CARD    = "#161D26"
CARD2   = "#1C2530"
NAV     = "#0B1017"
NAV2    = "#111820"
BORDER  = "#243040"
ACCENT  = "#2DD4BF"
ACCENT2 = "#5EEAD4"
SUCCESS = "#34D399"
WARNING = "#FBBF24"
DANGER  = "#F87171"
MUTED   = "#445566"
TEXT    = "#E8EEF4"
TEXT2   = "#7A92A8"
PURPLE  = "#A78BFA"
TEAL    = "#2DD4BF"
NAV_TEXT = TEXT
NAV_MUTED = TEXT2
PRIMARY_FG = "#0B1017"
FONT    = "Segoe UI"

S_COLORS = {
    "Open":         ("#5EEAD4", "#0D2E2A"),
    "In Progress":  ("#FDE68A", "#2E2200"),
    "Pending info": ("#94A3B8", "#1C2530"),
    "Resolved":     ("#6EE7B7", "#0A2818"),
}
P_COLORS = {
    "High":   ("#FCA5A5", "#2E1111"),
    "Medium": ("#FDE68A", "#2A1E00"),
    "Low":    ("#6EE7B7", "#0A2818"),
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f: return json.load(f)
        except: pass
    if os.path.exists(LEGACY_CONFIG_FILE):
        try:
            with open(LEGACY_CONFIG_FILE) as f:
                cfg=json.load(f)
            save_config(cfg)
            return cfg
        except:
            pass
    return {}

def load_shared_settings(excel_file):
    """Read query types and team members from the _Settings sheet in the shared tracker.
    Returns a dict with 'query_types' and 'team_members', or empty dict if not present."""
    if not excel_file or not os.path.exists(excel_file):
        return {}
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if "_Settings" not in wb.sheetnames:
            return {}
        ws = wb["_Settings"]
        result = {}
        section = None
        for row in ws.iter_rows(values_only=True):
            key = str(row[0] or "").strip()
            if not key:
                continue
            if key == "[query_types]":
                section = "query_types"; result[section] = []; continue
            if key == "[team_members]":
                section = "team_members"; result[section] = []; continue
            if key.startswith("["):
                section = None; continue
            if section and key:
                result[section].append(key)
        return result
    except Exception:
        return {}

def save_shared_settings(excel_file, query_types, team_members):
    """Write query types and team members into the _Settings sheet of the shared tracker."""
    if not excel_file or not os.path.exists(excel_file):
        return
    try:
        wb = openpyxl.load_workbook(excel_file)
        if "_Settings" in wb.sheetnames:
            del wb["_Settings"]
        ws = wb.create_sheet("_Settings")
        ws["A1"] = "[query_types]"
        for i, qt in enumerate(query_types, start=2):
            ws.cell(row=i, column=1, value=qt)
        row = len(query_types) + 3
        ws.cell(row=row, column=1, value="[team_members]")
        for i, tm in enumerate(team_members, start=row+1):
            ws.cell(row=i, column=1, value=tm)
        wb.save(excel_file)
    except Exception:
        pass

def get_query_types(excel_file=None):
    """Load query types from the shared tracker first, then local config, then defaults."""
    if excel_file:
        shared = load_shared_settings(excel_file)
        if shared.get("query_types"):
            return shared["query_types"]
    cfg = load_config()
    return cfg.get("query_types", list(DEFAULT_QUERY_TYPES))

def get_team_members(excel_file=None):
    """Load team members from the shared tracker first, then local config."""
    if excel_file:
        shared = load_shared_settings(excel_file)
        if shared.get("team_members"):
            return shared["team_members"]
    cfg = load_config()
    return cfg.get("team_members", [])

# Initialise from config — will be overridden in _launch once excel_file is known
QUERY_TYPES[:] = get_query_types()

def save_config(cfg):
    _ensure_config_folder()
    with open(CONFIG_FILE,"w") as f: json.dump(cfg,f,indent=2)

def today_str(): return date.today().isoformat()

def _to_iso_date_str(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial dates (1900 date system) occasionally appear after manual edits.
        try:
            if value > 0:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except:
            return ""
        return ""

    s = str(value).strip()
    if not s:
        return ""

    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
        "%d %b %Y", "%d %B %Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except:
            pass
    return ""

def fmt_date(d):
    iso = _to_iso_date_str(d)
    if not iso:
        return str(d or "")
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except:
        return iso

def stamp(username): return f"{today_str()} [{username}]"

def parse_iso_date(d_str):
    try:
        iso = _to_iso_date_str(d_str)
        return datetime.strptime(iso, "%Y-%m-%d").date() if iso else None
    except:
        return None

def is_pushback(old_date, new_date, last_updated_date=""):
    # If this query was already updated today, don't count another date move as a pushback.
    if _to_iso_date_str(last_updated_date) == today_str():
        return False
    o = parse_iso_date(old_date)
    n = parse_iso_date(new_date)
    return bool(o and n and n > o)

def pushback_event_entry(username, old_date, new_date):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"{SYS_PUSHBACK_TAG}{ts};{old_date};{new_date};{username}"

def parse_pushback_events(log_str):
    events = []
    if not log_str: return events
    for entry in [e.strip() for e in str(log_str).split(" | ") if e.strip()]:
        if not entry.startswith(SYS_PUSHBACK_TAG):
            continue
        payload = entry[len(SYS_PUSHBACK_TAG):]
        parts = payload.split(";", 3)
        if len(parts) == 4:
            events.append({"timestamp": parts[0], "from": parts[1], "to": parts[2], "by": parts[3]})
    return events

def pushback_count(query):
    return len(parse_pushback_events(query.get("log", "")))

def append_pushback_event(query, username, old_date, new_date):
    entry = pushback_event_entry(username, old_date, new_date)
    query["log"] = (query.get("log", "").strip() + " | " + entry).strip(" |")

def public_log_entries(log_str):
    if not log_str: return []
    return [e for e in [x.strip() for x in str(log_str).split(" | ") if x.strip()]
            if not e.startswith(SYS_PUSHBACK_TAG)]

def attach_text_spellcheck(widget):
    if SPELLCHECKER is None or not isinstance(widget, tk.Text):
        return

    widget.tag_configure("spell_error", underline=True, foreground="#FCA5A5")
    pending = {"id": None}

    def refresh():
        try:
            text = widget.get("1.0", "end-1c")
        except tk.TclError:
            return
        widget.tag_remove("spell_error", "1.0", "end")
        if not text.strip():
            return

        words = []
        matches = []
        for match in SPELLCHECK_WORD_RE.finditer(text):
            word = match.group(0)
            normalized = word.lower()
            if normalized in SPELLCHECK_SKIP_WORDS:
                continue
            if word.isupper() and len(word) <= 5:
                continue
            words.append(normalized)
            matches.append(match)

        if not words:
            return

        misspelled = SPELLCHECKER.unknown(words)
        for match in matches:
            normalized = match.group(0).lower()
            if normalized not in misspelled:
                continue
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            widget.tag_add("spell_error", start, end)

    def schedule_refresh(_=None):
        if pending["id"] is not None:
            try:
                widget.after_cancel(pending["id"])
            except tk.TclError:
                pass
        try:
            pending["id"] = widget.after(250, refresh)
        except tk.TclError:
            pending["id"] = None

    widget.bind("<KeyRelease>", schedule_refresh, add="+")
    widget.bind("<FocusOut>", schedule_refresh, add="+")

    def _replace_word(start_idx, end_idx, replacement):
        try:
            widget.delete(start_idx, end_idx)
            widget.insert(start_idx, replacement)
        except tk.TclError:
            return
        schedule_refresh()

    def _right_click_suggestions(event):
        try:
            idx = widget.index(f"@{event.x},{event.y}")
        except tk.TclError:
            return
        if "spell_error" not in widget.tag_names(idx):
            return

        start = widget.index(f"{idx} wordstart")
        end = widget.index(f"{idx} wordend")
        word = widget.get(start, end).strip()
        normalized = word.lower()
        if not normalized or normalized in SPELLCHECK_SKIP_WORDS:
            return

        menu = tk.Menu(widget, tearoff=0)
        candidates = list(SPELLCHECKER.candidates(normalized) or [])
        preferred = SPELLCHECKER.correction(normalized)
        ordered = []
        if preferred:
            ordered.append(preferred)
        for candidate in sorted(candidates):
            if candidate not in ordered:
                ordered.append(candidate)

        def _match_case(candidate):
            if word.isupper():
                return candidate.upper()
            if word.istitle():
                return candidate.title()
            return candidate

        for candidate in ordered[:6]:
            replacement = _match_case(candidate)
            menu.add_command(
                label=replacement,
                command=lambda s=start, e=end, r=replacement: _replace_word(s, e, r),
            )
        if not ordered:
            menu.add_command(label="No suggestions", state="disabled")

        menu.add_separator()
        menu.add_command(
            label=f"Ignore '{word}'",
            command=lambda w=normalized: (SPELLCHECK_SKIP_WORDS.add(w), schedule_refresh()),
        )

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    widget.bind("<Button-3>", _right_click_suggestions, add="+")
    widget.bind("<Button-2>", _right_click_suggestions, add="+")
    schedule_refresh()

def is_weekend(d_str):
    try: return datetime.strptime(d_str,"%Y-%m-%d").weekday()>=5
    except: return False

def _easter_sunday(year):
    # Anonymous Gregorian algorithm.
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def _first_weekday_in_month(year, month, weekday):
    d = date(year, month, 1)
    while d.weekday() != weekday:
        d += timedelta(days=1)
    return d

def _last_weekday_in_month(year, month, weekday):
    if month == 12:
        d = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        d = date(year, month + 1, 1) - timedelta(days=1)
    while d.weekday() != weekday:
        d -= timedelta(days=1)
    return d

def uk_bank_holidays(year):
    # England & Wales bank holidays.
    holidays = set()

    new_year = date(year, 1, 1)
    if new_year.weekday() == 5:
        holidays.add(date(year, 1, 3))
    elif new_year.weekday() == 6:
        holidays.add(date(year, 1, 2))
    else:
        holidays.add(new_year)

    easter = _easter_sunday(year)
    holidays.add(easter - timedelta(days=2))  # Good Friday
    holidays.add(easter + timedelta(days=1))  # Easter Monday

    holidays.add(_first_weekday_in_month(year, 5, 0))  # Early May
    holidays.add(_last_weekday_in_month(year, 5, 0))   # Spring
    holidays.add(_last_weekday_in_month(year, 8, 0))   # Summer

    christmas = date(year, 12, 25)
    boxing = date(year, 12, 26)
    if christmas.weekday() in (0, 1, 2, 3, 4):
        holidays.add(christmas)
        if boxing.weekday() in (0, 1, 2, 3, 4):
            holidays.add(boxing)
        else:
            holidays.add(date(year, 12, 28))
    elif christmas.weekday() == 5:  # Saturday
        holidays.add(date(year, 12, 27))
        holidays.add(date(year, 12, 28))
    else:  # Sunday
        holidays.add(date(year, 12, 27))
        holidays.add(date(year, 12, 28))

    return holidays

def is_bank_holiday(d_str):
    d = parse_iso_date(d_str)
    if not d:
        return False
    return d in uk_bank_holidays(d.year)

def is_working_day(d_value):
    d = parse_iso_date(d_value) if isinstance(d_value, str) else d_value
    if not d:
        return False
    return d.weekday() < 5 and d not in uk_bank_holidays(d.year)

def sla_intake_working_days(start_str, end_str):
    start = parse_iso_date(start_str)
    end = parse_iso_date(end_str)
    if not start or not end:
        return None
    if end <= start:
        return 0

    working_days = 0
    current = start
    while current < end:
        current += timedelta(days=1)
        if is_working_day(current):
            working_days += 1
    return working_days

def days_overdue(chase_date_str):
    try:
        chase = parse_iso_date(chase_date_str)
        if not chase:
            return 0
        delta=(date.today()-chase).days
        return delta if delta>0 else 0
    except: return 0

def next_ref(queries,qtype):
    prefixes={"Sub-Meter reads":"SM","Tenancy change":"TC","Recharge rework":"RR",
              "New instruction":"NI","Quote request":"QR","Methodology query":"MQ",
              "Consumption":"CO","Missing AMR data":"AM","Missing supplier invoice":"SI","Other":"QY"}
    prefix=prefixes.get(qtype,"QY"); nums=[]
    for q in queries:
        if q["ref"].startswith(prefix+"-"):
            try: nums.append(int(q["ref"].split("-")[1]))
            except: pass
    return f"{prefix}-{str(max(nums)+1 if nums else 1).zfill(3)}"

def load_site_data(sites_file):
    if not sites_file or not os.path.exists(sites_file): return [],[],{},{},{},{}
    try:
        wb=openpyxl.load_workbook(sites_file,data_only=True)
    except PermissionError:
        messagebox.showerror("File in use","Sites.xlsx is open in Excel.\nPlease close it and try again.")
        return [],[],{},{},{},{}
    except Exception as e:
        messagebox.showerror("Error loading sites",str(e))
        return [],[],{},{},{},{}

    # Find the Sites sheet — try common names
    ws=None
    for name in ["Sites","Site","sites","site","Sheet1","Sheet"]:
        if name in wb.sheetnames:
            ws=wb[name]; break
    if ws is None:
        ws=wb.active

    clients=[]; sites_by_client={}; meters={}; utilities_by_site={}
    funds_by_client={}; sites_by_fund={}

    # Detect the first real data row.
    # A row is a HEADER/INSTRUCTION row if column A contains any of:
    #   • known header words (client, fund, site, address...)
    #   • instruction phrases (must be, note, example, column...)
    #   • dashes, asterisks, or is blank
    # A row is a DATA row when column A looks like an actual client name
    # AND column D (site name) is non-empty and doesn't look like a header.
    header_fragments = {
        "client","fund","site","address","utility","meter","contact",
        "property","prop","must be","note:","example","column","—","*",
        "managing","agent","spid","serial","instruction","please","name"
    }

    def looks_like_header(val):
        if not val: return True
        s = str(val).strip().lower()
        if not s or s in ("", "-", "—", "*"): return True
        return any(frag in s for frag in header_fragments)

    data_start_row = 1
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        col_a = row[0] if row else None
        col_d = row[3] if len(row) > 3 else None
        # Skip if col A looks like a header/instruction
        if looks_like_header(col_a):
            data_start_row += 1
            continue
        # Skip if col D (site) is missing or looks like a header
        if looks_like_header(col_d):
            data_start_row += 1
            continue
        break  # found first real data row

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row[0]: continue
        col_a = str(row[0]).strip()
        if looks_like_header(col_a): continue  # skip any stray header rows mid-sheet
        n=len(row)
        def gv(i): return str(row[i] or "").strip() if i<n else ""
        client=col_a
        fund=gv(1); prop_code=gv(2); site_name=gv(3)
        address=gv(4); utility=gv(5); spid=gv(6); serial=gv(7); contact=gv(8)
        if not site_name: continue
        if looks_like_header(site_name): continue
        if client not in clients: clients.append(client)
        sites_by_client.setdefault(client,[])
        if site_name not in sites_by_client[client]: sites_by_client[client].append(site_name)
        funds_by_client.setdefault(client,[])
        if fund and fund not in funds_by_client[client]: funds_by_client[client].append(fund)
        fkey=(client,fund); sites_by_fund.setdefault(fkey,[])
        if site_name not in sites_by_fund[fkey]: sites_by_fund[fkey].append(site_name)
        key=(client,site_name)
        meters.setdefault(key,[]).append(dict(fund=fund,utility=utility,spid=spid,
            serial=serial,address=address,contact=contact,prop_code=prop_code))
        utilities_by_site.setdefault(key,[])
        if utility and utility not in utilities_by_site[key]: utilities_by_site[key].append(utility)
    # Sort everything A-Z so dropdowns are always alphabetical
    clients.sort(key=str.lower)
    for k in sites_by_client: sites_by_client[k].sort(key=str.lower)
    for k in funds_by_client: funds_by_client[k].sort(key=str.lower)
    for k in sites_by_fund:   sites_by_fund[k].sort(key=str.lower)
    return clients,sites_by_client,meters,utilities_by_site,funds_by_client,sites_by_fund

def load_queries(excel_file):
    if not excel_file or not os.path.exists(excel_file): return []
    try: wb=openpyxl.load_workbook(excel_file,data_only=True)
    except: return []
    if "Queries" not in wb.sheetnames: return []
    ws=wb["Queries"]; queries=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row[0]: continue
        n=len(row)
        def g(i,d=""): return str(row[i] or d) if i<n else d
        if n>=20:
            q={"id":g(0),"ref":g(1),"client":g(2),"fund":g(3),"site":g(4),"utility":g(5),
               "meter":g(6),"type":g(7),"status":g(8,"Open"),"priority":g(9,"Medium"),
             "desc":g(10),"opened":_to_iso_date_str(g(11)) or today_str(),
             "chase_date":_to_iso_date_str(g(12)),"resolved_date":_to_iso_date_str(g(13)),
               "log":g(14),"address":g(15),"spid":g(16),"serial":g(17),"contact":g(18),
             "prop_code":g(19),"last_by":g(20),"last_date":_to_iso_date_str(g(21)),"assigned_to":g(22),
             "raised_date":_to_iso_date_str(g(23))}
        else:
            q={"id":g(0),"ref":g(1),"client":g(2),"fund":"","site":g(3),"utility":g(4),
               "meter":g(5),"type":g(6),"status":g(7,"Open"),"priority":g(8,"Medium"),
             "desc":g(9),"opened":_to_iso_date_str(g(10)) or today_str(),
             "chase_date":_to_iso_date_str(g(11)),"resolved_date":_to_iso_date_str(g(12)),
               "log":g(13),"address":g(14),"spid":g(15),"serial":g(16),"contact":g(17),
               "prop_code":g(18),"last_by":"","last_date":"","assigned_to":"","raised_date":""}
        queries.append(q)
    return queries

def _merge_queries_for_save(local_queries, excel_file):
    """Merge local in-memory queries with current on-disk queries by ID.

    This reduces accidental overwrite when two users save around the same time
    and are editing different queries.
    """
    remote_queries = load_queries(excel_file) if excel_file else []
    remote_by_id = {str(q.get("id", "")): q for q in remote_queries if q.get("id")}
    local_by_id = {str(q.get("id", "")): q for q in local_queries if q.get("id")}

    merged_ids = []
    for q in remote_queries:
        qid = str(q.get("id", ""))
        if qid and qid not in merged_ids:
            merged_ids.append(qid)
    for q in local_queries:
        qid = str(q.get("id", ""))
        if qid and qid not in merged_ids:
            merged_ids.append(qid)

    merged = []
    for qid in merged_ids:
        if qid in local_by_id:
            merged.append(local_by_id[qid])
        elif qid in remote_by_id:
            merged.append(remote_by_id[qid])
    return merged

def save_all_queries(queries,excel_file):
    if not excel_file: return
    queries = _merge_queries_for_save(queries, excel_file)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Queries"
    thin=Side(style="thin",color="CCCCCC"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    hfil=PatternFill("solid",fgColor="0F1B2D"); hfnt=Font(bold=True,color="FFFFFF",name=FONT,size=10)
    widths=[10,12,22,22,24,18,22,18,14,10,44,12,12,14,50,30,20,20,28,14,18,14,18,12]
    for i,(col,w) in enumerate(zip(COLS,widths),1):
        c=ws.cell(row=1,column=i,value=col)
        c.font=hfnt; c.fill=hfil; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[1].height=20
    sfills={"Open":"DBEAFE","In Progress":"FEF3C7","Pending info":"F3F4F6","Resolved":"D1FAE5"}
    pfills={"High":"FEE2E2","Medium":"FEF9C3","Low":"F0FDF4"}
    for r,q in enumerate(queries,2):
        vals=[q["id"],q["ref"],q["client"],q.get("fund",""),q["site"],q["utility"],q["meter"],
              q["type"],q["status"],q["priority"],q["desc"],q["opened"],
              q.get("chase_date",""),q.get("resolved_date",""),q["log"],
              q["address"],q["spid"],q["serial"],q["contact"],q["prop_code"],
              q.get("last_by",""),q.get("last_date",""),q.get("assigned_to",""),
              q.get("raised_date","")]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=c,value=v)
            cell.font=Font(name=FONT,size=10)
            cell.alignment=Alignment(vertical="top",wrap_text=(c in (11,15))); cell.border=bdr
        ws.cell(row=r,column=9).fill=PatternFill("solid",fgColor=sfills.get(q["status"],"FFFFFF"))
        ws.cell(row=r,column=10).fill=PatternFill("solid",fgColor=pfills.get(q["priority"],"FFFFFF"))
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
    ds=wb.create_sheet("Dashboard"); ds.sheet_view.showGridLines=False
    ds["A1"]="Query Tracker — Dashboard"
    ds["A1"].font=Font(bold=True,size=14,name=FONT,color="0F1B2D")
    ds["A2"]=f"Last updated: {date.today().strftime('%d/%m/%Y')}"
    ds["A2"].font=Font(size=10,name=FONT,color="6B7280")
    today=today_str()
    stats=[("Total",len(queries)),("Open",sum(1 for q in queries if q["status"]!="Resolved")),
           ("Resolved",sum(1 for q in queries if q["status"]=="Resolved")),
           ("Action today",sum(1 for q in queries if q["status"]!="Resolved" and q.get("chase_date","")<=today and q.get("chase_date","")))]
    for i,(lbl,val) in enumerate(stats):
        col=i*2+1
        ds.cell(row=4,column=col,value=lbl).font=Font(name=FONT,size=9,color="6B7280")
        ds.cell(row=5,column=col,value=val).font=Font(name=FONT,size=18,bold=True,
            color="DC2626" if lbl=="Action today" and val>0 else "0F1B2D")
        ds.column_dimensions[get_column_letter(col)].width=18
    wb.save(excel_file)

def apply_styles():
    style=ttk.Style(); style.theme_use("clam")
    style.configure(".",font=(FONT,10),background=BG)
    # Combobox
    style.configure("Modern.TCombobox",
        fieldbackground=CARD2,background=CARD2,foreground=TEXT,
        bordercolor=BORDER,lightcolor=BORDER,darkcolor=BORDER,
        arrowcolor=TEXT2,arrowsize=13,padding=(9,7),relief="flat",borderwidth=1,
        selectbackground=CARD2,selectforeground=TEXT)
    style.map("Modern.TCombobox",
        fieldbackground=[("readonly",CARD2),("focus",CARD2)],
        bordercolor=[("focus",ACCENT),("!focus",BORDER)],
        arrowcolor=[("hover",ACCENT2)],
        foreground=[("focus",TEXT),("!focus",TEXT)])
    # Treeview
    style.configure("Modern.Treeview",
        background=CARD,fieldbackground=CARD,foreground=TEXT,
        rowheight=36,font=(FONT,10),borderwidth=0,relief="flat")
    style.configure("Modern.Treeview.Heading",
        background=NAV2,foreground=MUTED,
        font=(FONT,8,"bold"),borderwidth=0,relief="flat",padding=(12,10))
    style.map("Modern.Treeview",
        background=[("selected","#0D2E2A")],
        foreground=[("selected",ACCENT2)])
    style.map("Modern.Treeview.Heading",
        background=[("active",BORDER)])
    # Scrollbar — thin and subtle
    style.configure("Modern.Vertical.TScrollbar",
        background=BORDER,troughcolor=BG,borderwidth=0,arrowsize=0,width=4)
    style.map("Modern.Vertical.TScrollbar",background=[("active",MUTED)])

_combobox_wheel_bound=False

def _ignore_combobox_wheel(e):
    widget = e.widget
    if isinstance(widget, ttk.Combobox):
        return "break"
    wname = str(widget)
    # Combobox popup list path typically ends with .f.l or contains .popdown.
    if wname.endswith(".f.l") or ".popdown." in wname:
        return "break"
    return None


def _bind_combobox_wheel_once(widget):
    global _combobox_wheel_bound
    if _combobox_wheel_bound:
        return
    root = widget.winfo_toplevel()
    root.bind_all("<MouseWheel>", _ignore_combobox_wheel, add="+")
    root.bind_all("<Button-4>", _ignore_combobox_wheel, add="+")
    root.bind_all("<Button-5>", _ignore_combobox_wheel, add="+")
    _combobox_wheel_bound = True


def make_combo(parent,textvariable,values,readonly=False,width=30):
    cb=ttk.Combobox(parent,textvariable=textvariable,values=values,
                        style="Modern.TCombobox",width=width,
                        state="readonly" if readonly else "normal",font=(FONT,10))

    _bind_combobox_wheel_once(parent)

    # Also prevent manual wheel on the instanced combobox field.
    cb.bind("<MouseWheel>",lambda e: "break")
    cb.bind("<Button-4>",lambda e: "break")   # Linux scroll up
    cb.bind("<Button-5>",lambda e: "break")   # Linux scroll down

    # More robust: disable wheel at combobox class level too (handles variations)
    cb.bind_class("TCombobox", "<MouseWheel>", lambda e: "break")
    cb.bind_class("TCombobox", "<Button-4>", lambda e: "break")
    cb.bind_class("TCombobox", "<Button-5>", lambda e: "break")

    return cb

def make_btn(parent,text,command,style="default",padx=14,pady=6):
    configs={
        "default": {"bg":CARD2,      "fg":TEXT2,   "border":BORDER,  "hbg":BORDER,    "hfg":TEXT},
        "primary":  {"bg":ACCENT,     "fg":PRIMARY_FG,"border":ACCENT,  "hbg":ACCENT2,   "hfg":PRIMARY_FG},
        "success":  {"bg":"#14532D",  "fg":SUCCESS, "border":SUCCESS, "hbg":"#166534", "hfg":"#FFFFFF"},
        "danger":   {"bg":CARD2,      "fg":DANGER,  "border":DANGER,  "hbg":"#3B1111", "hfg":DANGER},
        "warning":  {"bg":"#3B2800",  "fg":WARNING, "border":WARNING, "hbg":"#4A3200", "hfg":WARNING},
        "nav":      {"bg":NAV,        "fg":NAV_MUTED,"border":BORDER,  "hbg":CARD2,     "hfg":NAV_TEXT},
        "active":   {"bg":"#0D2E2A",  "fg":ACCENT2, "border":ACCENT,  "hbg":"#113530", "hfg":ACCENT2},
        "pill":     {"bg":"#0D2E2A",  "fg":ACCENT2, "border":"#0D2E2A","hbg":ACCENT,   "hfg":PRIMARY_FG},
    }
    c=configs.get(style,configs["default"])
    btn=tk.Label(parent,text=text,font=(FONT,10),bg=c["bg"],fg=c["fg"],cursor="hand2",
                 padx=padx,pady=pady,relief="flat",bd=0,
                 highlightthickness=1,highlightbackground=c["border"])
    def on_enter(e): btn.configure(bg=c["hbg"],fg=c["hfg"])
    def on_leave(e): btn.configure(bg=c["bg"],fg=c["fg"])
    def on_click(e): command()
    btn.bind("<Enter>",on_enter); btn.bind("<Leave>",on_leave); btn.bind("<Button-1>",on_click)
    return btn

def divider(parent):
    tk.Frame(parent,bg=BORDER,height=1).pack(fill="x",pady=8)

def section_lbl(parent, text):
    """Labelled section divider with horizontal rule."""
    f=tk.Frame(parent,bg=BG); f.pack(fill="x",pady=(20,10))
    tk.Label(f,text=text,font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(side="left")
    tk.Frame(f,bg=BORDER,height=1).pack(side="left",fill="x",expand=True,padx=(10,0),pady=6)

def scrollable_frame(parent):
    canvas=tk.Canvas(parent,bg=BG,highlightthickness=0,bd=0)
    sb=ttk.Scrollbar(parent,orient="vertical",command=canvas.yview,style="Modern.Vertical.TScrollbar")
    sb.pack(side="right",fill="y"); canvas.pack(fill="both",expand=True)
    inner=tk.Frame(canvas,bg=BG,padx=24,pady=16)
    win=canvas.create_window((0,0),window=inner,anchor="nw")
    inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>",lambda e:canvas.itemconfig(win,width=e.width))

    def _scroll(e): canvas.yview_scroll(int(-1*(e.delta/120)),"units")
    canvas.bind("<MouseWheel>",_scroll)
    inner.bind("<MouseWheel>",_scroll)
    def _bind_children(widget):
        widget.bind("<MouseWheel>",_scroll)
        for child in widget.winfo_children(): _bind_children(child)
    inner.bind("<Configure>",lambda e:(_bind_children(inner),
                                        canvas.configure(scrollregion=canvas.bbox("all"))),add="+")
    return inner,canvas

def _safe_folder_name(s):
    for ch in r'\/:*?"<>|': s=s.replace(ch,"_")
    return s.strip()[:60] or "Unknown"

def get_attachments_root(sites_file):
    if not sites_file: return None
    return os.path.join(os.path.dirname(sites_file),"Attachments")

def get_attachment_folder(sites_file, q, create=True):
    # SharePoint URLs cannot be used for local file operations - return None gracefully
    if sites_file and (sites_file.startswith("http://") or sites_file.startswith("https://")):
        return None
    root=get_attachments_root(sites_file)
    if not root: return None
    site  =_safe_folder_name(q.get("site","Unknown") or "Unknown")
    utility = _safe_folder_name((q.get("utility","") or "").strip() or "No utility assigned")
    qtype = _safe_folder_name((q.get("type","") or "").strip() or "General")
    ref   =_safe_folder_name(q.get("ref","Unknown"))
    folder=os.path.join(root,site,utility,qtype,ref)
    if create:
        os.makedirs(folder,exist_ok=True)
    return folder

def get_attachment_folder_legacy(sites_file, q):
    if sites_file and (sites_file.startswith("http://") or sites_file.startswith("https://")):
        return None
    root=get_attachments_root(sites_file)
    if not root: return None
    client=_safe_folder_name(q.get("client","Unknown"))
    site  =_safe_folder_name(q.get("site","Unknown") or "Unknown")
    ref   =_safe_folder_name(q.get("ref","Unknown"))
    return os.path.join(root,client,site,ref)

def list_attachments(sites_file, q):
    files=[]

    folder=get_attachment_folder(sites_file,q,create=False)
    if folder and os.path.exists(folder):
        for f in sorted(os.listdir(folder)):
            fp=os.path.join(folder,f)
            if os.path.isfile(fp): files.append((f,fp))

    # Backward compatibility for previously saved attachments in old folder layout.
    if not files:
        legacy=get_attachment_folder_legacy(sites_file,q)
        if legacy and os.path.exists(legacy):
            for f in sorted(os.listdir(legacy)):
                fp=os.path.join(legacy,f)
                if os.path.isfile(fp): files.append((f,fp))

    return files

def save_attachment(sites_file, q, src_path):
    folder=get_attachment_folder(sites_file,q)
    if not folder: return None,None
    fname=os.path.basename(src_path)
    dest=os.path.join(folder,fname)
    base,ext=os.path.splitext(fname)
    counter=1
    while os.path.exists(dest):
        dest=os.path.join(folder,f"{base}_{counter}{ext}"); counter+=1
    shutil.copy2(src_path,dest)
    return os.path.basename(dest),dest

def open_file(path):
    try:
        if sys.platform=="win32": os.startfile(path)
        elif sys.platform=="darwin": subprocess.Popen(["open",path])
        else: subprocess.Popen(["xdg-open",path])
    except Exception as e:
        messagebox.showerror("Cannot open file",str(e))

def open_folder(path):
    try:
        if sys.platform=="win32": subprocess.Popen(["explorer",path])
        elif sys.platform=="darwin": subprocess.Popen(["open",path])
        else: subprocess.Popen(["xdg-open",path])
    except Exception as e:
        messagebox.showerror("Cannot open folder",str(e))

DROP_FOLDER_NAME = "_DROP_HERE"

def get_drop_inbox(sites_file):
    # Prefer attachment folder next to the sites list workbook.
    # If this is not available, fall back to an Attachments folder under the current working directory.
    root = None
    if sites_file and os.path.exists(sites_file):
        root = get_attachments_root(sites_file)
    if not root:
        root = os.path.join(os.getcwd(), "Attachments")

    inbox = os.path.join(root, DROP_FOLDER_NAME)
    try:
        os.makedirs(inbox, exist_ok=True)
    except Exception:
        # Fallback to current directory if creation here is blocked by permissions.
        inbox = os.path.join(os.getcwd(), DROP_FOLDER_NAME)
        os.makedirs(inbox, exist_ok=True)

    return inbox

def extract_ref_from_filename(filename, queries):
    name = os.path.splitext(filename)[0].upper()
    for q in queries:
        ref = q.get("ref","").upper()
        if ref and ref in name:
            return q
    m = re.search(r'([A-Z]{2}-\d+)', name)
    if m:
        code = m.group(1)
        for q in queries:
            if q.get("ref","").upper() == code:
                return q
    return None

def _show_toast(root_widget, message, color=None, duration=4000):
    color = color or "#1A2E48"
    toast = tk.Toplevel(root_widget)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.configure(bg=color)
    root_widget.update_idletasks()
    rx = root_widget.winfo_rootx() + root_widget.winfo_width()
    ry = root_widget.winfo_rooty() + root_widget.winfo_height()
    toast.geometry(f"340x60+{rx-360}+{ry-80}")
    tk.Label(toast, text=message, font=(FONT,9), bg=color, fg="white",
             wraplength=310, justify="left", padx=14, pady=10).pack(fill="both")
    toast.after(duration, toast.destroy)

def _show_cal(parent,date_var,get_day_load=None,confirm_day_selection=None,high_volume_threshold=HIGH_VOLUME_DAY_THRESHOLD,date_block_reason=None):
    try: current=datetime.strptime(date_var.get(),"%Y-%m-%d").date()
    except: current=date.today()
    cal=tk.Toplevel(parent); cal.title("Pick date"); cal.configure(bg=BG)
    cal.resizable(True,True); cal.grab_set()
    cw=min(420,max(340,parent.winfo_screenwidth()-80))
    ch=min(500,max(360,parent.winfo_screenheight()-140))
    sx=max(0,parent.winfo_screenwidth()-cw-20)
    sy=max(0,parent.winfo_screenheight()-ch-60)
    px=min(parent.winfo_rootx()+40,sx)
    py=min(parent.winfo_rooty()+40,sy)
    cal.geometry(f"{cw}x{ch}+{px}+{py}")

    wrap=tk.Frame(cal,bg=BG)
    wrap.pack(fill="both",expand=True)
    canvas=tk.Canvas(wrap,bg=BG,highlightthickness=0,bd=0)
    sb=ttk.Scrollbar(wrap,orient="vertical",command=canvas.yview,style="Modern.Vertical.TScrollbar")
    sb.pack(side="right",fill="y")
    canvas.pack(side="left",fill="both",expand=True)
    inner=tk.Frame(canvas,bg=BG)
    win=canvas.create_window((0,0),window=inner,anchor="nw")
    inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>",lambda e:canvas.itemconfig(win,width=e.width))
    canvas.configure(yscrollcommand=sb.set)

    def _scroll(e):
        delta=getattr(e,"delta",0)
        if delta:
            canvas.yview_scroll(int(-1*(delta/120)),"units")
        elif getattr(e,"num",None)==4:
            canvas.yview_scroll(-1,"units")
        elif getattr(e,"num",None)==5:
            canvas.yview_scroll(1,"units")
    for widget in (canvas,inner):
        widget.bind("<MouseWheel>",_scroll)
        widget.bind("<Button-4>",_scroll)
        widget.bind("<Button-5>",_scroll)

    state={"year":current.year,"month":current.month}

    def build(yr,mo):
        for w in inner.winfo_children(): w.destroy()
        hdr=tk.Frame(inner,bg=NAV,padx=12,pady=8); hdr.pack(fill="x")
        def prev():
            m=mo-1; y=yr
            if m<1: m=12; y-=1
            state["year"]=y; state["month"]=m; build(y,m)
        def nxt():
            m=mo+1; y=yr
            if m>12: m=1; y+=1
            state["year"]=y; state["month"]=m; build(y,m)
        make_btn(hdr,"◀",prev,"nav",padx=8,pady=4).pack(side="left")
        tk.Label(hdr,text=datetime(yr,mo,1).strftime("%B %Y"),font=(FONT,11,"bold"),
             bg=NAV,fg=NAV_TEXT).pack(side="left",expand=True)
        make_btn(hdr,"▶",nxt,"nav",padx=8,pady=4).pack(side="right")
        grid=tk.Frame(inner,bg=BG,padx=10,pady=8); grid.pack(fill="both",expand=True)
        for i,day in enumerate(["Mo","Tu","We","Th","Fr","Sa","Su"]):
            tk.Label(grid,text=day,font=(FONT,8,"bold"),bg=BG,
                     fg=DANGER if day in ("Sa","Su") else MUTED,width=3,anchor="center").grid(row=0,column=i,pady=(0,4))
        for i in range(7):
            grid.columnconfigure(i,weight=1)
        import calendar as _cal
        first_wd,days_in_mo=_cal.monthrange(yr,mo)
        row=1; col=first_wd
        threshold = max(1, int(high_volume_threshold or HIGH_VOLUME_DAY_THRESHOLD))
        for d in range(1,days_in_mo+1):
            wd=_cal.weekday(yr,mo,d)
            day_str=date(yr,mo,d).strftime("%Y-%m-%d")
            load=0
            if callable(get_day_load):
                try: load=int(get_day_load(day_str) or 0)
                except: load=0
            is_today=(date(yr,mo,d)==date.today())
            is_sel=(date(yr,mo,d)==current)
            is_weekend=(wd>=5)
            if is_sel and is_today:
                bg=ACCENT; fg="white"; border=ACCENT
            elif is_sel:
                bg=ACCENT; fg="white"; border=ACCENT
            elif is_today:
                bg=CARD2; fg=ACCENT2; border=ACCENT
            elif is_weekend:
                bg=CARD; fg=MUTED; border=BORDER
            elif load>=threshold:
                bg="#2A1015"; fg="#FCA5A5"; border="#5C2030"
            elif load>=max(1, threshold//2):
                bg="#231900"; fg="#FDE68A"; border="#5C4400"
            else:
                bg=CARD; fg=TEXT; border=BORDER
            btn=tk.Label(grid,text=str(d),font=(FONT,9,"bold" if is_today or is_sel else ""),
                         bg=bg,fg=fg,width=3,cursor="hand2",anchor="center",
                         highlightthickness=1,highlightbackground=border)
            btn.grid(row=row,column=col,padx=1,pady=1,ipady=3)
            def pick(dd=d,yy=yr,mm=mo):
                picked=date(yy,mm,dd).strftime("%Y-%m-%d")
                if callable(date_block_reason):
                    try:
                        block_msg = date_block_reason(picked)
                        if block_msg:
                            messagebox.showwarning("Date not allowed", block_msg, parent=cal)
                            return
                    except:
                        pass
                if callable(confirm_day_selection):
                    try:
                        if not confirm_day_selection(picked):
                            return
                    except:
                        pass
                date_var.set(picked); cal.destroy()
            btn.bind("<Button-1>",lambda e,p=pick:p())
            btn.bind("<Enter>",lambda e,b=btn,s=is_sel,w=is_weekend:b.configure(bg=ACCENT2 if s else (CARD2 if not w else CARD)))
            btn.bind("<Leave>",lambda e,b=btn,ob=bg:b.configure(bg=ob))
            col+=1
            if col>6: col=0; row+=1
        tk.Frame(inner,bg=BORDER,height=1).pack(fill="x")
        bf=tk.Frame(inner,bg=BG,padx=10,pady=8); bf.pack(fill="x")
        if callable(get_day_load):
            lg=tk.Frame(bf,bg=BG); lg.pack(side="left")
            tk.Label(lg,text="High",font=(FONT,7),bg=BG,fg="#FCA5A5").pack(side="left",padx=(0,6))
            tk.Label(lg,text="Medium",font=(FONT,7),bg=BG,fg="#FDE68A").pack(side="left",padx=(0,6))
            tk.Label(lg,text="Normal",font=(FONT,7),bg=BG,fg=MUTED).pack(side="left")
        def pick_today():
            today_pick = date.today().strftime("%Y-%m-%d")
            if callable(date_block_reason):
                try:
                    block_msg = date_block_reason(today_pick)
                    if block_msg:
                        messagebox.showwarning("Date not allowed", block_msg, parent=cal)
                        return
                except:
                    pass
            if callable(confirm_day_selection):
                try:
                    if not confirm_day_selection(today_pick):
                        return
                except:
                    pass
            date_var.set(today_pick); cal.destroy()
        make_btn(bf,"Today",pick_today,"default",padx=10,pady=4).pack(side="left")
        make_btn(bf,"Cancel",cal.destroy,"default",padx=10,pady=4).pack(side="right")
        inner.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.yview_moveto(0)

    build(state["year"],state["month"])

class SetupWizard(tk.Toplevel):
    def __init__(self,parent,existing_cfg=None,on_complete=None):
        super().__init__(parent)
        self.title("Query Tracker — Settings")
        sw=self.winfo_screenwidth(); sh=self.winfo_screenheight()
        ww=min(980,max(760,sw-120)); wh=min(860,max(700,sh-120))
        x=max(20,(sw-ww)//2); y=max(20,(sh-wh)//2)
        self.geometry(f"{ww}x{wh}+{x}+{y}")
        self.minsize(760,700)
        self.configure(bg=BG); self.resizable(True,True); self.grab_set()
        self.on_complete=on_complete; self.cfg=dict(existing_cfg or {})

        hdr=tk.Frame(self,bg=NAV,padx=24,pady=18); hdr.pack(fill="x")
        tk.Label(hdr,text="⚙  Settings",font=(FONT,14,"bold"),bg=NAV,fg=NAV_TEXT).pack(anchor="w")
        tk.Label(hdr,text="Configure file locations, your name, and query types.",
             font=(FONT,9),bg=NAV,fg=NAV_MUTED).pack(anchor="w",pady=(3,0))
        tk.Frame(self,bg=ACCENT,height=2).pack(fill="x")

        # ── Tab bar ───────────────────────────────────────────────────────────
        tab_bar=tk.Frame(self,bg=NAV2); tab_bar.pack(fill="x")
        tk.Frame(self,bg=BORDER,height=1).pack(fill="x")
        self._tab_frames={}; self._stab_btns={}
        self._active_stab=tk.StringVar(value="general")

        content=tk.Frame(self,bg=BG); content.pack(fill="both",expand=True)

        for tab_id,tab_lbl in [("general","General"),("query_types","Query Types"),
                                ("team","Team Members"),("trackers","Linked Trackers"),
                                ("escalation","Escalation"),("theme","Theme")]:
            f=tk.Frame(content,bg=BG); self._tab_frames[tab_id]=f
            btn=tk.Label(tab_bar,text=f"  {tab_lbl}  ",font=(FONT,10),bg=NAV2,
                         fg=NAV_MUTED,cursor="hand2",padx=6,pady=10)
            btn.pack(side="left")
            btn.bind("<Button-1>",lambda e,t=tab_id:self._switch_stab(t))
            self._stab_btns[tab_id]=btn

        self._switch_stab("general",init=True)

        # ── General tab ───────────────────────────────────────────────────────
        gen=self._tab_frames["general"]
        body=tk.Frame(gen,bg=BG,padx=28,pady=20); body.pack(fill="both",expand=True)

        def path_row(label,note,key,save_mode=False):
            tk.Label(body,text=label,font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(12,2))
            tk.Label(body,text=note,font=(FONT,9),bg=BG,fg=TEXT2).pack(anchor="w")
            row=tk.Frame(body,bg=BG); row.pack(fill="x",pady=(6,0))
            var=tk.StringVar(value=self.cfg.get(key,""))
            card=tk.Frame(row,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
            card.pack(side="left",fill="x",expand=True,padx=(0,10))
            tk.Entry(card,textvariable=var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                     relief="flat",bd=8,highlightthickness=0).pack(fill="x")
            def browse(v=var,sm=save_mode):
                current=v.get().strip()
                initial_dir=os.path.dirname(current) if current and os.path.dirname(current) else DEFAULT_DATA_DIR
                if sm:
                    # Always use open-file picker — never overwrite
                    p=filedialog.askopenfilename(
                        title="Select your query_tracker.xlsx",
                        initialdir=initial_dir,
                        filetypes=[("Excel files","*.xlsx"),("All files","*.*")],
                    )
                else:
                    p=filedialog.askopenfilename(
                        title="Find sites.xlsx",
                        initialdir=initial_dir,
                        filetypes=[("Excel files","*.xlsx")],
                    )
                if p: v.set(p)
            make_btn(row,"Browse",browse,"default",padx=14,pady=6).pack(side="left")
            return var

        # Only pre-fill the default path when there is no saved setting at all
        # (i.e. very first launch). Do not overwrite a saved path from a previous session.
        if not self.cfg.get("excel_file"):
            self.cfg["excel_file"] = DEFAULT_EXCEL_FILE
        self.excel_var=path_row("Query data file",
            "Click Browse to select your shared query_tracker.xlsx. Your choice is remembered for next time.","excel_file",save_mode=True)
        self.sites_var=path_row("Site list (sites.xlsx)",
            "Point to your sites.xlsx in the same shared folder.","sites_file")

        def test_sites():
            """Load the sites file right now and show what was found — for diagnosing issues."""
            sf=self.sites_var.get().strip()
            if not sf:
                messagebox.showwarning("No file set","Please choose a sites.xlsx file first.",parent=self); return
            if not os.path.exists(sf):
                messagebox.showerror("File not found",f"Cannot find:\n{sf}",parent=self); return
            try:
                import openpyxl as _opx
                wb=_opx.load_workbook(sf,data_only=True)
                sheet_names=wb.sheetnames
                ws=wb.active
                # Show first 5 data rows raw
                rows_preview=[]
                for i,row in enumerate(ws.iter_rows(min_row=1,max_row=6,values_only=True)):
                    rows_preview.append(f"Row {i+1}: {[str(c or '')[:20] for c in row[:6]]}")
                clients_found,sites_dict,*_=load_site_data(sf)
                total_sites=sum(len(v) for v in sites_dict.values())
                msg=(f"Sheets in file: {sheet_names}\n"
                     f"Active sheet: '{ws.title}'\n\n"
                     f"First rows (cols A–F):\n" + "\n".join(rows_preview) +
                     f"\n\nResult: {len(clients_found)} clients, {total_sites} sites loaded\n\n")
                if clients_found:
                    msg+=f"Clients found:\n"+"\n".join(f"  • {c} ({len(sites_dict.get(c,[]))} sites)" for c in clients_found[:10])
                else:
                    msg+=("⚠ No clients loaded.\n\nCommon causes:\n"
                          "• Data doesn't start in column A\n"
                          "• Sheet name is unexpected (check above)\n"
                          "• Column order is different from expected:\n"
                          "  A=Client, B=Fund, C=Prop code, D=Site, E=Address,\n"
                          "  F=Utility, G=SPID, H=Serial, I=Contact")
                messagebox.showinfo("Site file diagnostic",msg,parent=self)
            except Exception as e:
                messagebox.showerror("Error reading file",str(e),parent=self)

        make_btn(body,"🔍 Test sites file",test_sites,"default",padx=12,pady=5).pack(anchor="w",pady=(6,0))
        tk.Label(body,text="Your name",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(16,2))
        tk.Label(body,text="Shown in activity logs and used to identify which queries are assigned to you.",
                 font=(FONT,9),bg=BG,fg=TEXT2).pack(anchor="w")
        # Multi-user note
        note_f=tk.Frame(body,bg="#1A2E18",highlightthickness=1,highlightbackground="#2E5030",padx=10,pady=8)
        note_f.pack(fill="x",pady=(6,0))
        tk.Label(note_f,
                 text="ℹ  Multi-user setup: each person runs their own copy of this app on their PC.\n"
                      "   Settings (including your name) are saved locally on your PC only — not in the shared folder.\n"
                      "   The query data file (query_tracker.xlsx) lives in the shared OneDrive folder and is shared by everyone.\n"
                      "   Each person just needs to point their app at the same shared xlsx in Settings → General.",
                 font=(FONT,8),bg="#1A2E18",fg="#86EFAC",justify="left",wraplength=520).pack(anchor="w")
        self.name_var=tk.StringVar(value=self.cfg.get("username",""))
        nc=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); nc.pack(anchor="w",pady=(6,0))
        tk.Entry(nc,textvariable=self.name_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=8,width=28,highlightthickness=0).pack()

        tk.Label(body,text="High-volume warning threshold",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(16,2))
        tk.Label(body,text="Show workload warnings when a day has at least this many open queries.",
                 font=(FONT,9),bg=BG,fg=TEXT2).pack(anchor="w")
        self.high_volume_threshold_var=tk.StringVar(
            value=str(self.cfg.get("high_volume_threshold", HIGH_VOLUME_DAY_THRESHOLD))
        )
        hvt_row=tk.Frame(body,bg=BG); hvt_row.pack(anchor="w",pady=(6,0))
        hvt_card=tk.Frame(hvt_row,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        hvt_card.pack(side="left")
        tk.Entry(hvt_card,textvariable=self.high_volume_threshold_var,font=(FONT,10),bg=CARD2,fg=TEXT,
                 insertbackground=TEXT,relief="flat",bd=8,width=6,highlightthickness=0).pack()
        tk.Label(hvt_row,text="queries/day",font=(FONT,9),bg=BG,fg=MUTED).pack(side="left",padx=(8,0))

        # ── Query Types tab ───────────────────────────────────────────────────
        qt_frame=self._tab_frames["query_types"]
        qt_body=tk.Frame(qt_frame,bg=BG,padx=24,pady=16); qt_body.pack(fill="both",expand=True)
        tk.Label(qt_body,text="Manage query types",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w")
        tk.Label(qt_body,text="Add, rename or remove types. Changes are saved to the shared tracker so all teammates see the same list automatically.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=540,justify="left").pack(anchor="w",pady=(4,4))
        sync_note=tk.Frame(qt_body,bg="#1A2E18",highlightthickness=1,highlightbackground="#2E5030",padx=10,pady=6)
        sync_note.pack(fill="x",pady=(0,10))
        tk.Label(sync_note,text="🔄  Shared — any changes you save here will be picked up by all teammates within 30 seconds.",
                 font=(FONT,8),bg="#1A2E18",fg="#86EFAC",justify="left").pack(anchor="w")

        list_frame=tk.Frame(qt_body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        list_frame.pack(fill="both",expand=True,pady=(0,10))
        lb=tk.Listbox(list_frame,font=(FONT,10),bg=CARD2,fg=TEXT,relief="flat",
                      selectbackground="#0D2E2A",selectforeground=ACCENT2,
                      activestyle="none",bd=0,highlightthickness=0)
        lb_sb=ttk.Scrollbar(list_frame,orient="vertical",command=lb.yview,style="Modern.Vertical.TScrollbar")
        lb.configure(yscrollcommand=lb_sb.set)
        lb_sb.pack(side="right",fill="y"); lb.pack(fill="both",expand=True,padx=4,pady=4)

        # Seed from the shared tracker if available, otherwise local config
        _shared_qt=load_shared_settings(self.cfg.get("excel_file",""))
        current_types=list(_shared_qt.get("query_types") or self.cfg.get("query_types",DEFAULT_QUERY_TYPES))
        def refresh_lb():
            lb.delete(0,"end")
            for t in current_types: lb.insert("end",f"  {t}")
        refresh_lb()

        ctrl=tk.Frame(qt_body,bg=BG); ctrl.pack(fill="x",pady=(0,4))
        new_var=tk.StringVar()
        nf=tk.Frame(ctrl,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        nf.pack(side="left",fill="x",expand=True,padx=(0,8))
        tk.Entry(nf,textvariable=new_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=7,highlightthickness=0).pack(fill="x")

        def add_type():
            t=new_var.get().strip()
            if not t: return
            if t in current_types:
                messagebox.showwarning("Duplicate","That type already exists.",parent=self); return
            current_types.append(t); new_var.set(""); refresh_lb()

        def remove_type():
            sel=lb.curselection()
            if not sel: return
            t=current_types[sel[0]]
            if messagebox.askyesno("Remove type",
                f"Remove '{t}'?\n\nExisting queries of this type are unaffected.",parent=self):
                current_types.pop(sel[0]); refresh_lb()

        def move_up():
            sel=lb.curselection()
            if not sel or sel[0]==0: return
            i=sel[0]; current_types[i-1],current_types[i]=current_types[i],current_types[i-1]
            refresh_lb(); lb.selection_set(i-1)

        def move_down():
            sel=lb.curselection()
            if not sel or sel[0]>=len(current_types)-1: return
            i=sel[0]; current_types[i],current_types[i+1]=current_types[i+1],current_types[i]
            refresh_lb(); lb.selection_set(i+1)

        make_btn(ctrl,"Add",add_type,"primary",padx=12,pady=6).pack(side="left",padx=(0,4))
        make_btn(ctrl,"Remove selected",remove_type,"danger",padx=12,pady=6).pack(side="left",padx=(0,4))
        make_btn(ctrl,"▲",move_up,"default",padx=10,pady=6).pack(side="left",padx=(0,2))
        make_btn(ctrl,"▼",move_down,"default",padx=10,pady=6).pack(side="left")
        self._current_types=current_types

        # ── Team Members tab ──────────────────────────────────────────────────
        tm_frame=self._tab_frames["team"]
        tm_body=tk.Frame(tm_frame,bg=BG,padx=24,pady=16); tm_body.pack(fill="both",expand=True)
        tk.Label(tm_body,text="Team members",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w")
        tk.Label(tm_body,text="Add colleagues who work on queries. They appear as assignable people and get a box on the dashboard.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=540,justify="left").pack(anchor="w",pady=(4,12))

        tm_list_frame=tk.Frame(tm_body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        tm_list_frame.pack(fill="both",expand=True,pady=(0,10))
        tm_lb=tk.Listbox(tm_list_frame,font=(FONT,10),bg=CARD2,fg=TEXT,relief="flat",
                         selectbackground="#0D2E2A",selectforeground=ACCENT2,
                         activestyle="none",bd=0,highlightthickness=0)
        tm_sb=ttk.Scrollbar(tm_list_frame,orient="vertical",command=tm_lb.yview,style="Modern.Vertical.TScrollbar")
        tm_lb.configure(yscrollcommand=tm_sb.set)
        tm_sb.pack(side="right",fill="y"); tm_lb.pack(fill="both",expand=True,padx=4,pady=4)

        current_members=list(self.cfg.get("team_members",[]))
        def refresh_tm_lb():
            tm_lb.delete(0,"end")
            for m in current_members: tm_lb.insert("end",f"  {m}")
        refresh_tm_lb()

        tm_ctrl=tk.Frame(tm_body,bg=BG); tm_ctrl.pack(fill="x",pady=(0,4))
        tm_new_var=tk.StringVar()
        tm_nf=tk.Frame(tm_ctrl,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        tm_nf.pack(side="left",fill="x",expand=True,padx=(0,8))
        tk.Entry(tm_nf,textvariable=tm_new_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=7,highlightthickness=0).pack(fill="x")

        def add_member():
            m=tm_new_var.get().strip()
            if not m: return
            if m in current_members:
                messagebox.showwarning("Duplicate","That person is already in the list.",parent=self); return
            current_members.append(m); tm_new_var.set(""); refresh_tm_lb()

        def remove_member():
            sel=tm_lb.curselection()
            if not sel: return
            current_members.pop(sel[0]); refresh_tm_lb()

        make_btn(tm_ctrl,"Add",add_member,"primary",padx=12,pady=6).pack(side="left",padx=(0,4))
        make_btn(tm_ctrl,"Remove selected",remove_member,"danger",padx=12,pady=6).pack(side="left")
        self._current_members=current_members

        # ── Escalation Rules tab ──────────────────────────────────────────────
        esc_frame=self._tab_frames["escalation"]
        esc_body=tk.Frame(esc_frame,bg=BG,padx=24,pady=16); esc_body.pack(fill="both",expand=True)
        tk.Label(esc_body,text="Automated escalation rules",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w")
        tk.Label(esc_body,text="Set rules for automatically escalating query priorities based on age.\n"
                               "Rules are checked when the app loads and can be run manually from the dashboard.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=540,justify="left").pack(anchor="w",pady=(4,12))

        current_rules=list(self.cfg.get("escalation_rules", []))  # List of {priority, days_old, new_priority}

        esc_list=tk.Frame(esc_body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        esc_list.pack(fill="x",pady=(0,10))

        def refresh_esc():
            for w in esc_list.winfo_children(): w.destroy()
            if not current_rules:
                tk.Label(esc_list,text="No escalation rules configured.",font=(FONT,9),
                         bg=CARD2,fg=MUTED,pady=12).pack(anchor="w",padx=14)
                return
            for i,rule in enumerate(current_rules):
                row=tk.Frame(esc_list,bg=CARD2); row.pack(fill="x",padx=10,pady=4)
                desc = f"If {rule['priority']} priority query is {rule['days_old']} days old → escalate to {rule['new_priority']}"
                tk.Label(row,text=desc,font=(FONT,9),bg=CARD2,fg=TEXT).pack(side="left")
                def rm(idx=i):
                    current_rules.pop(idx); refresh_esc()
                make_btn(row,"✕",rm,"danger",padx=6,pady=2).pack(side="right")

        refresh_esc()

        # Add new rule
        esc_add=tk.Frame(esc_body,bg=BG); esc_add.pack(fill="x",pady=(0,2))
        
        rule_frame=tk.Frame(esc_add,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=12,pady=10)
        rule_frame.pack(fill="x",pady=(0,8))
        
        tk.Label(rule_frame,text="If a query with priority",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left")
        pri_var=tk.StringVar(value="Medium")
        make_combo(rule_frame,pri_var,PRIORITIES,readonly=True,width=10).pack(side="left",padx=(4,4))
        tk.Label(rule_frame,text="is",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,4))
        days_var=tk.StringVar(value="7")
        days_frame=tk.Frame(rule_frame,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        days_frame.pack(side="left",padx=(0,4))
        tk.Entry(days_frame,textvariable=days_var,font=(FONT,9),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=4,width=4,highlightthickness=0).pack()
        tk.Label(rule_frame,text="days old, escalate to",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,4))
        new_pri_var=tk.StringVar(value="High")
        make_combo(rule_frame,new_pri_var,PRIORITIES,readonly=True,width=10).pack(side="left",padx=(0,8))

        def add_rule():
            try:
                days = int(days_var.get().strip())
                if days <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Invalid days","Days must be a positive number.",parent=self)
                return
            
            priority = pri_var.get()
            new_priority = new_pri_var.get()
            
            if priority == new_priority:
                messagebox.showerror("Invalid rule","Priority and new priority cannot be the same.",parent=self)
                return
            
            # Check for duplicate
            for rule in current_rules:
                if (rule['priority'] == priority and rule['days_old'] == days and 
                    rule['new_priority'] == new_priority):
                    messagebox.showwarning("Duplicate rule","This rule already exists.",parent=self)
                    return
            
            current_rules.append({
                'priority': priority,
                'days_old': days,
                'new_priority': new_priority
            })
            refresh_esc()

        make_btn(esc_add,"Add Rule",add_rule,"primary",padx=12,pady=6).pack(side="left")
        self._current_rules=current_rules

        # ── Linked Trackers tab ───────────────────────────────────────────────
        lt_frame=self._tab_frames["trackers"]
        lt_body=tk.Frame(lt_frame,bg=BG,padx=24,pady=16); lt_body.pack(fill="both",expand=True)
        tk.Label(lt_body,text="Linked trackers",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w")
        tk.Label(lt_body,text="Add other teams' query tracker Excel files here.\nYou can then transfer queries directly to them using the ↗ Transfer button.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=540,justify="left").pack(anchor="w",pady=(4,12))

        current_trackers=list(self.cfg.get("linked_trackers",[]))  # [{name, excel_file}]

        lt_list=tk.Frame(lt_body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        lt_list.pack(fill="x",pady=(0,10))

        def refresh_lt():
            for w in lt_list.winfo_children(): w.destroy()
            if not current_trackers:
                tk.Label(lt_list,text="No linked trackers yet.",font=(FONT,9),
                         bg=CARD2,fg=MUTED,pady=12).pack(anchor="w",padx=14)
                return
            for i,t in enumerate(current_trackers):
                row=tk.Frame(lt_list,bg=CARD2); row.pack(fill="x",padx=10,pady=4)
                tk.Label(row,text=t.get("name","Unnamed"),font=(FONT,10,"bold"),
                         bg=CARD2,fg=TEXT).pack(side="left")
                detail=tk.Frame(row,bg=CARD2); detail.pack(side="left",padx=8)
                tk.Label(detail,text=t.get("excel_file","—"),font=(FONT,8),bg=CARD2,fg=TEXT2).pack(anchor="w")
                if t.get("sites_file"):
                    tk.Label(detail,text=f"Sites: {t['sites_file']}",font=(FONT,7),bg=CARD2,fg=MUTED).pack(anchor="w")
                def rm(idx=i):
                    current_trackers.pop(idx); refresh_lt()
                make_btn(row,"✕",rm,"danger",padx=6,pady=2).pack(side="right")

        refresh_lt()

        # Add row - two lines: name + excel file, then sites file below
        lt_add=tk.Frame(lt_body,bg=BG); lt_add.pack(fill="x",pady=(0,2))
        lt_name_var=tk.StringVar()
        lt_file_var=tk.StringVar()
        lt_sites_var=tk.StringVar()

        # Row 1: Name + tracker Excel
        nf2=tk.Frame(lt_add,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        nf2.pack(side="left",padx=(0,6))
        tk.Entry(nf2,textvariable=lt_name_var,font=(FONT,9),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=6,width=14,highlightthickness=0).pack()
        tk.Label(lt_add,text="Name",font=(FONT,7),bg=BG,fg=MUTED).pack(side="left",padx=(0,6))
        ff2=tk.Frame(lt_add,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        ff2.pack(side="left",padx=(0,4),fill="x",expand=True)
        tk.Entry(ff2,textvariable=lt_file_var,font=(FONT,9),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=6,width=22,highlightthickness=0).pack(fill="x")
        tk.Label(lt_add,text="Tracker",font=(FONT,7),bg=BG,fg=MUTED).pack(side="left",padx=(0,4))

        def browse_lt():
            p=filedialog.askopenfilename(title="Find tracker Excel",filetypes=[("Excel files","*.xlsx")])
            if p: lt_file_var.set(p)

        make_btn(lt_add,"📁",browse_lt,"nav",padx=7,pady=4).pack(side="left")

        # Row 2: sites.xlsx for that tracker
        lt_add2=tk.Frame(lt_body,bg=BG); lt_add2.pack(fill="x",pady=(0,4))
        tk.Label(lt_add2,text="Their sites.xlsx (optional — lets transferred queries auto-add the site):",
                 font=(FONT,8),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,3))
        lt_add2b=tk.Frame(lt_add2,bg=BG); lt_add2b.pack(fill="x")
        sf2=tk.Frame(lt_add2b,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        sf2.pack(side="left",fill="x",expand=True,padx=(0,4))
        tk.Entry(sf2,textvariable=lt_sites_var,font=(FONT,9),bg=CARD2,fg=TEXT,insertbackground=TEXT,
                 relief="flat",bd=6,highlightthickness=0).pack(fill="x")
        def browse_lt_sites():
            p=filedialog.askopenfilename(title="Find their sites.xlsx",filetypes=[("Excel files","*.xlsx")])
            if p: lt_sites_var.set(p)
        make_btn(lt_add2b,"📁",browse_lt_sites,"nav",padx=7,pady=4).pack(side="left",padx=(0,8))

        def add_tracker():
            name=lt_name_var.get().strip(); fpath=lt_file_var.get().strip()
            if not name or not fpath:
                messagebox.showwarning("Required","Please enter a name and choose the tracker Excel file.",parent=self); return
            current_trackers.append({"name":name,"excel_file":fpath,"sites_file":lt_sites_var.get().strip()})
            lt_name_var.set(""); lt_file_var.set(""); lt_sites_var.set(""); refresh_lt()

        make_btn(lt_add2b,"Add tracker",add_tracker,"primary",padx=12,pady=5).pack(side="left")
        tk.Label(lt_body,text="  Tip: give each tracker a short name like 'Savills COT team' or 'Recharge inbox'",
                 font=(FONT,8),bg=BG,fg=MUTED).pack(anchor="w",pady=(4,0))
        self._current_trackers=current_trackers

        # ── Theme tab ─────────────────────────────────────────────────────────
        th_frame=self._tab_frames["theme"]
        th_wrap=tk.Frame(th_frame,bg=BG)
        th_wrap.pack(fill="both",expand=True)
        th_canvas=tk.Canvas(th_wrap,bg=BG,highlightthickness=0,bd=0)
        th_scroll=ttk.Scrollbar(th_wrap,orient="vertical",command=th_canvas.yview,style="Modern.Vertical.TScrollbar")
        th_scroll.pack(side="right",fill="y")
        th_canvas.pack(side="left",fill="both",expand=True)
        th_body=tk.Frame(th_canvas,bg=BG,padx=24,pady=16)
        th_window=th_canvas.create_window((0,0),window=th_body,anchor="nw")
        th_canvas.configure(yscrollcommand=th_scroll.set)
        th_body.bind("<Configure>",lambda e:th_canvas.configure(scrollregion=th_canvas.bbox("all")))
        th_canvas.bind("<Configure>",lambda e:th_canvas.itemconfig(th_window,width=e.width))

        def _theme_scroll(e):
            if not (self.winfo_exists() and th_canvas.winfo_exists()):
                return "break"
            delta=getattr(e,"delta",0)
            try:
                if delta:
                    step=-1 if delta>0 else 1
                    th_canvas.yview_scroll(step,"units")
                elif getattr(e,"num",None)==4:
                    th_canvas.yview_scroll(-1,"units")
                elif getattr(e,"num",None)==5:
                    th_canvas.yview_scroll(1,"units")
            except tk.TclError:
                return "break"
            return "break"

        def _bind_theme_wheel(_e=None):
            if not self.winfo_exists():
                return
            self.bind_all("<MouseWheel>",_theme_scroll)
            self.bind_all("<Button-4>",_theme_scroll)
            self.bind_all("<Button-5>",_theme_scroll)

        def _unbind_theme_wheel(_e=None):
            if not self.winfo_exists():
                return
            self.unbind_all("<MouseWheel>")
            self.unbind_all("<Button-4>")
            self.unbind_all("<Button-5>")

        def _cleanup_theme_wheel(e):
            if e.widget is self:
                try: _unbind_theme_wheel()
                except: pass

        self._theme_bind_wheel=_bind_theme_wheel
        self._theme_unbind_wheel=_unbind_theme_wheel
        self._theme_canvas=th_canvas
        self.bind("<Destroy>",_cleanup_theme_wheel,add="+")

        tk.Label(th_body,text="Colour theme",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w")
        tk.Label(th_body,text="Choose a colour theme. The change takes effect when you Save & apply and reopen the app.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=620,justify="left").pack(anchor="w",pady=(4,8))
        tk.Label(th_body,text="Each option also includes a small layout preview so you can compare the overall feel of the interface, not just the colours.",
                 font=(FONT,9),bg=BG,fg=TEXT2,wraplength=620,justify="left").pack(anchor="w",pady=(0,14))

        self._selected_theme=tk.StringVar(value=self.cfg.get("theme","Slate & Teal"))

        theme_cards={}

        def select_theme(name):
            self._selected_theme.set(name)
            # Update card borders
            for n,card in theme_cards.items():
                try:
                    card.configure(
                        highlightbackground=ACCENT if n==name else BORDER,
                        highlightthickness=2 if n==name else 1
                    )
                except: pass

        # Single responsive column so the theme page always fits smaller settings windows.
        row_frame=None
        for i,(name,t) in enumerate(THEMES.items()):
            row_frame=tk.Frame(th_body,bg=BG)
            row_frame.pack(fill="x",pady=(0,10))

            is_selected=(name==self._selected_theme.get())
            card=tk.Frame(row_frame,bg=t["CARD"],
                          highlightthickness=2 if is_selected else 1,
                          highlightbackground=ACCENT if is_selected else BORDER,
                          cursor="hand2",padx=14,pady=12,height=190)
            card.pack(side="left",fill="x",expand=True)
            card.pack_propagate(False)
            theme_cards[name]=card

            # Theme name
            tk.Label(card,text=name,font=(FONT,9,"bold"),
                     bg=t["CARD"],fg=t["TEXT"]).pack(anchor="w")
            summary, direction = THEME_NOTES.get(name,("", ""))
            tk.Label(card,text=summary,font=(FONT,8),bg=t["CARD"],fg=t["TEXT2"],
                     wraplength=290,justify="left").pack(anchor="w",pady=(3,0))

            # Colour swatch strip
            swatch_row=tk.Frame(card,bg=t["CARD"]); swatch_row.pack(anchor="w",pady=(6,4))
            for col in [t["BG"],t["ACCENT"],t["ACCENT2"],t["SUCCESS"],t["DANGER"],t["WARNING"]]:
                sw=tk.Frame(swatch_row,bg=col,width=18,height=18); sw.pack(side="left",padx=1)

            # Mini preview showing the intended layout style for the theme.
            prev=tk.Frame(card,bg=t["BG"],padx=8,pady=8,
                          highlightthickness=1,highlightbackground=t["BORDER"])
            prev.pack(fill="x",pady=(4,6))
            nav_strip=tk.Frame(prev,bg=t["NAV"],height=22)
            nav_strip.pack(fill="x")
            tk.Frame(nav_strip,bg=t["ACCENT"],width=24,height=12).pack(side="left",padx=(8,6),pady=5)
            tk.Frame(nav_strip,bg=t.get("NAV_MUTED",t["TEXT2"]),width=48,height=4).pack(side="left",pady=9)
            tk.Frame(nav_strip,bg=t.get("NAV_MUTED",t["TEXT2"]),width=32,height=4).pack(side="left",padx=(6,0),pady=9)
            tk.Frame(nav_strip,bg=t["ACCENT"],width=42,height=12).pack(side="right",padx=8,pady=5)
            tabs_strip=tk.Frame(prev,bg=t["NAV2"],height=18)
            tabs_strip.pack(fill="x",pady=(4,6))
            tk.Frame(tabs_strip,bg=t["ACCENT2"],width=42,height=10).pack(side="left",padx=(8,6),pady=4)
            tk.Frame(tabs_strip,bg=t.get("NAV_MUTED",t["TEXT2"]),width=34,height=4).pack(side="left",pady=7)
            tk.Frame(tabs_strip,bg=t.get("NAV_MUTED",t["TEXT2"]),width=28,height=4).pack(side="left",padx=(6,0),pady=7)
            content_preview=tk.Frame(prev,bg=t["BG"])
            content_preview.pack(fill="both",expand=True)
            info_block=tk.Frame(content_preview,bg=t["CARD"],height=44,
                                highlightthickness=1,highlightbackground=t["BORDER"])
            info_block.pack(fill="x")
            info_block.pack_propagate(False)
            tk.Frame(info_block,bg=t["ACCENT"],width=4).pack(side="left",fill="y")
            text_stack=tk.Frame(info_block,bg=t["CARD"])
            text_stack.pack(side="left",fill="both",expand=True,padx=8,pady=8)
            tk.Frame(text_stack,bg=t["TEXT"],width=70,height=5).pack(anchor="w")
            tk.Frame(text_stack,bg=t["TEXT2"],width=110,height=4).pack(anchor="w",pady=(6,0))
            tk.Frame(text_stack,bg=t["TEXT2"],width=82,height=4).pack(anchor="w",pady=(5,0))
            action_row=tk.Frame(content_preview,bg=t["BG"])
            action_row.pack(fill="x",pady=(6,0))
            tk.Frame(action_row,bg=t["CARD2"],height=18,width=92,
                     highlightthickness=1,highlightbackground=t["BORDER"]).pack(side="left")
            tk.Frame(action_row,bg=t["ACCENT"],height=18,width=72).pack(side="right")
            tk.Label(card,text=direction,font=(FONT,8),bg=t["CARD"],fg=t["TEXT2"],
                     wraplength=290,justify="left").pack(anchor="w")

            def _click(e,n=name): select_theme(n)
            for w in [card]+list(card.winfo_children()):
                try: w.bind("<Button-1>",_click)
                except: pass
            # Recursive bind for swatch children
            for sw in swatch_row.winfo_children():
                sw.bind("<Button-1>",_click)
            for pw in prev.winfo_children():
                pw.bind("<Button-1>",_click)
                for pw2 in pw.winfo_children():
                    try: pw2.bind("<Button-1>",_click)
                    except: pass

        self.after(50,lambda:th_canvas.configure(scrollregion=th_canvas.bbox("all")))

        # ── Footer ────────────────────────────────────────────────────────────
        btns=tk.Frame(self,bg=CARD2,padx=20,pady=12); btns.pack(fill="x",side="bottom")
        tk.Frame(self,bg=BORDER,height=1).pack(fill="x",side="bottom")
        if existing_cfg:
            make_btn(btns,"Cancel",self.destroy,"default",padx=14,pady=7).pack(side="right",padx=(8,0))
            make_btn(btns,"Apply",lambda:self._save(close_after=False),"default",padx=14,pady=7).pack(side="right",padx=(8,0))
        make_btn(btns,"Save & apply",self._save,"primary",padx=18,pady=7).pack(side="right")

    def _switch_stab(self,tab_id,init=False):
        if getattr(self,"_active_stab",None) and self._active_stab.get()=="theme":
            try: self._theme_unbind_wheel()
            except: pass
        for tid,f in self._tab_frames.items(): f.pack_forget()
        self._tab_frames[tab_id].pack(fill="both",expand=True)
        self._active_stab.set(tab_id)
        if tab_id=="theme":
            try:
                self.update_idletasks()
                self._theme_canvas.configure(scrollregion=self._theme_canvas.bbox("all"))
                self._theme_bind_wheel()
            except:
                pass
        for tid,btn in self._stab_btns.items():
            if tid==tab_id:
                btn.configure(fg=ACCENT2,font=(FONT,10,"bold"),bg=CARD2)
            else:
                btn.configure(fg=NAV_MUTED,font=(FONT,10),bg=NAV2)

    def _save(self, close_after=True):
        excel=self.excel_var.get().strip(); name=self.name_var.get().strip()
        if not excel:
            messagebox.showwarning("Required","Please choose a location for the query data file.",parent=self); return
        if not name:
            messagebox.showwarning("Required","Please enter your name.",parent=self); return
        try:
            hvt=int(self.high_volume_threshold_var.get().strip())
            if hvt < 1: raise ValueError()
        except:
            messagebox.showwarning("Invalid threshold",
                "High-volume warning threshold must be a whole number greater than 0.",parent=self); return
        chosen_theme=getattr(self,"_selected_theme",None)
        chosen_theme=chosen_theme.get() if chosen_theme else self.cfg.get("theme","Slate & Teal")
        self.cfg.update({
            "excel_file":     excel,
            "sites_file":     self.sites_var.get().strip(),
            "username":       name,
            "high_volume_threshold": hvt,
            "query_types":    self._current_types,
            "team_members":   self._current_members,
            "linked_trackers":self._current_trackers,
            "escalation_rules": self._current_rules,
            "theme":          chosen_theme,
        })
        save_config(self.cfg)
        # Also write query types and team members to the shared tracker so all
        # teammates pick them up automatically on next reload.
        save_shared_settings(excel, self._current_types, self._current_members)
        global QUERY_TYPES
        QUERY_TYPES[:] = self._current_types
        theme_changed=(chosen_theme!=ACTIVE_THEME)
        if close_after:
            self.destroy()
            if self.on_complete: self.on_complete(self.cfg)
        else:
            messagebox.showinfo("Applied","Settings updated successfully.",parent=self)
        if theme_changed:
            messagebox.showinfo("Theme changed",
                f"Theme set to '{chosen_theme}'.\n\nClose and reopen the app to apply the new theme fully.")

class QueryTrackerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Query Tracker"); self.geometry("1400x840")
        self.configure(bg=BG); self.resizable(True,True); self.withdraw()
        
        # Set app icon
        try:
            icon_path = resource_path("qbox-icon-256.png")
            if os.path.exists(icon_path):
                photo = tk.PhotoImage(file=icon_path)
                self.iconphoto(False, photo)
                self._icon = photo  # Keep reference to prevent garbage collection
        except Exception as e:
            pass  # Icon loading is optional
        
        apply_styles()
        self.cfg=load_config()
        self._watcher_running=False
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        if not self.cfg.get("excel_file") or not self.cfg.get("username"):
            self._run_wizard(first_run=True)
        else:
            self._launch()

    def _on_close(self):
        self._watcher_running = False
        self._auto_reload_running = False
        self._backup_running = False
        self._do_daily_backup(forced=True)   # snapshot on exit
        self.destroy()

    def _toggle_mini_window(self):
        """Show or hide the compact always-on-top mini dashboard window."""
        if getattr(self,"_mini_win",None) and self._mini_win.winfo_exists():
            self._mini_win.destroy(); self._mini_win=None; return
        self._create_mini_window()

    def _create_mini_window(self):
        win=tk.Toplevel(self)
        win.title("QBOX Today"); win.configure(bg=NAV)
        win.attributes("-topmost",True)
        win.overrideredirect(True)
        win.resizable(False,False)
        self._mini_win=win

        # Position bottom-right of screen
        sw=win.winfo_screenwidth(); sh=win.winfo_screenheight()
        ww,wh=320,200  # Increased width and height for priority breakdown
        win.geometry(f"{ww}x{wh}+{sw-ww-20}+{sh-wh-60}")

        # Draggable
        _drag={"x":0,"y":0}
        def on_drag_start(e): _drag["x"]=e.x; _drag["y"]=e.y
        def on_drag(e):
            win.geometry(f"+{win.winfo_x()+e.x-_drag['x']}+{win.winfo_y()+e.y-_drag['y']}")
        win.bind("<ButtonPress-1>",on_drag_start)
        win.bind("<B1-Motion>",on_drag)

        # Header
        hdr=tk.Frame(win,bg=ACCENT,padx=10,pady=5); hdr.pack(fill="x")
        tk.Label(hdr,text=" QBOX  Today",font=(FONT,8,"bold"),bg=ACCENT,fg=NAV).pack(side="left")
        cl=tk.Label(hdr,text="✕",font=(FONT,9,"bold"),bg=ACCENT,fg=NAV,cursor="hand2",padx=8)
        cl.pack(side="right")
        cl.bind("<Button-1>",lambda e:self._toggle_mini_window())

        # Body
        body=tk.Frame(win,bg=NAV,padx=16,pady=10); body.pack(fill="both",expand=True)
        # Change layout to show priority breakdown
        pri_frame=tk.Frame(body,bg=NAV); pri_frame.pack(fill="x")
        
        # Priority breakdown
        priorities = ["High", "Medium", "Low"]
        pri_counts = {}
        t=today_str()
        due=[q for q in self.queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=t]
        mine=[q for q in due if q.get("assigned_to","")==self.username]
        
        for pri in priorities:
            pri_counts[pri] = len([q for q in mine if q.get("priority", "Low") == pri])
        
        total_mine = len(mine)
        total_team = len(due)
        
        # Show breakdown
        for pri in priorities:
            count = pri_counts[pri]
            if count > 0:
                pri_row=tk.Frame(pri_frame,bg=NAV); pri_row.pack(fill="x",pady=1)
                pc,pbg=P_COLORS.get(pri,("#94A3B8","#1E2D45"))
                pri_pill=tk.Label(pri_row,text=f"  {pri}  ",font=(FONT,7,"bold"),bg=pbg,fg=pc,padx=4,pady=2)
                pri_pill.pack(side="left")
                count_label=tk.Label(pri_row,text=str(count),font=(FONT,12,"bold"),bg=NAV,fg=pc)
                count_label.pack(side="left",padx=(4,0))
        
        # Total summary
        if total_mine > 0:
            summary_frame=tk.Frame(body,bg=NAV); summary_frame.pack(fill="x",pady=(6,0))
            tk.Label(summary_frame,text=f"Total: {total_mine} action{'s' if total_mine!=1 else ''} today",font=(FONT,8),bg=NAV,fg=TEXT2).pack(anchor="w")
            if total_team > total_mine:
                tk.Label(summary_frame,text=f"{total_team} total across team",font=(FONT,8),bg=NAV,fg=MUTED).pack(anchor="w")
        else:
            tk.Label(body,text="All clear ✓",font=(FONT,12,"bold"),bg=NAV,fg=SUCCESS).pack(anchor="w",pady=(10,0))
        
        btn_row=tk.Frame(body,bg=NAV); btn_row.pack(fill="x",side="bottom")
        def go():
            self.deiconify(); self.lift(); self.focus_force(); self._go_list("action")
        make_btn(btn_row,"View →",go,"active",padx=10,pady=3).pack(side="left")
        make_btn(btn_row,"＋ New",self._open_add_dialog,"primary",padx=8,pady=3).pack(side="left",padx=(6,0))

        def refresh():
            if not self._mini_win or not self._mini_win.winfo_exists(): return
            # Clear existing priority breakdown
            for w in pri_frame.winfo_children(): w.destroy()
            
            t=today_str()
            due=[q for q in self.queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=t]
            mine=[q for q in due if q.get("assigned_to","")==self.username]
            
            priorities = ["High", "Medium", "Low"]
            pri_counts = {}
            for pri in priorities:
                pri_counts[pri] = len([q for q in mine if q.get("priority", "Low") == pri])
            
            total_mine = len(mine)
            total_team = len(due)
            
            # Show breakdown
            for pri in priorities:
                count = pri_counts[pri]
                if count > 0:
                    pri_row=tk.Frame(pri_frame,bg=NAV); pri_row.pack(fill="x",pady=1)
                    pc,pbg=P_COLORS.get(pri,("#94A3B8","#1E2D45"))
                    pri_pill=tk.Label(pri_row,text=f"  {pri}  ",font=(FONT,7,"bold"),bg=pbg,fg=pc,padx=4,pady=2)
                    pri_pill.pack(side="left")
                    count_label=tk.Label(pri_row,text=str(count),font=(FONT,12,"bold"),bg=NAV,fg=pc)
                    count_label.pack(side="left",padx=(4,0))
            
            # Clear and update summary
            for w in body.winfo_children():
                if hasattr(w, 'winfo_children') and w != pri_frame and w != btn_row:
                    for sw in w.winfo_children(): sw.destroy()
                    w.destroy()
            
            # Recreate summary frame
            if total_mine > 0:
                summary_frame=tk.Frame(body,bg=NAV); summary_frame.pack(fill="x",pady=(6,0))
                tk.Label(summary_frame,text=f"Total: {total_mine} action{'s' if total_mine!=1 else ''} today",font=(FONT,8),bg=NAV,fg=TEXT2).pack(anchor="w")
                if total_team > total_mine:
                    tk.Label(summary_frame,text=f"{total_team} total across team",font=(FONT,8),bg=NAV,fg=MUTED).pack(anchor="w")
            else:
                tk.Label(body,text="All clear ✓",font=(FONT,12,"bold"),bg=NAV,fg=SUCCESS).pack(anchor="w",pady=(10,0))
            
            # Re-pack button row at bottom
            btn_row.pack_forget()
            btn_row.pack(fill="x",side="bottom")
            
            self._mini_win.after(30000,refresh)
        refresh()
        self._mini_refresh=refresh

    def _run_wizard(self,first_run=False):
        def on_complete(cfg): self.cfg=cfg; self._launch()
        w=SetupWizard(self,existing_cfg=self.cfg,on_complete=on_complete)
        if first_run: w.protocol("WM_DELETE_WINDOW",self.destroy)
        self.wait_window(w)

    def _launch(self):
        global QUERY_TYPES
        self.username=self.cfg.get("username","Unknown")
        self.excel_file=self.cfg.get("excel_file","")
        self.sites_file=self.cfg.get("sites_file","")
        # Load query types and team members from the shared tracker so all
        # teammates automatically pick up any changes made by others.
        shared=load_shared_settings(self.excel_file)
        QUERY_TYPES[:]=shared.get("query_types") or self.cfg.get("query_types",DEFAULT_QUERY_TYPES)
        self.team_members=shared.get("team_members") or self.cfg.get("team_members",[])
        self.linked_trackers=self.cfg.get("linked_trackers",[])
        self.out_of_office=self._normalize_out_of_office(self.cfg.get("out_of_office",[]))
        self._assignee_filter=""
        self._calendar_day_filter=""
        # Apply saved theme before building UI so all colour constants are set
        apply_theme(self.cfg.get("theme","Slate & Teal"))
        self.clients,self.sites_by_client,self.meters,self.utilities_by_site,\
            self.funds_by_client,self.sites_by_fund=load_site_data(self.sites_file)
        self.queries=load_queries(self.excel_file)
        self._apply_escalation_rules()  # Apply escalation rules after loading
        self.recent_ids=[]
        self._watcher_running=False
        self._watcher_thread=None
        self._watcher_seen=set()
        # Auto-reload state
        self._auto_reload_running=False
        self._auto_reload_thread=None
        self._excel_mtime=self._get_excel_mtime()
        self._reload_tick=0
        self._build_ui(); self.deiconify()
        self._refresh_table(); self._show_daily_banner()
        self._start_auto_reload()
        self._start_daily_backup()
        self._start_watcher()   # run from launch, not just when on list page
        try: self.sync_lbl.config(text="● Live")
        except: pass
        self.after(1200,self._check_notifications)  # slight delay so UI is fully drawn

    def _build_ui(self):
        # ── Top navigation bar ────────────────────────────────────────────────
        nav=tk.Frame(self,bg=NAV,pady=0); nav.pack(fill="x")
        inner_nav=tk.Frame(nav,bg=NAV); inner_nav.pack(fill="x",padx=20,pady=14)
        left=tk.Frame(inner_nav,bg=NAV); left.pack(side="left")
        # Logo pill
        logo_pill=tk.Frame(left,bg=ACCENT,padx=8,pady=3); logo_pill.pack(side="left",padx=(0,12))
        logo_img=None
        for icon_name in ["qbox-icon-64.png","qbox-icon-128.png","qbox-icon-256.png"]:
            icon_path=resource_path(icon_name)
            if os.path.exists(icon_path):
                try:
                    logo_img=tk.PhotoImage(file=icon_path)
                    break
                except Exception:
                    logo_img=None
        if logo_img:
            try:
                logo_img=logo_img.subsample(max(1,logo_img.width()//20),max(1,logo_img.height()//20))
            except Exception:
                pass
            self._brand_logo=logo_img
            tk.Label(logo_pill,image=self._brand_logo,bg=ACCENT).pack(side="left")
            tk.Label(logo_pill,text="QBOX",font=(FONT,10,"bold"),bg=ACCENT,fg=PRIMARY_FG,padx=6).pack(side="left")
        else:
            tk.Label(logo_pill,text="QBOX",font=(FONT,11,"bold"),bg=ACCENT,fg=PRIMARY_FG,pady=5,padx=8).pack()
        tk.Label(left,text="Query Tracker",font=(FONT,13,"bold"),bg=NAV,fg=NAV_TEXT).pack(side="left")
        tk.Label(left,text="  ·  ",font=(FONT,11),bg=NAV,fg=BORDER).pack(side="left")
        tk.Label(left,text=self.username,font=(FONT,10),bg=NAV,fg=NAV_MUTED).pack(side="left")
        right=tk.Frame(inner_nav,bg=NAV); right.pack(side="right")
        make_btn(right,"⚙  Settings",self._open_settings,"nav",padx=11,pady=5).pack(side="right",padx=2)
        make_btn(right,"↻  Reload",self._reload_sites,"nav",padx=11,pady=5).pack(side="right",padx=2)
        make_btn(right,"□  Mini",self._toggle_mini_window,"nav",padx=11,pady=5).pack(side="right",padx=2)
        tk.Frame(right,bg=BORDER,width=1,height=24).pack(side="right",padx=10)
        make_btn(right,"＋  New query",self._open_add_dialog,"primary",padx=15,pady=5).pack(side="right",padx=2)
        # Bottom accent stripe
        tk.Frame(self,bg=ACCENT,height=2).pack(fill="x")

        # ── Page navigation tabs ──────────────────────────────────────────────
        pnav=tk.Frame(self,bg=NAV2); pnav.pack(fill="x")
        tk.Frame(self,bg=BORDER,height=1).pack(fill="x")
        self._page_btns={}
        for text,val,icon in [("  Dashboard","dashboard","◈"),("  Query List","list","≡"),("  Calendar","calendar","🗓"),("  Reports","reports","⬡")]:
            btn=tk.Label(pnav,text=f"{icon}{text}  ",font=(FONT,10),bg=NAV2,
                         fg=NAV_MUTED,cursor="hand2",padx=8,pady=11)
            btn.pack(side="left")
            btn.bind("<Button-1>",lambda e,v=val:self._show_page(v))
            self._page_btns[val]=btn

        self.page_container=tk.Frame(self,bg=BG); self.page_container.pack(fill="both",expand=True)
        self.dashboard_page=tk.Frame(self.page_container,bg=BG)
        self.list_page=tk.Frame(self.page_container,bg=BG)
        self.calendar_page=tk.Frame(self.page_container,bg=BG)
        self.reports_page=tk.Frame(self.page_container,bg=BG)

        lp=self.list_page
        self.metrics_frame=tk.Frame(lp,bg=BG); self.metrics_frame.pack(fill="x",padx=20,pady=(18,0))
        self.banner_frame=tk.Frame(lp,bg=CARD2,highlightbackground=BORDER,highlightthickness=1)
        self.banner_strip=tk.Frame(self.banner_frame,bg=DANGER,width=4)
        self.banner_strip.pack(side="left",fill="y")
        self.banner_lbl=tk.Label(self.banner_frame,text="",font=(FONT,9,"bold"),bg=CARD2,
                     fg=TEXT,wraplength=1200,justify="left",anchor="w")
        self.banner_lbl.pack(side="left",fill="x",expand=True,padx=14,pady=10)

        ctrl_wrap=tk.Frame(lp,bg=BG); ctrl_wrap.pack(fill="x",padx=20,pady=(16,0))
        ctrl=tk.Frame(ctrl_wrap,bg=BG); ctrl.pack(side="left")
        self.tab_var=tk.StringVar(value="open"); self._tab_btns={}
        for text,val in [("Action today","action"),("Open","open"),("All queries","all"),("Resolved","resolved")]:
            btn=tk.Label(ctrl,text=text,font=(FONT,9),bg=BG,fg=MUTED,cursor="hand2",padx=14,pady=7)
            btn.pack(side="left",padx=1)
            btn.bind("<Button-1>",lambda e,v=val:self._set_tab(v))
            self._tab_btns[val]=btn
        self._set_tab("open",refresh=False)

        filters_wrap=tk.Frame(lp,bg=BG); filters_wrap.pack(fill="x",padx=20,pady=(8,0))
        frow_top=tk.Frame(filters_wrap,bg=BG); frow_top.pack(fill="x")
        frow_bottom=tk.Frame(filters_wrap,bg=BG); frow_bottom.pack(fill="x",pady=(8,0))

        def _lbl_combo(parent, label, widget_fn, side="left", pad_right=8):
            """Wrap a filter combo with a small label above it."""
            wrap=tk.Frame(parent,bg=BG); wrap.pack(side=side,padx=(0,pad_right))
            tk.Label(wrap,text=label,font=(FONT,7,"bold"),bg=BG,fg=MUTED,anchor="w").pack(fill="x")
            w=widget_fn(wrap); w.pack(fill="x")
            return w

        sc_wrap=tk.Frame(frow_top,bg=BG); sc_wrap.pack(side="left",padx=(0,12))
        tk.Label(sc_wrap,text="SEARCH",font=(FONT,7,"bold"),bg=BG,fg=MUTED,anchor="w").pack(fill="x")
        sc=tk.Frame(sc_wrap,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); sc.pack()
        tk.Label(sc,text="🔍",font=(FONT,10),bg=CARD2,fg=MUTED).pack(side="left",padx=(8,0))
        self.search_var=tk.StringVar()
        self.search_var.trace_add("write",lambda *_:self._refresh_table())
        tk.Entry(sc,textvariable=self.search_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,width=22,highlightthickness=0).pack(side="left")

        self.filter_client=_lbl_combo(frow_top,"CLIENT",
            lambda p: make_combo(p,tk.StringVar(),["All"]+self.clients,readonly=True,width=20))
        self.filter_client.set("All")
        self.filter_fund=_lbl_combo(frow_top,"FUND",
            lambda p: make_combo(p,tk.StringVar(),["All"],readonly=True,width=24))
        self.filter_fund.set("All")

        def on_filter_client(e=None):
            c=self.filter_client.get()
            funds=self.funds_by_client.get(c,[]) if c!="All" else sorted({f for fl in self.funds_by_client.values() for f in fl})
            self.filter_fund.configure(values=["All"]+funds); self.filter_fund.set("All"); self._refresh_table()
        self.filter_client.bind("<<ComboboxSelected>>",on_filter_client)
        self.filter_fund.bind("<<ComboboxSelected>>",lambda _:self._refresh_table())

        self.filter_type=_lbl_combo(frow_top,"TYPE",
            lambda p: make_combo(p,tk.StringVar(),["All"]+QUERY_TYPES,readonly=True,width=20))
        self.filter_type.set("All"); self.filter_type.bind("<<ComboboxSelected>>",lambda _:self._refresh_table())

        self.filter_status=_lbl_combo(frow_bottom,"STATUS",
            lambda p: make_combo(p,tk.StringVar(),["All"]+STATUSES,readonly=True,width=13))
        self.filter_status.set("All"); self.filter_status.bind("<<ComboboxSelected>>",lambda _:self._refresh_table())
        all_utils=sorted({q.get("utility","") for q in self.queries if q.get("utility","")})
        self.filter_utility=_lbl_combo(frow_bottom,"UTILITY",
            lambda p: make_combo(p,tk.StringVar(),["All"]+all_utils,readonly=True,width=14),pad_right=12)
        self.filter_utility.set("All"); self.filter_utility.bind("<<ComboboxSelected>>",lambda _:self._refresh_table())
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        self.filter_assignee=_lbl_combo(frow_bottom,"ASSIGNED TO",
            lambda p: make_combo(p,tk.StringVar(),["All"]+all_members,readonly=True,width=16))
        self.filter_assignee.set("All")
        self.filter_assignee.bind("<<ComboboxSelected>>",lambda _:self._refresh_table())
        btn_wrap=tk.Frame(frow_bottom,bg=BG); btn_wrap.pack(side="left",padx=(4,0))
        tk.Label(btn_wrap,text=" ",font=(FONT,7),bg=BG).pack()  # spacer to align buttons with combos
        make_btn(btn_wrap,"Clear",self._clear_filters,"default",padx=12,pady=6).pack(side="left")
        make_btn(btn_wrap,"Bulk Actions",self._open_bulk_actions,"primary",padx=12,pady=6).pack(side="left",padx=(8,0))

        # Query list colour key
        key_wrap=tk.Frame(lp,bg=BG)
        key_wrap.pack(fill="x",padx=20,pady=(8,0))
        key_card=tk.Frame(key_wrap,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=10,pady=6)
        key_card.pack(fill="x")
        tk.Label(key_card,text="Colour key:",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED).pack(side="left",padx=(0,8))

        def add_key_item(parent,text,bg_col,fg_col):
            item=tk.Frame(parent,bg=CARD2)
            item.pack(side="left",padx=(0,10))
            swatch=tk.Frame(item,bg=bg_col,highlightthickness=1,highlightbackground=fg_col,width=14,height=14)
            swatch.pack(side="left")
            swatch.pack_propagate(False)
            tk.Label(item,text=text,font=(FONT,8),bg=CARD2,fg=TEXT2).pack(side="left",padx=(5,0))

        add_key_item(key_card,"Overdue",   "#2A1010", "#FCA5A5")
        add_key_item(key_card,"Stale >30d", "#231600", "#FB923C")
        add_key_item(key_card,"Aging >14d", "#1E1900", "#FBBF24")
        add_key_item(key_card,"In Progress", "#0E1628", "#93C5FD")
        add_key_item(key_card,"High priority", "#1A1028", "#C084FC")
        add_key_item(key_card,"Resolved", CARD, MUTED)

        tw=tk.Frame(lp,bg=BG); tw.pack(fill="both",expand=True,padx=20,pady=(10,0))
        card=tk.Frame(tw,bg=CARD,highlightthickness=1,highlightbackground=BORDER); card.pack(fill="both",expand=True)
        cols=("ref","client","fund","site","utility","meter","type","status","priority","overdue","chase","raised","opened","assigned","last_by","att")
        self.tree=ttk.Treeview(card,columns=cols,show="headings",height=20,style="Modern.Treeview",selectmode="extended")
        self._sort_state={}
        for col,label,width,anchor in [
            ("ref","Reference",90,"w"),("client","Client",140,"w"),("fund","Fund",140,"w"),
            ("site","Site",155,"w"),("utility","Utility",90,"w"),("meter","Meter",130,"w"),
            ("type","Type",140,"w"),("status","Status",100,"w"),("priority","Priority",75,"center"),
            ("overdue","Overdue",75,"center"),("chase","Next Action",110,"w"),
            ("raised","Raised",85,"w"),("opened","Logged",85,"w"),
            ("assigned","Assigned to",110,"w"),
            ("last_by","Last updated by",120,"w"),
            ("att","📎",40,"center")]:
            self.tree.heading(col,text=label,command=lambda c=col:self._sort_by(c))
            self.tree.column(col,width=width,minwidth=40,anchor=anchor,stretch=False)
        # Row colour priority (highest to lowest):
        # overdue   = past action date + not resolved  → red tint
        # stale     = open query >30 days old          → orange tint
        # aging     = open query >14 days old          → amber tint
        # in_prog   = status "In Progress"             → blue tint
        # high      = High priority, not overdue       → soft purple tint
        # open      = normal open query                → plain card
        # resolved  = resolved                         → dimmed text
        if ACTIVE_THEME == "Mist Light":
            self.tree.tag_configure("overdue",  background="#F9D8D8", foreground="#7A1A1A")
            self.tree.tag_configure("stale",    background="#FAF0CC", foreground="#7A4800")
            self.tree.tag_configure("aging",    background="#FEFBE8", foreground="#6B5000")
            self.tree.tag_configure("in_prog",  background="#DCEEFF", foreground="#1A4070")
            self.tree.tag_configure("high",     background="#EDE0FF", foreground="#4A1A7A")
            self.tree.tag_configure("open",     background=CARD,      foreground=TEXT)
            self.tree.tag_configure("resolved", background=CARD,      foreground=MUTED)
        else:
            self.tree.tag_configure("overdue",  background="#2A1010", foreground="#FCA5A5")
            self.tree.tag_configure("stale",    background="#231600", foreground="#FB923C")
            self.tree.tag_configure("aging",    background="#1E1900", foreground="#FBBF24")
            self.tree.tag_configure("in_prog",  background="#0E1628", foreground="#93C5FD")
            self.tree.tag_configure("high",     background="#1A1028", foreground="#C084FC")
            self.tree.tag_configure("open",     background=CARD,      foreground=TEXT)
            self.tree.tag_configure("resolved", background=CARD,      foreground=MUTED)
        vsb=ttk.Scrollbar(card,orient="vertical",command=self.tree.yview,style="Modern.Vertical.TScrollbar")
        hsb=ttk.Scrollbar(card,orient="horizontal",command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,xscrollcommand=hsb.set)
        vsb.pack(side="right",fill="y",padx=(0,2),pady=4)
        hsb.pack(side="bottom",fill="x",pady=(0,2))
        self.tree.pack(fill="both",expand=True,padx=2,pady=2)
        self.tree.bind("<Double-1>",self._open_detail)
        self.tree.bind("<Control-c>", self._copy_selected_rows)
        self.tree.bind("<Control-C>", self._copy_selected_rows)

        sbar=tk.Frame(lp,bg=NAV2,pady=6); sbar.pack(fill="x",side="bottom")
        tk.Frame(lp,bg=BORDER,height=1).pack(fill="x",side="bottom")
        self.status_lbl=tk.Label(sbar,text="",font=(FONT,9),bg=NAV2,fg=MUTED,anchor="w",padx=16)
        self.status_lbl.pack(side="left")
        make_btn(sbar,"↻ Sync now",self._sync_now,"default",padx=10,pady=4).pack(side="right",padx=(0,8))
        self.sync_lbl=tk.Label(sbar,text="",font=(FONT,8),bg=NAV2,fg=MUTED,anchor="e",padx=16)
        self.sync_lbl.pack(side="right")

        self._build_dashboard_page()
        self._build_calendar_page()
        self._show_page("dashboard",init=True)

    def _show_page(self,page,init=False):
        self.dashboard_page.pack_forget(); self.list_page.pack_forget(); self.calendar_page.pack_forget(); self.reports_page.pack_forget()
        if page=="dashboard":
            self.dashboard_page.pack(fill="both",expand=True)
            if getattr(self,"_dash_dirty",True):
                self._refresh_dashboard()
                self._dash_dirty=False
        elif page=="calendar":
            self.calendar_page.pack(fill="both",expand=True)
            self._refresh_calendar_page()
        elif page=="reports":
            self.reports_page.pack(fill="both",expand=True)
            if getattr(self,"_rpt_dirty",True):
                self._refresh_reports()
                self._rpt_dirty=False
        else:
            self.list_page.pack(fill="both",expand=True)
            if not init:
                self._refresh_table()
        # Watcher runs on ALL pages — drop inbox should work wherever you are
        self._start_watcher()
        for v,btn in self._page_btns.items():
            if v==page:
                btn.configure(fg=ACCENT2,font=(FONT,10,"bold"),bg=NAV2,
                              highlightthickness=0)
                # Underline effect via border bottom trick — pad bottom more
                btn.configure(pady=9)
            else:
                btn.configure(fg=NAV_MUTED,font=(FONT,10),bg=NAV2,pady=11,
                              highlightthickness=0)

    def _section_lbl(self, parent, text):
        f=tk.Frame(parent,bg=BG); f.pack(fill="x",pady=(20,10))
        tk.Label(f,text=text,font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(side="left")
        tk.Frame(f,bg=BORDER,height=1).pack(side="left",fill="x",expand=True,padx=(10,0),pady=6)

    def _build_dashboard_page(self):
        dp=self.dashboard_page
        self._dash_inner,self._dash_canvas_obj=scrollable_frame(dp)
        self._rpt_inner,_=scrollable_frame(self.reports_page)

    def _build_calendar_page(self):
        cp=self.calendar_page
        body,self.cal_canvas=scrollable_frame(cp)
        self.cal_inner=body
        hdr=tk.Frame(body,bg=BG); hdr.pack(fill="x",padx=20,pady=(18,6))
        tk.Label(hdr,text="Calendar Workload",font=(FONT,15,"bold"),bg=BG,fg=TEXT).pack(anchor="center")

        self.cal_focus_date=date.today()
        self.cal_view_var=tk.StringVar(value="month")
        self.cal_member_var=tk.StringVar(value="All")

        ctrl=tk.Frame(body,bg=BG); ctrl.pack(fill="x",padx=20,pady=(0,8))
        nav=tk.Frame(ctrl,bg=BG); nav.pack(fill="x")
        make_btn(nav,"◀",lambda:self._shift_calendar_period(-1),"default",padx=12,pady=5).pack(side="left")
        self.cal_title_lbl=tk.Label(nav,text="",font=(FONT,18,"bold"),bg=BG,fg=TEXT)
        self.cal_title_lbl.pack(side="left",expand=True)
        make_btn(nav,"▶",lambda:self._shift_calendar_period(1),"default",padx=12,pady=5).pack(side="right")

        filters=tk.Frame(body,bg=BG); filters.pack(fill="x",padx=20,pady=(0,8))
        tk.Label(filters,text="View",font=(FONT,9),bg=BG,fg=MUTED).pack(side="left",padx=(0,8))
        for txt,val in [("Month","month"),("Week","week")]:
            rb=tk.Radiobutton(filters,text=txt,variable=self.cal_view_var,value=val,
                              font=(FONT,9),bg=BG,fg=TEXT,selectcolor=CARD,activebackground=BG,
                              command=self._refresh_calendar_page)
            rb.pack(side="left",padx=(0,8))

        tk.Frame(filters,bg=BORDER,width=1,height=24).pack(side="left",padx=(4,12))
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        tk.Label(filters,text="Team member",font=(FONT,9),bg=BG,fg=MUTED).pack(side="left",padx=(0,6))
        self.cal_member_cb=make_combo(filters,self.cal_member_var,["All"]+all_members,readonly=True,width=20)
        self.cal_member_cb.pack(side="left")
        self.cal_member_cb.bind("<<ComboboxSelected>>",lambda _:self._refresh_calendar_page())
        make_btn(filters,"Availability",self._open_out_of_office_dialog,"default",padx=12,pady=5).pack(side="right")

        self.cal_hint_lbl=tk.Label(body,text="",font=(FONT,8),bg=BG,fg=MUTED,justify="left")
        self.cal_hint_lbl.pack(anchor="w",padx=20,pady=(0,8))

        self.cal_grid=tk.Frame(body,bg=BG)
        self.cal_grid.pack(fill="both",expand=True,padx=20,pady=(0,14))

    def _normalize_out_of_office(self, entries):
        normalized=[]
        seen=set()
        for item in entries or []:
            if not isinstance(item, dict):
                continue
            member=str(item.get("member","")).strip()
            day=str(item.get("date","")).strip()
            kind=str(item.get("type","Out of office")).strip() or "Out of office"
            note=str(item.get("note","")).strip()
            if not member or not parse_iso_date(day):
                continue
            key=(member,day,kind,note)
            if key in seen:
                continue
            seen.add(key)
            normalized.append({"member":member,"date":day,"type":kind,"note":note})
        normalized.sort(key=lambda x:(x["date"],x["member"],x["type"],x["note"]))
        return normalized

    def _save_out_of_office(self):
        self.out_of_office=self._normalize_out_of_office(self.out_of_office)
        self.cfg["out_of_office"]=self.out_of_office
        save_config(self.cfg)

    def _open_out_of_office_dialog(self):
        dlg=tk.Toplevel(self)
        dlg.title("Team availability")
        dlg.configure(bg=BG)
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(True,True)
        cw=min(860,max(700,self.winfo_screenwidth()-120))
        ch=min(640,max(500,self.winfo_screenheight()-160))
        px=max(0,min(self.winfo_rootx()+40,self.winfo_screenwidth()-cw-20))
        py=max(0,min(self.winfo_rooty()+40,self.winfo_screenheight()-ch-60))
        dlg.geometry(f"{cw}x{ch}+{px}+{py}")
        dlg.minsize(700,500)

        hdr=tk.Frame(dlg,bg=NAV,padx=20,pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr,text="Team availability",font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
        tk.Label(hdr,text="Mark a team member as unavailable for one day or a full date range, for example out of office or site visit.",font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
        tk.Frame(dlg,bg=ACCENT,height=2).pack(fill="x")

        body=tk.Frame(dlg,bg=BG,padx=20,pady=16)
        body.pack(fill="both",expand=True)
        body.grid_columnconfigure(0,weight=1)
        body.grid_rowconfigure(1,weight=1)

        all_members=list(dict.fromkeys([self.username]+self.team_members))
        member_var=tk.StringVar(value=all_members[0] if all_members else self.username)
        date_from_var=tk.StringVar(value=today_str())
        date_to_var=tk.StringVar(value=today_str())
        type_var=tk.StringVar(value="Out of office")
        note_var=tk.StringVar()

        form_card=tk.Frame(body,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=14,pady=14)
        form_card.grid(row=0,column=0,sticky="ew",pady=(0,12))
        form_card.grid_columnconfigure(0,weight=1)
        form_card.grid_columnconfigure(1,weight=1)
        form_card.grid_columnconfigure(2,weight=1)

        row=tk.Frame(form_card,bg=CARD)
        row.grid(row=0,column=0,columnspan=3,sticky="ew")
        row.grid_columnconfigure(1,weight=1)
        row.grid_columnconfigure(3,weight=1)

        tk.Label(row,text="Team member",font=(FONT,9),bg=CARD,fg=MUTED).grid(row=0,column=0,sticky="w",padx=(0,8))
        member_cb=make_combo(row,member_var,all_members,readonly=True,width=18)
        member_cb.grid(row=0,column=1,sticky="ew",padx=(0,14))
        tk.Label(row,text="Type",font=(FONT,9),bg=CARD,fg=MUTED).grid(row=0,column=2,sticky="w",padx=(0,8))
        type_cb=make_combo(row,type_var,["Out of office","Site visit","Annual leave","Training","Sickness","Other"],readonly=True,width=18)
        type_cb.grid(row=0,column=3,sticky="ew")

        row2=tk.Frame(form_card,bg=CARD)
        row2.grid(row=1,column=0,columnspan=3,sticky="ew",pady=(12,0))
        row2.grid_columnconfigure(1,weight=1)
        row2.grid_columnconfigure(4,weight=1)
        row2.grid_columnconfigure(7,weight=3)

        tk.Label(row2,text="From",font=(FONT,9),bg=CARD,fg=MUTED).grid(row=0,column=0,sticky="w",padx=(0,8))
        from_wrap=tk.Frame(row2,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        from_wrap.grid(row=0,column=1,sticky="ew")
        from_entry=tk.Entry(from_wrap,textvariable=date_from_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,highlightthickness=0)
        from_entry.pack(fill="x")
        make_btn(row2,"📅",lambda:_show_cal(dlg,date_from_var),"default",padx=7,pady=3).grid(row=0,column=2,sticky="w",padx=(4,14))

        tk.Label(row2,text="To",font=(FONT,9),bg=CARD,fg=MUTED).grid(row=0,column=3,sticky="w",padx=(0,8))
        to_wrap=tk.Frame(row2,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        to_wrap.grid(row=0,column=4,sticky="ew")
        to_entry=tk.Entry(to_wrap,textvariable=date_to_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,highlightthickness=0)
        to_entry.pack(fill="x")
        make_btn(row2,"📅",lambda:_show_cal(dlg,date_to_var),"default",padx=7,pady=3).grid(row=0,column=5,sticky="w",padx=(4,14))

        tk.Label(row2,text="Note",font=(FONT,9),bg=CARD,fg=MUTED).grid(row=0,column=6,sticky="w",padx=(0,8))
        note_wrap=tk.Frame(row2,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        note_wrap.grid(row=0,column=7,sticky="ew")
        note_entry=tk.Entry(note_wrap,textvariable=note_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,highlightthickness=0)
        note_entry.pack(fill="x")

        list_frame=tk.Frame(body,bg=CARD,highlightthickness=1,highlightbackground=BORDER)
        list_frame.grid(row=1,column=0,sticky="nsew")
        cols=("member","date","type","note")
        tree=ttk.Treeview(list_frame,columns=cols,show="headings",height=12,style="Modern.Treeview")
        for col,label,width in [("member","Team member",140),("date","Date",100),("type","Type",120),("note","Note",220)]:
            tree.heading(col,text=label)
            tree.column(col,width=width,anchor="w",stretch=(col=="note"))
        ysb=ttk.Scrollbar(list_frame,orient="vertical",command=tree.yview,style="Modern.Vertical.TScrollbar")
        tree.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right",fill="y")
        tree.pack(fill="both",expand=True,padx=4,pady=4)

        def refresh_entries():
            tree.delete(*tree.get_children())
            for idx,entry in enumerate(self._normalize_out_of_office(self.out_of_office)):
                note=entry.get("note","")
                tree.insert("","end",iid=str(idx),values=(entry["member"],fmt_date(entry["date"]),entry["type"],note or "-"))

        def add_entry():
            member=member_var.get().strip()
            start_day=date_from_var.get().strip()
            end_day=date_to_var.get().strip()
            kind=type_var.get().strip() or "Out of office"
            note=note_var.get().strip()
            if not member:
                messagebox.showwarning("Required","Please select a team member.",parent=dlg)
                return
            start_date=parse_iso_date(start_day)
            end_date=parse_iso_date(end_day)
            if not start_date or not end_date:
                messagebox.showwarning("Invalid date","Please enter valid From and To dates in YYYY-MM-DD format.",parent=dlg)
                return
            if end_date < start_date:
                messagebox.showwarning("Invalid range","The To date must be on or after the From date.",parent=dlg)
                return

            existing_keys={
                (
                    str(entry.get("member","")),
                    str(entry.get("date","")),
                    str(entry.get("type","")),
                    str(entry.get("note","")).strip(),
                )
                for entry in self.out_of_office
            }
            days_added=0
            current_day=start_date
            while current_day <= end_date:
                day=current_day.isoformat()
                key=(member,day,kind,note)
                if key not in existing_keys:
                    self.out_of_office.append({"member":member,"date":day,"type":kind,"note":note})
                    existing_keys.add(key)
                    days_added += 1
                current_day += timedelta(days=1)

            if days_added == 0:
                messagebox.showwarning("Duplicate","That availability range already exists.",parent=dlg)
                return

            self._save_out_of_office()
            note_var.set("")
            refresh_entries()
            self._refresh_calendar_page()
            note_entry.focus_set()

        def remove_selected():
            sel=tree.selection()
            if not sel:
                return
            entries=self._normalize_out_of_office(self.out_of_office)
            for iid in sorted(sel, key=lambda x:int(x), reverse=True):
                idx=int(iid)
                if 0 <= idx < len(entries):
                    target=entries[idx]
                    for existing in list(self.out_of_office):
                        if existing.get("member")==target["member"] and existing.get("date")==target["date"] and existing.get("type")==target["type"] and str(existing.get("note","")).strip()==target.get("note",""):
                            self.out_of_office.remove(existing)
                            break
            self._save_out_of_office()
            refresh_entries()
            self._refresh_calendar_page()

        btns=tk.Frame(body,bg=BG)
        btns.grid(row=2,column=0,sticky="ew",pady=(10,0))
        make_btn(btns,"Add",add_entry,"primary",padx=12,pady=6).pack(side="left")
        make_btn(btns,"Remove selected",remove_selected,"danger",padx=12,pady=6).pack(side="left",padx=(8,0))
        make_btn(btns,"Close",dlg.destroy,"default",padx=12,pady=6).pack(side="right")

        for entry_widget in (from_entry,to_entry,note_entry):
            entry_widget.bind("<Return>",lambda e:(add_entry(),"break")[1])
        tree.bind("<Delete>",lambda e:(remove_selected(),"break")[1])
        dlg.bind("<Escape>",lambda e:dlg.destroy())
        refresh_entries()
        note_entry.focus_set()

    def _shift_calendar_period(self,delta):
        if self.cal_view_var.get()=="week":
            self.cal_focus_date=self.cal_focus_date+timedelta(days=7*delta)
        else:
            y,m=self.cal_focus_date.year,self.cal_focus_date.month+delta
            if m<1: m=12; y-=1
            if m>12: m=1; y+=1
            # Keep a safe day number when crossing month boundaries
            d=min(self.cal_focus_date.day,28)
            self.cal_focus_date=date(y,m,d)
        self._refresh_calendar_page()

    def _refresh_calendar_page(self):
        import calendar as _cal
        for w in self.cal_grid.winfo_children(): w.destroy()

        cal_palette={
            "high_bg":"#2E1111","high_fg":TEXT,
            "med_bg":"#2A1E00","med_fg":TEXT,
            "low_bg":CARD,"low_fg":TEXT,
            "out_bg":"#121922","out_fg":MUTED,
            "high_pri":"#FCA5A5","med_pri":"#FDE68A","low_pri":"#86EFAC",
            "ooo":"#93C5FD",
            "today_bg":"#0D2E2A","today_fg":ACCENT2,
        }
        if ACTIVE_THEME=="Mist Light":
            cal_palette.update({
                "high_bg":"#F9D8D8","high_fg":"#4A1818",
                "med_bg":"#FAF0CC","med_fg":"#5C3B0A",
                "low_bg":"#F4F9F4","low_fg":TEXT,
                "out_bg":"#DCE6EA","out_fg":"#4A6470",
                "high_pri":"#B83232","med_pri":"#A05E0A","low_pri":"#196B3A",
                "ooo":"#1A4FBA",
                "today_bg":"#C2E2EA","today_fg":"#155668",
            })

        def ordinal(n):
            if 10 <= (n % 100) <= 20:
                suffix = "th"
            else:
                suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
            return f"{n}{suffix}"

        focus=self.cal_focus_date
        view=self.cal_view_var.get()
        if view=="week":
            week_start=focus-timedelta(days=focus.weekday())
            week_end=week_start+timedelta(days=6)
            self.cal_title_lbl.config(text=f"Week: {fmt_date(week_start.isoformat())} - {fmt_date(week_end.isoformat())}")
        else:
            month=focus.replace(day=1)
            self.cal_title_lbl.config(text=month.strftime("%B %Y"))

        member_filter=self.cal_member_var.get().strip()
        if member_filter and member_filter!="All":
            self.cal_hint_lbl.config(text=f"Showing workload for: {member_filter}. Availability markers show when they are out of office or on a site visit. Click any day to open Query List filtered to that action date.")
        else:
            self.cal_hint_lbl.config(text="Showing all open queries. Each day shows total, priority levels, team breakdown, and unavailable team members. Click any day to jump to Query List.")

        for c,day_name in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
            tk.Label(self.cal_grid,text=day_name,font=(FONT,9,"bold"),bg=BG,fg=MUTED).grid(row=0,column=c,sticky="nsew",padx=4,pady=(0,6))
            self.cal_grid.columnconfigure(c,weight=1)
        if view=="week":
            max_rows=1
        else:
            month=focus.replace(day=1)
            first_wd,days_in_month=_cal.monthrange(month.year,month.month)
            max_rows=max(4,(first_wd+days_in_month+6)//7)
        for r in range(1,max_rows+1):
            self.cal_grid.rowconfigure(r,weight=1,minsize=150 if view=="week" else 108)

        today=today_str()
        threshold=self._get_high_volume_threshold()

        def queries_for_day(day_date):
            out=[]
            for q in self.queries:
                if q.get("status")=="Resolved":
                    continue
                d=parse_iso_date(q.get("chase_date",""))
                if not d or d!=day_date:
                    continue
                if member_filter!="All" and q.get("assigned_to","")!=member_filter:
                    continue
                out.append(q)
            return out

        def out_of_office_for_day(day_str):
            items=[]
            for entry in getattr(self,"out_of_office",[]):
                if entry.get("date")!=day_str:
                    continue
                if member_filter!="All" and entry.get("member")!=member_filter:
                    continue
                items.append(entry)
            items.sort(key=lambda x:(x.get("member",""),x.get("type",""),x.get("note","")))
            return items

        if view=="week":
            day_cells=[week_start+timedelta(days=i) for i in range(7)]
        else:
            month=focus.replace(day=1)
            first_wd,days_in_month=_cal.monthrange(month.year,month.month)
            total_slots=max_rows*7
            first_cell=month-timedelta(days=first_wd)
            day_cells=[first_cell+timedelta(days=i) for i in range(total_slots)]

        for idx,day_date in enumerate(day_cells):
            row=(idx//7)+1
            col=idx%7
            qs=queries_for_day(day_date)
            total=len(qs)
            day_str=day_date.isoformat()
            ooo_items=out_of_office_for_day(day_str)
            in_focus_month=(day_date.month==focus.month) if view=="month" else True

            if total>=threshold:
                bg=cal_palette["high_bg"]; fg=cal_palette["high_fg"]
            elif total>=max(1,threshold//2):
                bg=cal_palette["med_bg"]; fg=cal_palette["med_fg"]
            else:
                bg=cal_palette["low_bg"]; fg=cal_palette["low_fg"]
            if not in_focus_month:
                bg=cal_palette["out_bg"]; fg=cal_palette["out_fg"]
            if day_str<today and total>0:
                border_col=DANGER
            elif day_str==today:
                border_col=ACCENT2
            elif ooo_items:
                border_col="#60A5FA"
            else:
                border_col=BORDER

            cell=tk.Frame(self.cal_grid,bg=bg,highlightthickness=1,highlightbackground=border_col,padx=6,pady=5,cursor="hand2")
            cell.grid(row=row,column=col,sticky="nsew",padx=3,pady=3)
            top=tk.Frame(cell,bg=bg); top.pack(fill="x",anchor="nw")
            tk.Label(top,text=ordinal(day_date.day),font=(FONT,12,"bold"),bg=bg,fg=fg).pack(side="left",anchor="nw")
            if day_str==today:
                tk.Label(top,text="Today",font=(FONT,7,"bold"),bg=cal_palette["today_bg"],fg=cal_palette["today_fg"],padx=4,pady=1).pack(side="right")

            content=tk.Frame(cell,bg=bg)
            content.pack(fill="x",pady=(2,0))
            left_box=tk.Frame(content,bg=bg)
            left_box.pack(side="left",fill="x",expand=True,anchor="nw")
            tk.Label(left_box,text=f"Total: {total}",font=(FONT,9,"bold"),bg=bg,fg=ACCENT2 if total else MUTED).pack(anchor="nw")

            h=sum(1 for q in qs if q.get("priority")=="High")
            m=sum(1 for q in qs if q.get("priority")=="Medium")
            l=sum(1 for q in qs if q.get("priority")=="Low")
            prio_box=tk.Frame(content,bg=bg)
            prio_box.pack(side="right",anchor="ne",padx=(6,0))
            tk.Label(prio_box,text=f"High: {h}",font=(FONT,8),bg=bg,fg=cal_palette["high_pri"] if h else TEXT2,anchor="e").pack(anchor="ne")
            tk.Label(prio_box,text=f"Medium: {m}",font=(FONT,8),bg=bg,fg=cal_palette["med_pri"] if m else TEXT2,anchor="e").pack(anchor="ne")
            tk.Label(prio_box,text=f"Low: {l}",font=(FONT,8),bg=bg,fg=cal_palette["low_pri"] if l else TEXT2,anchor="e").pack(anchor="ne")

            if member_filter=="All" and qs:
                counts={}
                for q in qs:
                    k=q.get("assigned_to","") or "Unassigned"
                    counts[k]=counts.get(k,0)+1
                top=sorted(counts.items(),key=lambda kv:(-kv[1],kv[0]))[:2]
                members_line=" | ".join(f"{member[:8]}:{count}" for member,count in top)
                if members_line:
                    tk.Label(cell,text=members_line,font=(FONT,7),bg=bg,fg=MUTED,anchor="w",justify="left").pack(anchor="nw",pady=(2,0))

            if ooo_items:
                if member_filter!="All":
                    summary=", ".join(sorted({item.get("type","Out of office") for item in ooo_items}))
                    tk.Label(cell,text=f"Unavailable: {summary}",font=(FONT,7,"bold"),bg=bg,fg=cal_palette["ooo"],anchor="w",justify="left",wraplength=160).pack(anchor="nw",pady=(2,0))
                else:
                    names=", ".join(sorted({item.get("member","") for item in ooo_items if item.get("member","")}))
                    short_names=(names[:24]+"...") if len(names)>24 else names
                    tk.Label(cell,text=f"OOO: {short_names}",font=(FONT,7),bg=bg,fg=cal_palette["ooo"],anchor="w",justify="left",wraplength=160).pack(anchor="nw",pady=(2,0))

            def open_day(e,d=day_str):
                self._open_day_in_list(d)

            def hover_on(e,c=cell):
                c.configure(highlightbackground=ACCENT)

            def hover_off(e,c=cell,bc=border_col):
                c.configure(highlightbackground=bc)

            cell.bind("<Button-1>",open_day)
            cell.bind("<Enter>",hover_on)
            cell.bind("<Leave>",hover_off)
            def bind_descendants(widget):
                for ch in widget.winfo_children():
                    ch.bind("<Button-1>",open_day)
                    ch.bind("<Enter>",hover_on)
                    ch.bind("<Leave>",hover_off)
                    bind_descendants(ch)
            bind_descendants(cell)

        try:
            self.cal_canvas.yview_moveto(0)
        except:
            pass

    def _refresh_dashboard(self):
        for w in self._dash_inner.winfo_children(): w.destroy()
        dp=self._dash_inner; today=today_str()
        
        # Apply assignee filter if set
        af = getattr(self, "_assignee_filter", "")
        filtered_queries = [q for q in self.queries if not af or q.get("assigned_to", "") == af]
        
        summary=tk.Frame(dp,bg=BG); summary.pack(fill="x",pady=(4,0))
        open_n=sum(1 for q in filtered_queries if q["status"]!="Resolved")
        chase_n=sum(1 for q in filtered_queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today)

        # SLA metric — queries with raised_date set where intake was >1 working day
        def intake_d(q):
            try:
                rd=q.get("raised_date",""); od=q.get("opened","")
                if not rd or not od:
                    return None
                return sla_intake_working_days(rd, od)
            except:
                return None

        qs_with_raised=[q for q in filtered_queries if q.get("raised_date","")]
        sla_breached=sum(1 for q in qs_with_raised if (intake_d(q) or 0)>1)
        sla_met=sum(1 for q in qs_with_raised if (intake_d(q) or 0)<=1)
        sla_pct=int(sla_met/len(qs_with_raised)*100) if qs_with_raised else None

        for label,val,color,action in [
            ("Total queries", len(filtered_queries),                                         ACCENT,  lambda:self._go_list("all")),
            ("Open",          open_n,                                                    ACCENT2, lambda:self._go_list("open")),
            ("Action today",  chase_n,                                                   DANGER if chase_n else MUTED, lambda:self._go_list("action")),
            ("Resolved",      sum(1 for q in filtered_queries if q["status"]=="Resolved"),   SUCCESS, lambda:self._go_list("resolved"))]:
            outer=tk.Frame(summary,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,cursor="hand2")
            outer.pack(side="left",padx=(0,10))
            tk.Frame(outer,bg=color,height=3).pack(fill="x")
            inner=tk.Frame(outer,bg=CARD2,padx=20,pady=16); inner.pack(fill="both",expand=True)
            # Number always in ACCENT colour (teal) so readable on dark card — not red/danger
            num_col=ACCENT if color==DANGER else color
            tk.Label(inner,text=str(val),font=(FONT,26,"bold"),bg=CARD2,fg=num_col).pack(anchor="w")
            tk.Label(inner,text=label.upper(),font=(FONT,7,"bold"),bg=CARD2,fg=MUTED).pack(anchor="w",pady=(3,0))
            def bind_card(c=outer,a=action,col=color):
                def _enter(e): c.configure(highlightbackground=col)
                def _leave(e): c.configure(highlightbackground=BORDER)
                for w in [c]+list(c.winfo_children())+[w2 for ch in c.winfo_children() for w2 in ch.winfo_children()]:
                    try: w.bind("<Button-1>",lambda e,aa=a:aa()); w.bind("<Enter>",_enter); w.bind("<Leave>",_leave)
                    except: pass
            bind_card()

            bind_card()

        # SLA card — only show when there are queries with raised_date set
        if qs_with_raised:
            sla_col=SUCCESS if sla_pct>=80 else WARNING if sla_pct>=50 else DANGER
            sla_outer=tk.Frame(summary,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
            sla_outer.pack(side="left",padx=(0,10))
            tk.Frame(sla_outer,bg=sla_col,height=3).pack(fill="x")
            sla_inner=tk.Frame(sla_outer,bg=CARD2,padx=20,pady=16); sla_inner.pack(fill="both",expand=True)
            tk.Label(sla_inner,text=f"{sla_pct}%",font=(FONT,26,"bold"),bg=CARD2,fg=sla_col).pack(anchor="w")
            tk.Label(sla_inner,text="1-WORKING-DAY SLA MET",font=(FONT,7,"bold"),bg=CARD2,fg=MUTED).pack(anchor="w",pady=(3,0))
            if sla_breached:
                tk.Label(sla_inner,text=f"{sla_breached} breached",font=(FONT,7),
                         bg=CARD2,fg=DANGER).pack(anchor="w")

        # Escalation actions
        if self.cfg.get("escalation_rules"):
            esc_actions = tk.Frame(summary, bg=BG)
            esc_actions.pack(side="left", padx=(0, 10))
            def run_escalation():
                escalated = self._apply_escalation_rules()
                self._refresh_dashboard()
                self._refresh_table()
                msg = f"Escalation complete. {escalated} quer{'y' if escalated == 1 else 'ies'} escalated." if escalated > 0 else "Escalation complete. No queries needed escalation."
                messagebox.showinfo("Escalation Complete", msg, parent=self)
            make_btn(esc_actions, "⚡ Run Escalation", run_escalation, "warning", padx=12, pady=8).pack()

        recent_qs=[q for q in filtered_queries if q["id"] in self.recent_ids]
        recent_qs.sort(key=lambda q:self.recent_ids.index(q["id"]) if q["id"] in self.recent_ids else 99)

        # ── Team member boxes ─────────────────────────────────────────────────
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        if len(all_members)>0:
            self._section_lbl(dp,"TEAM")
            tm_row=tk.Frame(dp,bg=BG); tm_row.pack(fill="x",pady=(0,4))
            af=getattr(self,"_assignee_filter","")

            def set_person_filter(person):
                if getattr(self,"_assignee_filter","")==person:
                    self._assignee_filter=""  # toggle off — stay on dashboard
                    try: self.filter_assignee.set("All")
                    except: pass
                    try: self.cal_member_var.set("All")
                    except: pass
                    self._dash_dirty=True
                    self._refresh_dashboard()
                    self._refresh_table()
                else:
                    self._assignee_filter=person
                    try: self.filter_assignee.set(person)
                    except: pass
                    try: self.cal_member_var.set(person)
                    except: pass
                    # Stay on dashboard and filter there instead of going to list
                    self._dash_dirty=True
                    self._refresh_dashboard()
                    self._refresh_table()

            for person in all_members:
                pqs=[q for q in filtered_queries if q.get("assigned_to","")==person]
                p_open=sum(1 for q in pqs if q["status"]!="Resolved")
                p_action=sum(1 for q in pqs if q["status"]!="Resolved"
                             and q.get("chase_date","") and q["chase_date"]<=today)
                is_active=(af==person)
                outer=tk.Frame(tm_row,bg=ACCENT if is_active else CARD,
                               highlightthickness=1,
                               highlightbackground=ACCENT if is_active else BORDER,
                               cursor="hand2")
                outer.pack(side="left",padx=(0,10))
                tk.Frame(outer,bg=ACCENT2 if is_active else MUTED,width=3).pack(side="left",fill="y")
                inner_p=tk.Frame(outer,bg=ACCENT if is_active else CARD,padx=14,pady=10)
                inner_p.pack(side="left")
                name_col=NAV if is_active else TEXT
                tk.Label(inner_p,text=person,font=(FONT,9,"bold"),
                         bg=ACCENT if is_active else CARD,fg=name_col).pack(anchor="w")
                stats=tk.Frame(inner_p,bg=ACCENT if is_active else CARD); stats.pack(anchor="w",pady=(4,0))
                for lbl,val,col in [("open",p_open,ACCENT2),("action",p_action,DANGER if p_action else MUTED)]:
                    sf=tk.Frame(stats,bg=ACCENT if is_active else CARD)
                    sf.pack(side="left",padx=(0,12))
                    tk.Label(sf,text=str(val),font=(FONT,13,"bold"),
                             bg=ACCENT if is_active else CARD,
                             fg=(DANGER if col==DANGER and val>0 else ACCENT2 if is_active else col) if val>0 else MUTED
                             ).pack(anchor="w")
                    tk.Label(sf,text=lbl,font=(FONT,7),
                             bg=ACCENT if is_active else CARD,
                             fg=NAV if is_active else MUTED).pack(anchor="w")
                def bind_person(o=outer,p=person):
                    def _click(e): set_person_filter(p)
                    def _enter(e): o.configure(highlightbackground=ACCENT2)
                    def _leave(e): o.configure(highlightbackground=ACCENT if af==p else BORDER)
                    for w in [o]+list(o.winfo_children())+[w2 for ch in o.winfo_children() for w2 in ch.winfo_children()]+\
                              [w3 for ch in o.winfo_children() for w2 in ch.winfo_children() for w3 in (w2.winfo_children() if hasattr(w2,'winfo_children') else [])]:
                        try: w.bind("<Button-1>",_click); w.bind("<Enter>",_enter); w.bind("<Leave>",_leave)
                        except: pass
                bind_person()

            # Clear filter button if one is active
            if af:
                def clear_pf():
                    self._assignee_filter=""
                    try: self.filter_assignee.set("All")
                    except: pass
                    try: self.cal_member_var.set("All")
                    except: pass
                    self._refresh_dashboard(); self._refresh_table()
                make_btn(tm_row,"✕ Clear person filter",clear_pf,"danger",padx=10,pady=6).pack(side="left",padx=(4,0))

        # ── Workload Balancing ────────────────────────────────────────────────
        all_members = list(dict.fromkeys([self.username] + self.team_members))
        if len(all_members) > 1:  # Only show if there are multiple team members
            self._section_lbl(dp, "WORKLOAD BALANCE")
            wl_frame = tk.Frame(dp, bg=BG)
            wl_frame.pack(fill="x", pady=(0, 4))
            
            # Calculate workload stats
            member_stats = []
            total_open = 0
            total_action = 0
            
            for person in all_members:
                pqs = [q for q in filtered_queries if q.get("assigned_to", "") == person]
                p_open = sum(1 for q in pqs if q["status"] != "Resolved")
                p_action = sum(1 for q in pqs if q["status"] != "Resolved" 
                              and q.get("chase_date", "") and q["chase_date"] <= today)
                member_stats.append((person, p_open, p_action))
                total_open += p_open
                total_action += p_action
            
            # Calculate averages
            avg_open = total_open / len(all_members) if all_members else 0
            avg_action = total_action / len(all_members) if all_members else 0
            
            # Show workload distribution
            for person, p_open, p_action in member_stats:
                workload_level = "low" if p_open < avg_open * 0.8 else "high" if p_open > avg_open * 1.2 else "balanced"
                bg_color = SUCCESS if workload_level == "balanced" else WARNING if workload_level == "high" else MUTED
                
                wl_card = tk.Frame(wl_frame, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
                wl_card.pack(side="left", padx=(0, 10))
                tk.Frame(wl_card, bg=bg_color, width=3).pack(side="left", fill="y")
                wl_inner = tk.Frame(wl_card, bg=CARD, padx=14, pady=10)
                wl_inner.pack(side="left")
                
                tk.Label(wl_inner, text=person, font=(FONT, 9, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")
                wl_stats = tk.Frame(wl_inner, bg=CARD)
                wl_stats.pack(anchor="w", pady=(4, 0))
                
                for lbl, val, col in [("open", p_open, ACCENT2), ("action", p_action, DANGER if p_action else MUTED)]:
                    wl_sf = tk.Frame(wl_stats, bg=CARD)
                    wl_sf.pack(side="left", padx=(0, 12))
                    tk.Label(wl_sf, text=str(val), font=(FONT, 13, "bold"), bg=CARD, 
                            fg=(DANGER if col == DANGER and val > 0 else col) if val > 0 else MUTED).pack(anchor="w")
                    tk.Label(wl_sf, text=lbl, font=(FONT, 7), bg=CARD, fg=MUTED).pack(anchor="w")

        if recent_qs:
            self._section_lbl(dp,"RECENTLY OPENED")
            rrow=tk.Frame(dp,bg=BG); rrow.pack(fill="x")
            for q in recent_qs[:5]:
                sc,spill=S_COLORS.get(q["status"],("#94A3B8","#1E2D45"))
                rc=tk.Frame(rrow,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=14,pady=12,cursor="hand2")
                rc.pack(side="left",padx=(0,10))
                tk.Label(rc,text=q["ref"],font=(FONT,10,"bold"),bg=CARD,fg=TEXT).pack(anchor="w")
                tk.Label(rc,text=q["client"],font=(FONT,8),bg=CARD,fg=TEXT2).pack(anchor="w",pady=(2,0))
                # Status pill
                pill=tk.Frame(rc,bg=spill); pill.pack(anchor="w",pady=(6,0))
                tk.Label(pill,text=f"  {q['status']}  ",font=(FONT,8,"bold"),bg=spill,fg=sc,pady=2).pack()
                def open_q(e,qid=q["id"]):
                    found=next((x for x in self.queries if x["id"]==qid),None)
                    if found: self._show_page("list"); self._open_detail_query(found)
                for w in [rc]+list(rc.winfo_children())+[pill.winfo_children()[0]]:
                    try:
                        w.bind("<Button-1>",open_q)
                        w.bind("<Enter>",lambda e,c=rc:c.configure(highlightbackground=ACCENT))
                        w.bind("<Leave>",lambda e,c=rc:c.configure(highlightbackground=BORDER))
                    except: pass

        self._section_lbl(dp,"QUERIES BY TYPE")
        grid=tk.Frame(dp,bg=BG); grid.pack(fill="x")
        for i,qt in enumerate(QUERY_TYPES):
            total=sum(1 for q in filtered_queries if q["type"]==qt)
            open_=sum(1 for q in filtered_queries if q["type"]==qt and q["status"]!="Resolved")
            chase=sum(1 for q in filtered_queries if q["type"]==qt and q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today)
            row=i//3; col=i%3
            card=tk.Frame(grid,bg=CARD,highlightthickness=1,
                          highlightbackground="#3B1111" if chase>0 else BORDER,
                          padx=16,pady=14,cursor="hand2")
            card.grid(row=row,column=col,padx=(0,10),pady=(0,10),sticky="ew")
            grid.columnconfigure(col,weight=1)
            top=tk.Frame(card,bg=CARD); top.pack(fill="x")
            tk.Label(top,text=qt,font=(FONT,10,"bold"),bg=CARD,fg=TEXT).pack(side="left")
            # Total badge
            badge_bg="#0D2E2A" if open_>0 else CARD2
            badge_fg=ACCENT2 if open_>0 else MUTED
            tk.Label(top,text=f" {total} ",font=(FONT,8,"bold"),bg=badge_bg,fg=badge_fg,padx=5,pady=2).pack(side="right")
            # Small action indicator — text not a block
            if chase>0:
                tk.Label(top,text=f"⚠ {chase}",font=(FONT,8,"bold"),bg=CARD,fg=DANGER,padx=2).pack(side="right")
            # Mini stats row
            sub=tk.Frame(card,bg=CARD); sub.pack(fill="x",pady=(10,0))
            for lbl,n,col2 in [("open",open_,ACCENT2),("action",chase,DANGER)]:
                sf=tk.Frame(sub,bg=CARD); sf.pack(side="left",padx=(0,18))
                tk.Label(sf,text=str(n),font=(FONT,16,"bold"),bg=CARD,
                         fg=col2 if n>0 else BORDER).pack(anchor="w")
                tk.Label(sf,text=lbl,font=(FONT,7),bg=CARD,fg=MUTED).pack(anchor="w")
            ql=tk.Label(card,text="＋ log",font=(FONT,8),bg=CARD,fg=MUTED,cursor="hand2")
            ql.pack(anchor="e",pady=(6,0))
            ql.bind("<Button-1>",lambda e,qt=qt:self._open_add_dialog(preset_type=qt))
            def on_enter(e,c=card,ch=chase): c.configure(highlightbackground=DANGER if ch>0 else ACCENT)
            def on_leave(e,c=card,ch=chase): c.configure(highlightbackground="#3B1111" if ch>0 else BORDER)
            def on_click(e,qt=qt):
                self._show_page("list"); self.filter_type.set(qt); self._set_tab("all")
            all_w=[card,top,sub]+list(top.winfo_children())+list(sub.winfo_children())
            for sf in sub.winfo_children(): all_w+=list(sf.winfo_children())
            for w in all_w:
                try: w.bind("<Enter>",on_enter); w.bind("<Leave>",on_leave); w.bind("<Button-1>",on_click)
                except: pass

        self._section_lbl(dp,"BY STATUS")
        srow=tk.Frame(dp,bg=BG); srow.pack(fill="x",pady=(0,24))
        for status in STATUSES:
            n=sum(1 for q in filtered_queries if q["status"]==status)
            sc,spill=S_COLORS.get(status,("#94A3B8","#1E2D45"))
            outer=tk.Frame(srow,bg=CARD,highlightthickness=1,highlightbackground=BORDER,cursor="hand2")
            outer.pack(side="left",padx=(0,10))
            tk.Frame(outer,bg=sc,width=3).pack(side="left",fill="y")
            inner=tk.Frame(outer,bg=CARD,padx=18,pady=14); inner.pack(side="left")
            tk.Label(inner,text=str(n),font=(FONT,20,"bold"),bg=CARD,fg=sc if n>0 else MUTED).pack(anchor="w")
            tk.Label(inner,text=status,font=(FONT,8),bg=CARD,fg=TEXT2).pack(anchor="w",pady=(3,0))
            def on_click_s(e,s=status):
                self._show_page("list"); self.filter_status.set(s); self._set_tab("all")
            def on_enter_s(e,c=outer,col=sc): c.configure(highlightbackground=col)
            def on_leave_s(e,c=outer): c.configure(highlightbackground=BORDER)
            for w in [outer]+list(outer.winfo_children())+list(inner.winfo_children()):
                try: w.bind("<Button-1>",on_click_s); w.bind("<Enter>",on_enter_s); w.bind("<Leave>",on_leave_s)
                except: pass

    def _refresh_reports(self):
        for w in self._rpt_inner.winfo_children(): w.destroy()
        rp=self._rpt_inner

        # ── Persistent filter state ───────────────────────────────────────────
        if not hasattr(self,"_rpt_period"):  self._rpt_period="all"
        if not hasattr(self,"_rpt_client"):  self._rpt_client="All"
        if not hasattr(self,"_rpt_from"):    self._rpt_from=""
        if not hasattr(self,"_rpt_to"):      self._rpt_to=""
        if not hasattr(self,"_rpt_person"):  self._rpt_person="All"

        def apply_filters():
            self._rpt_period  = period_var.get()
            self._rpt_client  = client_var.get()
            self._rpt_from    = from_var.get().strip()
            self._rpt_to      = to_var.get().strip()
            self._refresh_reports()

        def set_period(p):
            self._rpt_period=p
            self._rpt_from=""; self._rpt_to=""
            self._refresh_reports()

        def clear_all():
            self._rpt_period="all"; self._rpt_client="All"
            self._rpt_from=""; self._rpt_to=""; self._rpt_person="All"
            self._refresh_reports()

        # ── Header row ────────────────────────────────────────────────────────
        hrow=tk.Frame(rp,bg=BG); hrow.pack(fill="x",pady=(0,10))
        tk.Label(hrow,text="Reports",font=(FONT,15,"bold"),bg=BG,fg=TEXT).pack(side="left")
        btn_right=tk.Frame(hrow,bg=BG); btn_right.pack(side="right")
        make_btn(btn_right,"Clear filters",clear_all,"default",padx=10,pady=4).pack(side="right",padx=(6,0))
        export_btn=make_btn(btn_right,"↓  Export report",lambda:self._export_report(
            self._rpt_qs, self._rpt_label),"primary",padx=12,pady=4)
        export_btn.pack(side="right")

        # ── Filter bar card ───────────────────────────────────────────────────
        fbar=tk.Frame(rp,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=12)
        fbar.pack(fill="x",pady=(0,4))

        # Row 1 — period pills
        pill_row=tk.Frame(fbar,bg=CARD2); pill_row.pack(fill="x",pady=(0,10))
        tk.Label(pill_row,text="Period",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED,width=8,anchor="w").pack(side="left")
        period_var=tk.StringVar(value=self._rpt_period)
        for lbl,val in [("This week","week"),("Last week","last_week"),
                        ("This month","month"),("Last month","last_month"),("All time","all")]:
            is_active=(self._rpt_period==val and not (self._rpt_from or self._rpt_to))
            b=tk.Label(pill_row,text=lbl,font=(FONT,9,"bold" if is_active else ""),
                       bg=ACCENT if is_active else CARD,
                       fg="#0B1017" if is_active else TEXT2,
                       cursor="hand2",padx=10,pady=4,
                       highlightthickness=1,
                       highlightbackground=ACCENT if is_active else BORDER)
            b.pack(side="left",padx=(0,4))
            def _hi(e,btn=b,act=is_active):
                if not act: btn.configure(bg=BORDER,fg=TEXT)
            def _lo(e,btn=b,act=is_active,obg=ACCENT if is_active else CARD,ofg="#0B1017" if is_active else TEXT2):
                btn.configure(bg=obg,fg=ofg)
            b.bind("<Enter>",_hi); b.bind("<Leave>",_lo)
            b.bind("<Button-1>",lambda e,v=val:set_period(v))

        # Row 2 — custom dates + client + person
        row2=tk.Frame(fbar,bg=CARD2); row2.pack(fill="x")

        # Custom date range
        tk.Label(row2,text="From",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED).pack(side="left",padx=(0,6))
        from_var=tk.StringVar(value=self._rpt_from)
        from_frame=tk.Frame(row2,bg=CARD,highlightthickness=1,
                            highlightbackground=ACCENT if self._rpt_from else BORDER)
        from_frame.pack(side="left")
        from_entry=tk.Entry(from_frame,textvariable=from_var,font=(FONT,9),bg=CARD,fg=TEXT,
                            insertbackground=TEXT,relief="flat",bd=6,width=11,highlightthickness=0)
        from_entry.pack(side="left")
        make_btn(from_frame,"📅",lambda:_show_cal(rp.winfo_toplevel(),from_var),"nav",padx=5,pady=3).pack(side="left")

        tk.Label(row2,text="  To",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED).pack(side="left",padx=(8,6))
        to_var=tk.StringVar(value=self._rpt_to)
        to_frame=tk.Frame(row2,bg=CARD,highlightthickness=1,
                          highlightbackground=ACCENT if self._rpt_to else BORDER)
        to_frame.pack(side="left")
        to_entry=tk.Entry(to_frame,textvariable=to_var,font=(FONT,9),bg=CARD,fg=TEXT,
                          insertbackground=TEXT,relief="flat",bd=6,width=11,highlightthickness=0)
        to_entry.pack(side="left")
        make_btn(to_frame,"📅",lambda:_show_cal(rp.winfo_toplevel(),to_var),"nav",padx=5,pady=3).pack(side="left")

        make_btn(row2,"Apply",apply_filters,"primary",padx=10,pady=4).pack(side="left",padx=(10,0))
        from_entry.bind("<Return>",lambda e:apply_filters())
        to_entry.bind("<Return>",lambda e:apply_filters())

        # Separator
        tk.Frame(row2,bg=BORDER,width=1,height=28).pack(side="left",padx=12)

        # Client filter
        tk.Label(row2,text="Client",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED).pack(side="left",padx=(0,6))
        client_var=tk.StringVar(value=self._rpt_client)
        client_cb=make_combo(row2,client_var,["All"]+self.clients,readonly=True,width=18)
        client_cb.pack(side="left")
        client_cb.bind("<<ComboboxSelected>>",lambda e:apply_filters())

        # Separator
        tk.Frame(row2,bg=BORDER,width=1,height=28).pack(side="left",padx=12)

        # Person filter
        tk.Label(row2,text="Person",font=(FONT,8,"bold"),bg=CARD2,fg=MUTED).pack(side="left",padx=(0,6))
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        person_var=tk.StringVar(value=self._rpt_person)
        person_cb=make_combo(row2,person_var,["All"]+all_members,readonly=True,width=16)
        person_cb.pack(side="left")
        def on_person_select(e=None):
            self._rpt_person=person_var.get(); self._refresh_reports()
        person_cb.bind("<<ComboboxSelected>>",on_person_select)

        # ── Resolve which filter is active ────────────────────────────────────
        today=date.today(); td=today_str()

        # Custom date range takes priority over period pills
        if self._rpt_from or self._rpt_to:
            f_from=self._rpt_from or "0000-01-01"
            f_to  =self._rpt_to   or "9999-12-31"
            in_period=lambda q,a=f_from,b=f_to: a<=q.get("opened","")<=b
            # Build label
            lbl_from=fmt_date(self._rpt_from) if self._rpt_from else "start"
            lbl_to  =fmt_date(self._rpt_to)   if self._rpt_to   else "today"
            period_label=f"Custom range  ({lbl_from} – {lbl_to})"
        elif self._rpt_period=="week":
            monday=today-timedelta(days=today.weekday())
            in_period=lambda q: q.get("opened","")>=monday.isoformat()
            period_label=f"This week  ({monday.strftime('%d/%m')} – {today.strftime('%d/%m/%Y')})"
        elif self._rpt_period=="last_week":
            this_mon=today-timedelta(days=today.weekday())
            last_mon=this_mon-timedelta(days=7); last_sun=this_mon-timedelta(days=1)
            in_period=lambda q: last_mon.isoformat()<=q.get("opened","")<=last_sun.isoformat()
            period_label=f"Last week  ({last_mon.strftime('%d/%m')} – {last_sun.strftime('%d/%m/%Y')})"
        elif self._rpt_period=="month":
            first=date(today.year,today.month,1)
            in_period=lambda q: q.get("opened","")>=first.isoformat()
            period_label=f"This month  ({first.strftime('%B %Y')})"
        elif self._rpt_period=="last_month":
            if today.month==1: lm_year=today.year-1; lm_month=12
            else: lm_year=today.year; lm_month=today.month-1
            lm_first=date(lm_year,lm_month,1)
            lm_last=date(today.year,today.month,1)-timedelta(days=1)
            in_period=lambda q: lm_first.isoformat()<=q.get("opened","")<=lm_last.isoformat()
            period_label=f"Last month  ({lm_first.strftime('%B %Y')})"
        else:
            in_period=lambda q: True
            period_label="All time"

        # Apply client + person filter on top of period filter
        rpt_client=self._rpt_client
        rpt_person=getattr(self,"_rpt_person","All")
        qs=[q for q in self.queries
            if in_period(q)
            and (rpt_client=="All" or q["client"]==rpt_client)
            and (rpt_person=="All" or q.get("assigned_to","")==rpt_person)]
        total_n=len(qs)

        # Active filter summary line
        client_suffix=f"  ·  {rpt_client}" if rpt_client!="All" else ""
        person_suffix=f"  ·  {rpt_person}" if rpt_person!="All" else ""
        self._rpt_qs=qs
        self._rpt_label=f"{period_label}{client_suffix}{person_suffix}"

        # Active filter summary line
        tk.Label(rp,text=f"{total_n} quer{'y' if total_n==1 else 'ies'}  ·  {period_label}{client_suffix}  ·  refreshed {today.strftime('%d/%m/%Y')}",
                 font=(FONT,9),bg=BG,fg=TEXT2).pack(anchor="w",pady=(8,16))

        if total_n==0:
            empty=tk.Frame(rp,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=24,pady=32)
            empty.pack(fill="x",pady=20)
            tk.Label(empty,text="No queries in this period.",font=(FONT,11),bg=CARD,fg=MUTED).pack()
            return

        # Soft pastel chart colours — muted, not neon
        CHART_COLORS=["#7EC8C8","#8FC98F","#E8C07A","#E89090","#B8A0E0",
                      "#7AB8D8","#E8A87A","#D898C8","#A0C878","#90C8C8"]

        # Open vs resolved — clearly distinct: muted amber vs muted teal
        OPEN_COL    = "#C89050"   # warm amber — open
        RESOLVED_COL= "#4A9E8A"   # muted teal — resolved
        DIVIDER_COL = BG          # gap between segments

        def draw_donut(parent, title, data, size=220):
            import math
            card=tk.Frame(parent,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=16)
            card.pack(side="left",padx=(0,12),pady=(0,12),anchor="n")
            tk.Label(card,text=title,font=(FONT,10,"bold"),bg=CARD,fg=TEXT).pack(anchor="w",pady=(0,12))
            t=sum(v for _,v in data) or 1
            cv=tk.Canvas(card,width=size,height=size,bg=CARD,highlightthickness=0); cv.pack()
            cx=cy=size//2; R=size//2-12; r=R-42
            start=90.0
            for i,(lbl,val) in enumerate(data):
                if val==0: continue
                extent=-(val/t)*360
                col=CHART_COLORS[i%len(CHART_COLORS)]
                cv.create_arc(cx-R,cy-R,cx+R,cy+R,start=start,extent=extent,
                              fill=col,outline=CARD,width=3,style="pieslice")
                if abs(extent)>22:
                    mid=math.radians(-(start+extent/2))
                    lx=cx+int((R*0.70)*math.cos(mid)); ly=cy+int((R*0.70)*math.sin(mid))
                    # Dark text on pastels reads much better
                    cv.create_text(lx,ly,text=f"{round(val/t*100)}%",font=(FONT,7,"bold"),fill="#1C2530")
                start+=extent
            # Donut hole
            cv.create_oval(cx-r,cy-r,cx+r,cy+r,fill=CARD,outline=CARD)
            cv.create_text(cx,cy-10,text=str(t),font=(FONT,15,"bold"),fill=TEXT)
            cv.create_text(cx,cy+10,text="total",font=(FONT,8),fill=TEXT2)
            # Legend
            leg=tk.Frame(card,bg=CARD); leg.pack(anchor="w",pady=(10,0))
            for i,(lbl,val) in enumerate(data):
                if val==0: continue
                row=tk.Frame(leg,bg=CARD); row.pack(anchor="w",pady=2)
                dot=tk.Frame(row,bg=CHART_COLORS[i%len(CHART_COLORS)],width=8,height=8)
                dot.pack(side="left",padx=(0,8))
                tk.Label(row,text=f"{lbl}",font=(FONT,8),bg=CARD,fg=TEXT2,width=14,anchor="w").pack(side="left")
                tk.Label(row,text=f"{val}",font=(FONT,8,"bold"),bg=CARD,fg=TEXT).pack(side="left",padx=(0,4))
                tk.Label(row,text=f"({round(val/t*100)}%)",font=(FONT,8),bg=CARD,fg=MUTED).pack(side="left")

        def draw_hbar(parent, title, data, bar_color=None):
            card=tk.Frame(parent,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=18,pady=16)
            card.pack(fill="x",pady=(0,12))
            tk.Label(card,text=title,font=(FONT,10,"bold"),bg=CARD,fg=TEXT).pack(anchor="w",pady=(0,12))
            if not data: return
            max_v=max(v for _,v in data) or 1
            for i,(lbl,val) in enumerate(data):
                row=tk.Frame(card,bg=CARD); row.pack(fill="x",pady=3)
                tk.Label(row,text=lbl,font=(FONT,9),bg=CARD,fg=TEXT2,width=28,anchor="e").pack(side="left",padx=(0,12))
                track=tk.Frame(row,bg=CARD2,height=20); track.pack(side="left",fill="x",expand=True)
                track.pack_propagate(False)
                cv=tk.Canvas(track,bg=CARD2,highlightthickness=0,height=20); cv.pack(fill="both",expand=True)
                col=bar_color or CHART_COLORS[i%len(CHART_COLORS)]
                fp=max(0.02,val/max_v); vv=val
                def redraw(e,c=cv,f=fp,cl=col,v=vv):
                    w=c.winfo_width(); bw=max(4,int(w*f))
                    c.delete("all")
                    c.create_rectangle(0,0,bw,20,fill=cl,outline="")
                    if bw+36<w: c.create_text(bw+8,10,text=str(v),font=(FONT,8,"bold"),fill=TEXT2,anchor="w")
                    else: c.create_text(max(bw-6,2),10,text=str(v),font=(FONT,8,"bold"),fill="#1C2530",anchor="e")
                def redraw_debounced(e,c=cv,f=fp,cl=col,v=vv):
                    if hasattr(c,"_rd_after"): c.after_cancel(c._rd_after)
                    c._rd_after=c.after(150,lambda:redraw(None,c,f,cl,v))
                cv.bind("<Configure>",redraw_debounced)
                track.after(30,lambda c=cv,f=fp,cl=col,v=vv: redraw(None,c,f,cl,v))

        def section_label(text):
            f=tk.Frame(rp,bg=BG); f.pack(fill="x",pady=(18,10))
            tk.Label(f,text=text,font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(side="left")
            tk.Frame(f,bg=BORDER,height=1).pack(side="left",fill="x",expand=True,padx=(10,0),pady=6)

        # ── Row 1: donut charts ───────────────────────────────────────────────
        section_label("BREAKDOWN")
        row1=tk.Frame(rp,bg=BG); row1.pack(fill="x")
        status_data=[(s,sum(1 for q in qs if q["status"]==s)) for s in STATUSES]
        draw_donut(row1,"By status",status_data)
        pri_data=[(p,sum(1 for q in qs if q["priority"]==p)) for p in PRIORITIES]
        draw_donut(row1,"By priority",pri_data)
        utils=sorted({q.get("utility","Unknown") or "Unknown" for q in qs})
        util_data=[(u,sum(1 for q in qs if (q.get("utility","") or "Unknown")==u)) for u in utils if any((qq.get("utility","") or "Unknown")==u for qq in qs)]
        if util_data: draw_donut(row1,"By utility",util_data)

        # ── SLA performance (1 working day intake) ────────────────────────────
        def _intake_d(q):
            try:
                rd=q.get("raised_date",""); od=q.get("opened","")
                if not rd or not od:
                    return None
                return sla_intake_working_days(rd, od)
            except:
                return None

        qs_raised=[q for q in qs if q.get("raised_date","")]
        if qs_raised:
            section_label("1-WORKING-DAY SLA PERFORMANCE (RAISED → LOGGED)")
            sla_card=tk.Frame(rp,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=14)
            sla_card.pack(fill="x",pady=(0,16))

            n_total=len(qs_raised)
            n_met=sum(1 for q in qs_raised if (_intake_d(q) or 0)<=1)
            n_breached=n_total-n_met
            pct=int(n_met/n_total*100) if n_total else 0
            sla_col=SUCCESS if pct>=80 else WARNING if pct>=50 else DANGER

            # Top-line numbers
            toprow=tk.Frame(sla_card,bg=CARD); toprow.pack(fill="x",pady=(0,10))
            for val,lbl,col in [
                (f"{pct}%","SLA met (1 working day)",sla_col),
                (str(n_met),"Within 1 working day",SUCCESS),
                (str(n_breached),"Breached",DANGER if n_breached else MUTED),
                (str(n_total),"Total tracked",MUTED),
            ]:
                stat_f=tk.Frame(toprow,bg=CARD,padx=16,pady=0); stat_f.pack(side="left")
                tk.Label(stat_f,text=val,font=(FONT,22,"bold"),bg=CARD,fg=col).pack(anchor="w")
                tk.Label(stat_f,text=lbl,font=(FONT,8),bg=CARD,fg=TEXT2).pack(anchor="w")
                tk.Frame(toprow,bg=BORDER,width=1,height=40).pack(side="left",padx=(0,0))

            tk.Frame(sla_card,bg=BORDER,height=1).pack(fill="x",pady=(8,10))

            # Per-type SLA breakdown
            tk.Label(sla_card,text="By query type",font=(FONT,8,"bold"),bg=CARD,fg=MUTED).pack(anchor="w",pady=(0,6))
            for qt in QUERY_TYPES:
                qt_qs=[q for q in qs_raised if q["type"]==qt]
                if not qt_qs: continue
                qt_met=sum(1 for q in qt_qs if (_intake_d(q) or 0)<=1)
                qt_pct=int(qt_met/len(qt_qs)*100)
                qt_col=SUCCESS if qt_pct>=80 else WARNING if qt_pct>=50 else DANGER
                row_s=tk.Frame(sla_card,bg=CARD); row_s.pack(fill="x",pady=2)
                tk.Label(row_s,text=qt,font=(FONT,9),bg=CARD,fg=TEXT2,width=28,anchor="w").pack(side="left")
                # Progress bar
                bar_outer=tk.Frame(row_s,bg=CARD2,height=14); bar_outer.pack(side="left",fill="x",expand=True)
                bar_outer.pack_propagate(False)
                bar_c=tk.Canvas(bar_outer,bg=CARD2,highlightthickness=0,height=14); bar_c.pack(fill="both",expand=True)
                def draw_sla_bar(e,c=bar_c,p=qt_pct,col=qt_col):
                    w=c.winfo_width(); bw=max(4,int(w*p/100))
                    c.delete("all"); c.create_rectangle(0,2,bw,12,fill=col,outline="")
                def draw_sla_db(e,c=bar_c,p=qt_pct,col=qt_col):
                    if hasattr(c,"_rd"): c.after_cancel(c._rd)
                    c._rd=c.after(150,lambda:draw_sla_bar(None,c,p,col))
                bar_c.bind("<Configure>",draw_sla_db)
                bar_outer.after(30,lambda c=bar_c,p=qt_pct,col=qt_col:draw_sla_bar(None,c,p,col))
                tk.Label(row_s,text=f"  {qt_pct}%  ({qt_met}/{len(qt_qs)})",
                         font=(FONT,8,"bold"),bg=CARD,fg=qt_col,width=14,anchor="w").pack(side="left")

        # ── Row 2: bar by type ────────────────────────────────────────────────
        section_label("BY QUERY TYPE")
        type_data=sorted([(qt,sum(1 for q in qs if q["type"]==qt)) for qt in QUERY_TYPES if any(q["type"]==qt for q in qs)],key=lambda x:-x[1])
        draw_hbar(rp,"Queries by type",type_data)

        # ── Row 3: bar by client ──────────────────────────────────────────────
        section_label("BY CLIENT")
        client_data=sorted([(c,sum(1 for q in qs if q["client"]==c)) for c in {q["client"] for q in qs}],key=lambda x:-x[1])
        draw_hbar(rp,"Queries by client (top 12)",client_data[:12],"#7EC8C8")

        # ── Row 4: stacked open vs resolved ──────────────────────────────────
        section_label("OPEN vs RESOLVED BY TYPE")
        # Legend row
        leg_row=tk.Frame(rp,bg=BG); leg_row.pack(anchor="w",pady=(0,8))
        for lcol,ltxt in [(OPEN_COL,"  Open  "),(RESOLVED_COL,"  Resolved  ")]:
            tk.Frame(leg_row,bg=lcol,width=12,height=12).pack(side="left",padx=(0,6))
            tk.Label(leg_row,text=ltxt,font=(FONT,8),bg=BG,fg=TEXT2).pack(side="left",padx=(0,16))

        ov_card=tk.Frame(rp,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=14,pady=14)
        ov_card.pack(fill="x",pady=(0,16))
        for qt in QUERY_TYPES:
            on=sum(1 for q in qs if q["type"]==qt and q["status"]!="Resolved")
            rn=sum(1 for q in qs if q["type"]==qt and q["status"]=="Resolved")
            tt=on+rn
            if tt==0: continue
            row=tk.Frame(ov_card,bg=CARD); row.pack(fill="x",pady=4)
            tk.Label(row,text=qt,font=(FONT,9),bg=CARD,fg=TEXT2,width=26,anchor="w").pack(side="left")
            track=tk.Frame(row,bg=CARD2,height=20); track.pack(side="left",fill="x",expand=True)
            track.pack_propagate(False)
            cv2=tk.Canvas(track,bg=CARD2,highlightthickness=0,height=20); cv2.pack(fill="both",expand=True)
            def draw_stacked(e=None,c=cv2,o=on,rv=rn,t=tt):
                w=c.winfo_width() or 300
                ow=int((o/t)*w); rw=w-ow
                c.delete("all")
                if ow>0: c.create_rectangle(0,0,ow,20,fill=OPEN_COL,outline="")
                gap=2
                if ow>0 and rw>0: c.create_rectangle(ow,0,ow+gap,20,fill=BG,outline="")
                if rw>gap: c.create_rectangle(ow+gap,0,w,20,fill=RESOLVED_COL,outline="")
                if ow>30: c.create_text(ow//2,10,text=str(o),font=(FONT,7,"bold"),fill="#1C2530")
                if rw>30: c.create_text(ow+gap+(rw-gap)//2,10,text=str(rv),font=(FONT,7,"bold"),fill="#0F1419")
            def draw_stacked_db(e=None,c=cv2,o=on,rv=rn,t=tt):
                if hasattr(c,"_rd_after"): c.after_cancel(c._rd_after)
                c._rd_after=c.after(150,lambda:draw_stacked(None,c,o,rv,t))
            cv2.bind("<Configure>",draw_stacked_db); track.after(20,draw_stacked)
            tk.Label(row,text=f"  {on} / {rn}",font=(FONT,8),bg=CARD,fg=MUTED).pack(side="left",padx=8)

        # ── Row 5: next action summary ────────────────────────────────────────
        section_label("NEXT ACTION DATE SUMMARY")
        overdue_n=sum(1 for q in qs if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<td)
        due_today=sum(1 for q in qs if q["status"]!="Resolved" and q.get("chase_date","")==td)
        due_7d=sum(1 for q in qs if q["status"]!="Resolved" and q.get("chase_date","")>td and q["chase_date"]<=(today+timedelta(days=7)).isoformat())
        no_date=sum(1 for q in qs if q["status"]!="Resolved" and not q.get("chase_date",""))
        nad_row=tk.Frame(rp,bg=BG); nad_row.pack(fill="x",pady=(0,24))
        for lbl,val,col in [("Overdue",overdue_n,DANGER),("Due today",due_today,WARNING),("Due in 7 days",due_7d,ACCENT2),("No date set",no_date,MUTED)]:
            c=tk.Frame(nad_row,bg=CARD,highlightthickness=1,
                       highlightbackground=DANGER if lbl=="Overdue" and val>0 else BORDER,padx=18,pady=14)
            c.pack(side="left",padx=(0,12))
            tk.Label(c,text=str(val),font=(FONT,20,"bold"),bg=CARD,fg=col if val>0 else MUTED).pack(anchor="w")
            tk.Label(c,text=lbl,font=(FONT,9),bg=CARD,fg=TEXT2).pack(anchor="w",pady=(2,0))

        # ── Helper: days between two date strings ─────────────────────────────
        def sla_start(q):
            """Use raised_date if set (true SLA clock), otherwise fall back to opened."""
            rd=q.get("raised_date","")
            return rd if rd else q.get("opened","")

        def age_days(q):
            try: return (today-datetime.strptime(sla_start(q),"%Y-%m-%d").date()).days
            except: return 0

        def resolve_days(q):
            try:
                o=datetime.strptime(sla_start(q),"%Y-%m-%d").date()
                r=datetime.strptime(q.get("resolved_date",""),"%Y-%m-%d").date()
                return max(0,(r-o).days)
            except: return None

        # ── Row 6: age of open queries by type ───────────────────────────────
        section_label("AGE OF OPEN QUERIES BY TYPE")

        age_data=[]
        for qt in QUERY_TYPES:
            open_qs=[q for q in qs if q["type"]==qt and q["status"]!="Resolved"]
            if not open_qs: continue
            ages=[age_days(q) for q in open_qs]
            age_data.append((qt, len(open_qs), max(ages), int(sum(ages)/len(ages)), min(ages)))

        if not age_data:
            tk.Label(rp,text="No open queries in this period.",font=(FONT,9),
                     bg=BG,fg=MUTED).pack(anchor="w",pady=(0,12))
        else:
            age_card=tk.Frame(rp,bg=CARD,highlightthickness=1,highlightbackground=BORDER,
                              padx=16,pady=14); age_card.pack(fill="x",pady=(0,12))
            # Column headers
            hdr_row=tk.Frame(age_card,bg=CARD); hdr_row.pack(fill="x",pady=(0,6))
            for txt,w,anc in [("Query type",24,"w"),("Open",6,"center"),
                               ("Oldest (days)",11,"center"),("Avg age (days)",12,"center"),
                               ("Newest (days)",12,"center"),("Age bar (oldest)",0,"w")]:
                tk.Label(hdr_row,text=txt,font=(FONT,8,"bold"),bg=CARD,fg=MUTED,
                         width=w,anchor=anc).pack(side="left",padx=(0,4))
            tk.Frame(age_card,bg=BORDER,height=1).pack(fill="x",pady=(4,8))

            max_oldest=max(d[2] for d in age_data) or 1
            for qt,count,oldest,avg,newest in sorted(age_data,key=lambda x:-x[2]):
                age_col=(DANGER if oldest>30 else WARNING if oldest>14 else ACCENT2)
                row=tk.Frame(age_card,bg=CARD); row.pack(fill="x",pady=3)
                tk.Label(row,text=qt,    font=(FONT,9),bg=CARD,fg=TEXT2, width=24,anchor="w").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(count),  font=(FONT,9,"bold"),bg=CARD,fg=TEXT,   width=6, anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(oldest), font=(FONT,9,"bold"),bg=CARD,fg=age_col,width=11,anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(avg),    font=(FONT,9),bg=CARD,fg=TEXT2,         width=12,anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(newest), font=(FONT,9),bg=CARD,fg=TEXT2,         width=12,anchor="center").pack(side="left",padx=(0,4))
                # Mini age bar
                bar_outer=tk.Frame(row,bg=CARD2,height=14); bar_outer.pack(side="left",fill="x",expand=True)
                bar_outer.pack_propagate(False)
                bar_cv=tk.Canvas(bar_outer,bg=CARD2,highlightthickness=0,height=14)
                bar_cv.pack(fill="both",expand=True)
                fp=oldest/max_oldest
                def draw_age_bar(e,c=bar_cv,f=fp,col=age_col):
                    w=c.winfo_width(); bw=max(4,int(w*f))
                    c.delete("all"); c.create_rectangle(0,2,bw,12,fill=col,outline="")
                def draw_age_bar_db(e,c=bar_cv,f=fp,col=age_col):
                    if hasattr(c,"_rd_after"): c.after_cancel(c._rd_after)
                    c._rd_after=c.after(150,lambda:draw_age_bar(None,c,f,col))
                bar_cv.bind("<Configure>",draw_age_bar_db)
                bar_outer.after(30,lambda c=bar_cv,f=fp,col=age_col:draw_age_bar(None,c,f,col))

            # Legend
            leg=tk.Frame(age_card,bg=CARD); leg.pack(anchor="w",pady=(10,0))
            tk.Label(leg,text="Age colour:  ",font=(FONT,8),bg=CARD,fg=MUTED).pack(side="left")
            for col,lbl in [(ACCENT2,"≤14 days"),(WARNING,"15–30 days"),(DANGER,"30+ days")]:
                tk.Frame(leg,bg=col,width=10,height=10).pack(side="left",padx=(0,4))
                tk.Label(leg,text=lbl,font=(FONT,8),bg=CARD,fg=TEXT2).pack(side="left",padx=(0,12))

        # ── Row 7: resolution time by type ───────────────────────────────────
        section_label("AVERAGE RESOLUTION TIME BY TYPE")

        res_data=[]
        for qt in QUERY_TYPES:
            resolved_qs=[q for q in qs if q["type"]==qt and q["status"]=="Resolved"
                         and q.get("resolved_date","")]
            days_list=[d for d in (resolve_days(q) for q in resolved_qs) if d is not None]
            if not days_list: continue
            res_data.append((qt, len(days_list), int(sum(days_list)/len(days_list)),
                             min(days_list), max(days_list)))

        if not res_data:
            tk.Label(rp,text="No resolved queries with resolution dates in this period.",
                     font=(FONT,9),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,24))
        else:
            res_card=tk.Frame(rp,bg=CARD,highlightthickness=1,highlightbackground=BORDER,
                              padx=16,pady=14); res_card.pack(fill="x",pady=(0,24))
            # Column headers
            hdr_row2=tk.Frame(res_card,bg=CARD); hdr_row2.pack(fill="x",pady=(0,6))
            for txt,w,anc in [("Query type",24,"w"),("Resolved",8,"center"),
                               ("Avg days",9,"center"),("Fastest",8,"center"),
                               ("Slowest",8,"center"),("Avg time bar",0,"w")]:
                tk.Label(hdr_row2,text=txt,font=(FONT,8,"bold"),bg=CARD,fg=MUTED,
                         width=w,anchor=anc).pack(side="left",padx=(0,4))
            tk.Frame(res_card,bg=BORDER,height=1).pack(fill="x",pady=(4,8))

            max_avg=max(d[2] for d in res_data) or 1
            for qt,count,avg,fastest,slowest in sorted(res_data,key=lambda x:-x[2]):
                res_col=(DANGER if avg>30 else WARNING if avg>14 else SUCCESS)
                row=tk.Frame(res_card,bg=CARD); row.pack(fill="x",pady=3)
                tk.Label(row,text=qt,          font=(FONT,9),bg=CARD,fg=TEXT2,  width=24,anchor="w").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(count),  font=(FONT,9,"bold"),bg=CARD,fg=TEXT,    width=8, anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(avg),    font=(FONT,9,"bold"),bg=CARD,fg=res_col, width=9, anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(fastest),font=(FONT,9),bg=CARD,fg=SUCCESS,        width=8, anchor="center").pack(side="left",padx=(0,4))
                tk.Label(row,text=str(slowest),font=(FONT,9),bg=CARD,fg=DANGER,         width=8, anchor="center").pack(side="left",padx=(0,4))
                # Avg time bar
                bar_outer2=tk.Frame(row,bg=CARD2,height=14); bar_outer2.pack(side="left",fill="x",expand=True)
                bar_outer2.pack_propagate(False)
                bar_cv2=tk.Canvas(bar_outer2,bg=CARD2,highlightthickness=0,height=14)
                bar_cv2.pack(fill="both",expand=True)
                fp2=avg/max_avg
                def draw_res_bar(e,c=bar_cv2,f=fp2,col=res_col):
                    w=c.winfo_width(); bw=max(4,int(w*f))
                    c.delete("all"); c.create_rectangle(0,2,bw,12,fill=col,outline="")
                def draw_res_bar_db(e,c=bar_cv2,f=fp2,col=res_col):
                    if hasattr(c,"_rd_after"): c.after_cancel(c._rd_after)
                    c._rd_after=c.after(150,lambda:draw_res_bar(None,c,f,col))
                bar_cv2.bind("<Configure>",draw_res_bar_db)
                bar_outer2.after(30,lambda c=bar_cv2,f=fp2,col=res_col:draw_res_bar(None,c,f,col))

            # Legend
            leg2=tk.Frame(res_card,bg=CARD); leg2.pack(anchor="w",pady=(10,0))
            tk.Label(leg2,text="Avg resolution:  ",font=(FONT,8),bg=CARD,fg=MUTED).pack(side="left")
            for col,lbl in [(SUCCESS,"≤14 days"),(WARNING,"15–30 days"),(DANGER,"30+ days")]:
                tk.Frame(leg2,bg=col,width=10,height=10).pack(side="left",padx=(0,4))
                tk.Label(leg2,text=lbl,font=(FONT,8),bg=CARD,fg=TEXT2).pack(side="left",padx=(0,12))


    def _export_report(self, qs, period_label):
        """Export the legacy multi-sheet report workbook."""
        if not qs:
            messagebox.showwarning("Nothing to export","No queries match the current filters.",parent=self)
            return

        # Build a meaningful filename from active filters
        fname_parts=["Query_Report"]
        rpt_client=getattr(self,"_rpt_client","All")
        rpt_person=getattr(self,"_rpt_person","All")
        if rpt_client!="All":
            safe=rpt_client.replace(" ","_").replace("/","_")[:30]
            fname_parts.append(safe)
        if rpt_person!="All":
            safe=rpt_person.replace(" ","_").replace("/","_")[:20]
            fname_parts.append(safe)
        fname_parts.append(date.today().strftime("%Y%m%d"))
        default_name="_".join(fname_parts)+".xlsx"
        path=filedialog.asksaveasfilename(
            title="Save report as",defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],initialfile=default_name)
        if not path: return

        try:
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.series import SeriesLabel
            from openpyxl.chart.label import DataLabelList

            wb=openpyxl.Workbook()

            # ── Shared styles ─────────────────────────────────────────────────
            thin=Side(style="thin",color="CCCCCC")
            bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
            hdr_fill=PatternFill("solid",fgColor="1C2530")
            hdr_font=Font(bold=True,color="5EEAD4",name=FONT,size=10)
            title_font=Font(bold=True,size=13,name=FONT,color="0F1419")
            sub_font=Font(size=9,name=FONT,color="445566")
            body_font=Font(name=FONT,size=10)
            status_fills={"Open":"1A3560","In Progress":"2E2200",
                          "Pending info":"243040","Resolved":"0A2818"}
            priority_fills={"High":"2E1111","Medium":"2A1E00","Low":"0A2818"}
            today=date.today()

            def style_hdr(ws,row,cols,widths):
                for i,(col,w) in enumerate(zip(cols,widths),1):
                    c=ws.cell(row=row,column=i,value=col)
                    c.font=hdr_font; c.fill=hdr_fill
                    c.alignment=Alignment(horizontal="center",vertical="center")
                    c.border=bdr
                    ws.column_dimensions[get_column_letter(i)].width=w

            def title_block(ws,title,subtitle):
                ws["A1"]=title; ws["A1"].font=title_font
                ws["A2"]=subtitle; ws["A2"].font=sub_font
                ws["A3"]=f"Exported: {today.strftime('%d/%m/%Y')}  ·  {len(qs)} queries"
                ws["A3"].font=sub_font
                ws.row_dimensions[1].height=22
                ws.sheet_view.showGridLines=False

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 0: Charts Overview (first sheet)
            # ═══════════════════════════════════════════════════════════════════
            wsc=wb.active; wsc.title="Charts Overview"
            wsc.sheet_view.showGridLines=False
            wsc["A1"]="Query Report — Charts Overview"
            wsc["A1"].font=Font(bold=True,size=14,name=FONT,color="0F1419")
            wsc["A2"]=period_label
            wsc["A2"].font=Font(size=9,name=FONT,color="445566")
            wsc["A3"]=f"Exported: {today.strftime('%d/%m/%Y')}  ·  {len(qs)} queries"
            wsc["A3"].font=Font(size=9,name=FONT,color="445566")
            # SLA summary in header
            qs_raised_exp=[q for q in qs if q.get("raised_date","")]
            if qs_raised_exp:
                def _exp_intake(q):
                    try:
                        return sla_intake_working_days(q.get("raised_date", ""), q.get("opened", "")) or 0
                    except: return 0
                n_met_exp=sum(1 for q in qs_raised_exp if _exp_intake(q)<=1)
                pct_exp=int(n_met_exp/len(qs_raised_exp)*100)
                wsc["A4"]=f"1-working-day SLA: {pct_exp}% met  ·  {n_met_exp}/{len(qs_raised_exp)} queries logged within 1 working day of client raising"
                sla_color="166534" if pct_exp>=80 else "92400E" if pct_exp>=50 else "991B1B"
                wsc["A4"].font=Font(size=9,name=FONT,color=sla_color,bold=True)

            # ── Chart data helper sheet (hidden) ──────────────────────────────
            wsd=wb.create_sheet("_ChartData")
            wsd.sheet_state="hidden"

            # ── Data block 1: By Status (no header row — avoids "Count" label) ─
            status_row_start=1
            for i,s in enumerate(STATUSES,status_row_start):
                wsd.cell(row=i,column=1,value=s)
                wsd.cell(row=i,column=2,value=sum(1 for q in qs if q["status"]==s))
            status_row_end=status_row_start+len(STATUSES)-1

            # ── Data block 2: By Priority ──────────────────────────────────────
            pri_row_start=1
            for i,p in enumerate(PRIORITIES,pri_row_start):
                wsd.cell(row=i,column=4,value=p)
                wsd.cell(row=i,column=5,value=sum(1 for q in qs if q["priority"]==p))
            pri_row_end=pri_row_start+len(PRIORITIES)-1

            # ── Data block 3: By Query Type (no header — avoids "Count" label) ─
            type_row_start=1; type_rows=[]
            for qt in QUERY_TYPES:
                type_qs=[q for q in qs if q["type"]==qt]
                if not type_qs: continue
                r=type_row_start+len(type_rows)
                wsd.cell(row=r,column=7,value=qt)
                wsd.cell(row=r,column=8,value=len(type_qs))
                wsd.cell(row=r,column=9,value=sum(1 for q in type_qs if q["status"]!="Resolved"))
                wsd.cell(row=r,column=10,value=sum(1 for q in type_qs if q["status"]=="Resolved"))
                type_rows.append(r)
            type_row_end=type_rows[-1] if type_rows else type_row_start

            # ── Data block 5: Avg resolution time ─────────────────────────────
            res_rows=[]
            for qt in QUERY_TYPES:
                res_qs=[q for q in qs if q["type"]==qt and q["status"]=="Resolved"
                        and q.get("resolved_date") and q.get("opened")]
                days_list=[]
                for q in res_qs:
                    try:
                        o=datetime.strptime(q["opened"],"%Y-%m-%d").date()
                        rv=datetime.strptime(q["resolved_date"],"%Y-%m-%d").date()
                        days_list.append(max(0,(rv-o).days))
                    except: pass
                if not days_list: continue
                r=1+len(res_rows)
                wsd.cell(row=r,column=12,value=qt)
                wsd.cell(row=r,column=13,value=int(sum(days_list)/len(days_list)))
                res_rows.append(r)

            from openpyxl.chart.series import DataPoint as ChartDataPoint, SeriesLabel

            BAR_COLOURS=["2DD4BF","7EC8C8","8FC98F","E8C07A","E89090",
                         "B8A0E0","7AB8D8","E8A87A","A0C878","90B8C8",
                         "D898C8","C8B890"]

            def colour_bars(series, n):
                pts=[]
                for i in range(n):
                    dp=ChartDataPoint(idx=i)
                    col=BAR_COLOURS[i%len(BAR_COLOURS)]
                    dp.spPr.solidFill=col
                    dp.spPr.ln.solidFill=col
                    pts.append(dp)
                series.dPt=pts

            def val_labels():
                dl=DataLabelList()
                dl.showVal=True; dl.showCatName=False
                dl.showSerName=False; dl.showPercent=False
                return dl

            def axis_font(sz_pt=9):
                """RichText for category axis label font size."""
                from openpyxl.chart.text import RichText
                from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                                    ParagraphProperties, CharacterProperties)
                cp=CharacterProperties(sz=sz_pt*100)
                para=Paragraph(pPr=ParagraphProperties(defRPr=cp))
                return RichText(p=[para])

            # All charts placed in a single column, each starting well below the last.
            # Each chart is 32cm wide x 18cm tall.
            # openpyxl row unit ≈ 0.635cm at height=18, so 18cm needs ~28 rows.
            # We space anchors 32 rows apart to give a clear gap between charts.

            CHART_W=32; CHART_H=18

            # ── Chart 1: Status Pie — row 6 ──────────────────────────────────
            pie1=PieChart()
            pie1.title="Queries by Status"
            pie1.style=2; pie1.width=CHART_W; pie1.height=CHART_H
            pie1.dataLabels=DataLabelList()
            pie1.dataLabels.showPercent=True; pie1.dataLabels.showCatName=True
            pie1.dataLabels.showVal=False; pie1.dataLabels.showSerName=False
            data1=Reference(wsd,min_col=2,min_row=status_row_start,max_row=status_row_end)
            cats1=Reference(wsd,min_col=1,min_row=status_row_start,max_row=status_row_end)
            pie1.add_data(data1,titles_from_data=False)
            pie1.set_categories(cats1); pie1.series[0].title=None
            wsc.add_chart(pie1,"B6")

            # ── Chart 2: Priority Pie — row 40 ───────────────────────────────
            pie2=PieChart()
            pie2.title="Queries by Priority"
            pie2.style=2; pie2.width=CHART_W; pie2.height=CHART_H
            pie2.dataLabels=DataLabelList()
            pie2.dataLabels.showPercent=True; pie2.dataLabels.showCatName=True
            pie2.dataLabels.showVal=False; pie2.dataLabels.showSerName=False
            data2=Reference(wsd,min_col=5,min_row=pri_row_start,max_row=pri_row_end)
            cats2=Reference(wsd,min_col=4,min_row=pri_row_start,max_row=pri_row_end)
            pie2.add_data(data2,titles_from_data=False)
            pie2.set_categories(cats2); pie2.series[0].title=None
            wsc.add_chart(pie2,"B40")

            # ── Chart 3: Total by type — coloured bar — row 74 ───────────────
            if type_rows:
                n_types=len(type_rows)
                bar1=BarChart()
                bar1.type="bar"; bar1.grouping="clustered"
                bar1.title="Queries by Type — Total"
                bar1.style=2; bar1.width=CHART_W
                bar1.height=max(CHART_H, n_types*2.0)
                bar1.legend=None
                # horizontal bar: x_axis = category axis (query type names on left)
                #                 y_axis = value axis (counts along bottom)
                bar1.x_axis.delete=False
                bar1.x_axis.tickLblPos="low"
                bar1.x_axis.txPr=axis_font(9)
                bar1.y_axis.numFmt="0"; bar1.y_axis.title="Count"
                bar1.y_axis.delete=False
                bar1.dataLabels=val_labels()
                bdata1=Reference(wsd,min_col=8,max_col=8,
                                 min_row=type_row_start,max_row=type_row_end)
                bcats1=Reference(wsd,min_col=7,
                                 min_row=type_row_start,max_row=type_row_end)
                bar1.add_data(bdata1,titles_from_data=False)
                bar1.set_categories(bcats1)
                bar1.series[0].title=None
                colour_bars(bar1.series[0], n_types)
                wsc.add_chart(bar1,"B74")

            # ── Chart 4: Open vs Resolved stacked — row 114 ──────────────────
            if type_rows:
                bar2=BarChart()
                bar2.type="bar"; bar2.grouping="stacked"
                bar2.title="Open vs Resolved by Type"
                bar2.style=2; bar2.width=CHART_W
                bar2.height=max(CHART_H, n_types*2.0)
                bar2.x_axis.delete=False
                bar2.x_axis.tickLblPos="low"
                bar2.x_axis.txPr=axis_font(9)
                bar2.y_axis.numFmt="0"; bar2.y_axis.title="Count"
                bar2.y_axis.delete=False
                bar2.dataLabels=val_labels()
                bdata2=Reference(wsd,min_col=9,max_col=10,
                                 min_row=type_row_start,max_row=type_row_end)
                bcats2=Reference(wsd,min_col=7,
                                 min_row=type_row_start,max_row=type_row_end)
                bar2.add_data(bdata2,titles_from_data=False)
                bar2.set_categories(bcats2)
                bar2.series[0].title=SeriesLabel(v="Open")
                bar2.series[1].title=SeriesLabel(v="Resolved")
                bar2.series[0].graphicalProperties.solidFill="C89050"
                bar2.series[0].graphicalProperties.line.solidFill="C89050"
                bar2.series[1].graphicalProperties.solidFill="4A9E8A"
                bar2.series[1].graphicalProperties.line.solidFill="4A9E8A"
                wsc.add_chart(bar2,"B114")

            # ── Chart 5: Avg resolution time — coloured bar — row 154 ────────
            if res_rows:
                res_row_end=res_rows[-1]
                n_res=len(res_rows)
                bar3=BarChart()
                bar3.type="bar"; bar3.grouping="clustered"
                bar3.title="Average Days to Resolve by Type"
                bar3.style=2; bar3.width=CHART_W
                bar3.height=max(CHART_H, n_res*2.0)
                bar3.legend=None
                bar3.x_axis.delete=False
                bar3.x_axis.tickLblPos="low"
                bar3.x_axis.txPr=axis_font(9)
                bar3.y_axis.numFmt="0"; bar3.y_axis.title="Days"
                bar3.y_axis.delete=False
                bar3.dataLabels=val_labels()
                bdata3=Reference(wsd,min_col=13,max_col=13,
                                 min_row=1,max_row=res_row_end)
                bcats3=Reference(wsd,min_col=12,min_row=1,max_row=res_row_end)
                bar3.add_data(bdata3,titles_from_data=False)
                bar3.set_categories(bcats3)
                bar3.series[0].title=None
                colour_bars(bar3.series[0], n_res)
                wsc.add_chart(bar3,"B154")

            # Chart sheet formatting — wide cols, consistent row heights
            for col_letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                wsc.column_dimensions[col_letter].width=6
            for row_n in range(1,200):
                wsc.row_dimensions[row_n].height=15

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 1: Query List (full detail, client-ready)
            # ═══════════════════════════════════════════════════════════════════
            ws1=wb.create_sheet("Query List")
            title_block(ws1,"Query Report — Full Detail",period_label)

            cols1=["Ref","Client","Fund","Site","Address","Utility","Meter",
                   "Type","Status","Priority","Query Raised","Logged","Action Date","Resolved",
                   "SLA Intake","Days Open","Description","Activity Log",
                   "Supply Point ID","Meter Serial","Contact","Prop Code",
                   "Assigned To","Last Updated By"]
            widths1=[12,24,22,26,30,14,20,
                     20,14,10,12,12,12,12,
                     12,10,50,60,
                     18,14,24,12,
                     16,16]
            style_hdr(ws1,5,cols1,widths1)

            def fmt_log_bullets(log_str):
                """Convert pipe-separated log entries to bullet-point lines."""
                if not log_str: return ""
                entries=public_log_entries(log_str)
                return "\n".join(f"• {e}" for e in entries)

            def days_open_str(q):
                try:
                    opened=datetime.strptime(q["opened"],"%Y-%m-%d").date()
                    if q.get("resolved_date"):
                        end=datetime.strptime(q["resolved_date"],"%Y-%m-%d").date()
                    else:
                        end=date.today()
                    return (end-opened).days
                except: return ""

            def intake_days(q):
                try:
                    rd=q.get("raised_date",""); od=q.get("opened","")
                    if not rd or not od: return ""
                    return sla_intake_working_days(rd, od)
                except: return ""

            for r,q in enumerate(qs,6):
                log_bullets=fmt_log_bullets(q.get("log",""))
                days=days_open_str(q)
                intake=intake_days(q)
                vals=[
                    q["ref"], q["client"], q.get("fund",""), q["site"],
                    q.get("address",""), q.get("utility",""), q.get("meter",""),
                    q["type"], q["status"], q["priority"],
                    fmt_date(q.get("raised_date","")), fmt_date(q["opened"]),
                    fmt_date(q.get("chase_date","")),
                    fmt_date(q.get("resolved_date","")), intake, days,
                    q["desc"], log_bullets,
                    q.get("spid",""), q.get("serial",""), q.get("contact",""),
                    q.get("prop_code",""), q.get("assigned_to",""),
                    q.get("last_by",""),
                ]
                for c,v in enumerate(vals,1):
                    cell=ws1.cell(row=r,column=c,value=v)
                    cell.font=body_font; cell.border=bdr
                    is_log=(c==18); is_desc=(c==17)
                    cell.alignment=Alignment(vertical="top",wrap_text=(is_log or is_desc))
                sf=status_fills.get(q["status"],"243040")
                pf=priority_fills.get(q["priority"],"243040")
                ws1.cell(row=r,column=9).fill=PatternFill("solid",fgColor=sf)
                ws1.cell(row=r,column=9).font=Font(name=FONT,size=10,color="FFFFFF")
                ws1.cell(row=r,column=10).fill=PatternFill("solid",fgColor=pf)
                ws1.cell(row=r,column=10).font=Font(name=FONT,size=10,color="FFFFFF")
                if intake != "":
                    c15=ws1.cell(row=r,column=15)
                    sla_label="BREACHED" if intake>1 else ("1 day" if intake==1 else "MET")
                    c15.value=f"{intake}d — {sla_label}"
                    if intake>1: c15.fill=PatternFill("solid",fgColor="2E1111"); c15.font=Font(name=FONT,size=10,color="FCA5A5")
                    elif intake==1: c15.fill=PatternFill("solid",fgColor="2A1E00"); c15.font=Font(name=FONT,size=10,color="FDE68A")
                    else: c15.fill=PatternFill("solid",fgColor="0A2818"); c15.font=Font(name=FONT,size=10,color="6EE7B7")
                n_entries=log_bullets.count("\n")+1 if log_bullets else 1
                ws1.row_dimensions[r].height=max(18, min(n_entries*14, 120))

            ws1.freeze_panes="A6"
            ws1.auto_filter.ref=f"A5:{get_column_letter(len(cols1))}5"

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 2: By Type
            # ═══════════════════════════════════════════════════════════════════
            ws2=wb.create_sheet("By Type")
            title_block(ws2,"Breakdown by Query Type",period_label)
            cols2=["Query Type","Total","Open","In Progress","Pending","Resolved","% Resolved"]
            widths2=[24,8,8,12,10,10,12]
            style_hdr(ws2,5,cols2,widths2)
            for r,qt in enumerate(QUERY_TYPES,6):
                tqs=[q for q in qs if q["type"]==qt]
                if not tqs: continue
                total=len(tqs)
                res_n=sum(1 for q in tqs if q["status"]=="Resolved")
                pct=f"{round(res_n/total*100)}%" if total else "—"
                for c,v in enumerate([qt,total,
                    sum(1 for q in tqs if q["status"]=="Open"),
                    sum(1 for q in tqs if q["status"]=="In Progress"),
                    sum(1 for q in tqs if q["status"]=="Pending info"),
                    res_n,pct],1):
                    cell=ws2.cell(row=r,column=c,value=v)
                    cell.font=body_font; cell.border=bdr
                    cell.alignment=Alignment(horizontal="center" if c>1 else "left")

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 3: Open Query Age
            # ═══════════════════════════════════════════════════════════════════
            ws3=wb.create_sheet("Open Query Age")
            title_block(ws3,"Age of Open Queries by Type",period_label)
            cols3=["Query Type","Open Count","Oldest (days)","Avg Age (days)","Newest (days)"]
            widths3=[24,12,14,14,14]
            style_hdr(ws3,5,cols3,widths3)
            r=6
            for qt in QUERY_TYPES:
                open_qs=[q for q in qs if q["type"]==qt and q["status"]!="Resolved"]
                if not open_qs: continue
                ages=[(today-datetime.strptime(q["opened"],"%Y-%m-%d").date()).days
                      for q in open_qs if q.get("opened")]
                if not ages: continue
                for c,v in enumerate([qt,len(open_qs),max(ages),int(sum(ages)/len(ages)),min(ages)],1):
                    cell=ws3.cell(row=r,column=c,value=v)
                    cell.font=body_font; cell.border=bdr
                    cell.alignment=Alignment(horizontal="center" if c>1 else "left")
                if max(ages)>30:
                    ws3.cell(row=r,column=3).fill=PatternFill("solid",fgColor="2E1111")
                    ws3.cell(row=r,column=3).font=Font(name=FONT,size=10,color="FCA5A5")
                r+=1

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 4: Resolution Time
            # ═══════════════════════════════════════════════════════════════════
            ws4=wb.create_sheet("Resolution Time")
            title_block(ws4,"Average Resolution Time by Type",period_label)
            cols4=["Query Type","Resolved Count","Avg Days","Fastest (days)","Slowest (days)"]
            widths4=[24,14,10,14,14]
            style_hdr(ws4,5,cols4,widths4)
            r=6
            for qt in QUERY_TYPES:
                res_qs=[q for q in qs if q["type"]==qt and q["status"]=="Resolved"
                        and q.get("resolved_date") and q.get("opened")]
                days_list=[]
                for q in res_qs:
                    try:
                        o=datetime.strptime(q["opened"],"%Y-%m-%d").date()
                        rv=datetime.strptime(q["resolved_date"],"%Y-%m-%d").date()
                        days_list.append(max(0,(rv-o).days))
                    except: pass
                if not days_list: continue
                avg=int(sum(days_list)/len(days_list))
                for c,v in enumerate([qt,len(days_list),avg,min(days_list),max(days_list)],1):
                    cell=ws4.cell(row=r,column=c,value=v)
                    cell.font=body_font; cell.border=bdr
                    cell.alignment=Alignment(horizontal="center" if c>1 else "left")
                if avg>30:
                    ws4.cell(row=r,column=3).fill=PatternFill("solid",fgColor="2E1111")
                    ws4.cell(row=r,column=3).font=Font(name=FONT,size=10,color="FCA5A5")
                r+=1

            # ═══════════════════════════════════════════════════════════════════
            # SHEET 5: Dashboard Upload (flat, clean, no merges — Power BI ready)
            # ═══════════════════════════════════════════════════════════════════
            ws5=wb.create_sheet("Dashboard Upload")
            ws5.sheet_view.showGridLines=False
            # Plain header row — no title block, no merges, column names only
            # One row per query, all fields, log as bullet points in one cell
            upload_cols=[
                "Reference","Client","Fund","Site","Address",
                "Utility","Meter","Query Type","Status","Priority",
                "Query Raised","Logged","Action Date","Resolved Date",
                "SLA Intake","Days Open",
                "Description","Activity Notes",
                "Supply Point ID","Meter Serial","Contact","Property Code",
                "Assigned To","Last Updated By","Last Updated Date",
                "Export Date"
            ]
            upload_widths=[
                12,26,22,28,32,
                14,20,22,14,10,
                12,12,12,12,
                12,10,
                52,62,
                18,14,26,12,
                16,16,14,
                12
            ]
            # Style header — light background, dark text (easy to read in Power BI)
            upload_hdr_fill=PatternFill("solid",fgColor="1C2530")
            upload_hdr_font=Font(bold=True,color="5EEAD4",name=FONT,size=10)
            for i,(col,w) in enumerate(zip(upload_cols,upload_widths),1):
                hc=ws5.cell(row=1,column=i,value=col)
                hc.font=upload_hdr_font; hc.fill=upload_hdr_fill
                hc.alignment=Alignment(horizontal="center",vertical="center")
                hc.border=bdr
                ws5.column_dimensions[get_column_letter(i)].width=w
            ws5.row_dimensions[1].height=20

            export_date=date.today().strftime("%d/%m/%Y")
            for r,q in enumerate(qs,2):
                log_bullets=fmt_log_bullets(q.get("log",""))
                days=days_open_str(q)
                intake=intake_days(q)
                row_vals=[
                    q["ref"], q["client"], q.get("fund",""), q["site"],
                    q.get("address",""),
                    q.get("utility",""), q.get("meter",""),
                    q["type"], q["status"], q["priority"],
                    fmt_date(q.get("raised_date","")),
                    fmt_date(q["opened"]),
                    fmt_date(q.get("chase_date","")),
                    fmt_date(q.get("resolved_date","")),
                    intake, days,
                    q["desc"], log_bullets,
                    q.get("spid",""), q.get("serial",""),
                    q.get("contact",""), q.get("prop_code",""),
                    q.get("assigned_to",""),
                    q.get("last_by",""), fmt_date(q.get("last_date","")),
                    export_date,
                ]
                for c,v in enumerate(row_vals,1):
                    cell=ws5.cell(row=r,column=c,value=v)
                    cell.font=Font(name=FONT,size=10)
                    cell.border=bdr
                    cell.alignment=Alignment(
                        vertical="top",
                        wrap_text=(c in (17,18))  # desc + log cols shifted by 2
                    )
                # Colour intake days
                if intake != "":
                    ci=ws5.cell(row=r,column=15)
                    sla_lbl="BREACHED" if intake>1 else ("1 day" if intake==1 else "MET")
                    ci.value=f"{intake}d — {sla_lbl}"
                    if intake>1: ci.fill=PatternFill("solid",fgColor="2E1111"); ci.font=Font(name=FONT,size=10,color="FCA5A5")
                    elif intake==1: ci.fill=PatternFill("solid",fgColor="2A1E00"); ci.font=Font(name=FONT,size=10,color="FDE68A")
                    else: ci.fill=PatternFill("solid",fgColor="0A2818"); ci.font=Font(name=FONT,size=10,color="6EE7B7")
                n_entries=log_bullets.count("\n")+1 if log_bullets else 1
                ws5.row_dimensions[r].height=max(18, min(n_entries*14, 120))

            ws5.freeze_panes="A2"
            ws5.auto_filter.ref=f"A1:{get_column_letter(len(upload_cols))}1"

            # Make Charts Overview the active sheet (first tab selected on open)
            wb.active=wsc

            wb.save(path)
            messagebox.showinfo("Report exported",
                f"Saved {len(qs)} queries to:\n{path}\n\n"
                f"Sheets:\n"
                f"  • Charts Overview\n"
                f"  • Query List (full detail + bullet-point notes)\n"
                f"  • By Type\n"
                f"  • Open Query Age\n"
                f"  • Resolution Time\n"
                f"  • Dashboard Upload (flat format for Power BI / client portals)")
            open_file(path)

        except PermissionError:
            messagebox.showerror("File in use",
                "Cannot save — the file may already be open in Excel.\nClose it and try again.")
        except Exception as e:
            messagebox.showerror("Export failed",str(e))

    def _get_excel_mtime(self):
        try:
            # Use both mtime and file size to improve change detection on synced files.
            if not self.excel_file or not os.path.exists(self.excel_file):
                return (0,0)
            return (os.path.getmtime(self.excel_file), os.path.getsize(self.excel_file))
        except: return (0,0)

    def _start_auto_reload(self):
        if self._auto_reload_running: return
        self._auto_reload_running=True
        self._auto_reload_thread=threading.Thread(target=self._auto_reload_loop,daemon=True)
        self._auto_reload_thread.start()

    def _stop_auto_reload(self):
        self._auto_reload_running=False

    # ── Daily automatic backup ────────────────────────────────────────────────
    def _start_daily_backup(self):
        """Start a background thread that takes one backup per calendar day."""
        if getattr(self, "_backup_running", False):
            return
        self._backup_running = True
        self._last_backup_date = None   # force first-run backup today
        t = threading.Thread(target=self._daily_backup_loop, daemon=True)
        t.start()

    def _daily_backup_loop(self):
        """Wake every 60 seconds; fire a backup whenever the calendar date rolls over."""
        sleep_event = threading.Event()
        while getattr(self, "_backup_running", False):
            today = datetime.now().date()
            if today != self._last_backup_date:
                self._do_daily_backup()
            sleep_event.wait(60)
            sleep_event.clear()

    def _do_daily_backup(self, forced=False):
        """Copy query_tracker.xlsx → Data/Backups/query_tracker_YYYY-MM-DD.xlsx.

        forced=True: used on app close — always run even if we already backed
        up today, so the file contains this session's latest saves.
        """
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return
            today = datetime.now().date()
            if not forced and today == self._last_backup_date:
                return   # already backed up today

            backup_dir = os.path.join(os.path.dirname(self.excel_file), "Backups")
            os.makedirs(backup_dir, exist_ok=True)

            stamp = datetime.now().strftime("%Y-%m-%d")
            dest = os.path.join(backup_dir, f"query_tracker_{stamp}.xlsx")
            import shutil
            shutil.copy2(self.excel_file, dest)
            self._last_backup_date = today

            # Prune backups older than 30 days
            cutoff = datetime.now().timestamp() - 30 * 86400
            for f in os.listdir(backup_dir):
                fp = os.path.join(backup_dir, f)
                try:
                    if os.path.isfile(fp) and os.path.getmtime(fp) < cutoff:
                        os.remove(fp)
                except Exception:
                    pass

            if not forced:
                self.after(0, lambda d=dest: _show_toast(self, f"Backup saved: {os.path.basename(d)}"))
        except Exception:
            pass
    # ─────────────────────────────────────────────────────────────────────────

    def _auto_reload_loop(self):
        """Poll the Excel file frequently; force periodic reloads to handle sync lag."""
        sleep_event=threading.Event()
        while self._auto_reload_running:
            sleep_event.wait(10)
            sleep_event.clear()
            if not self._auto_reload_running: break
            try:
                self._reload_tick += 1
                new_mtime=self._get_excel_mtime()
                # Also force a reload every 60 seconds in case SharePoint sync
                # updates content without a detectable local timestamp bump.
                if (new_mtime and new_mtime!=self._excel_mtime) or (self._reload_tick % 6 == 0):
                    self._excel_mtime=new_mtime
                    self.after(0,self._silent_reload)
            except: pass

    def _sync_now(self):
        """Manual on-demand refresh for users who need immediate updates."""
        try:
            self.sync_lbl.config(text="↻ Syncing...")
        except: pass
        self.after(0,self._silent_reload)

    def _silent_reload(self):
        """Reload queries from disk without disrupting any open dialogs."""
        try:
            fresh=load_queries(self.excel_file)
            if fresh is not None:
                self.queries=fresh
                self._excel_mtime=self._get_excel_mtime()
                self._dash_dirty=True
                self._rpt_dirty=True
                # Re-read shared query types / team members in case another user updated them
                shared=load_shared_settings(self.excel_file)
                if shared.get("query_types"):
                    QUERY_TYPES[:]=shared["query_types"]
                if shared.get("team_members"):
                    self.team_members=shared["team_members"]
                self._refresh_table()
                self._show_daily_banner()
                # Check for incoming transfer notifications on every reload
                self.after(200,self._check_notifications)
                # Refresh mini window if open
                try:
                    if getattr(self,"_mini_refresh",None): self._mini_refresh()
                except: pass
                ts=datetime.now().strftime("%H:%M:%S")
                try:
                    self.sync_lbl.config(text=f"↻ Synced {ts}")
                    self.after(5000,lambda:self.sync_lbl.config(text="● Live") if self.sync_lbl.winfo_exists() else None)
                except: pass
        except: pass

    def _excel_mtime_now(self):
        """Return current mtime — used by detail dialog to detect stale edits."""
        return self._get_excel_mtime()

    def _start_watcher(self):
        if self._watcher_running: return
        inbox = get_drop_inbox(self.sites_file)
        if not inbox: return
        self._watcher_running = True
        try:
            self._watcher_seen = set(os.listdir(inbox))
        except: self._watcher_seen = set()
        self._watcher_event = threading.Event()
        self._watcher_thread = threading.Thread(
            target=self._watcher_loop, daemon=True)
        self._watcher_thread.start()
        if not hasattr(self,"_inbox_hint_shown"):
            self._inbox_hint_shown=True
            _show_toast(self,f"📂 Drop inbox active\n{inbox}",
                        color="#1A2E48", duration=5000)

    def _stop_watcher(self):
        self._watcher_running = False
        if hasattr(self,"_watcher_event"):
            self._watcher_event.set()  # wake sleeping thread so it exits cleanly

    def _watcher_loop(self):
        reuse_event = threading.Event()
        while self._watcher_running:
            try:
                inbox = get_drop_inbox(self.sites_file)
                if inbox and os.path.exists(inbox):
                    current = set(os.listdir(inbox))
                    new_files = current - self._watcher_seen
                    for fname in sorted(new_files):
                        fpath = os.path.join(inbox, fname)
                        if not os.path.isfile(fpath): continue
                        # Wait until file is fully written (size stable for 0.5s)
                        if not self._wait_for_file_ready(fpath): continue
                        self._watcher_seen.add(fname)
                        self.after(0, lambda f=fpath, n=fname:
                                   self._process_inbox_file(f, n))
            except: pass
            reuse_event.wait(2)
            reuse_event.clear()

    def _wait_for_file_ready(self, fpath, timeout=8.0):
        """Wait until a file's size is stable (fully written). Returns True if ready."""
        import time
        deadline = time.time() + timeout
        prev_size = -1
        while time.time() < deadline:
            try:
                size = os.path.getsize(fpath)
                if size == prev_size and size > 0:
                    # Try opening to confirm not locked
                    try:
                        with open(fpath, 'rb'): pass
                        return True
                    except (PermissionError, OSError):
                        pass  # still locked, keep waiting
                prev_size = size
            except OSError:
                return False
            threading.Event().wait(0.5)
        return os.path.exists(fpath)  # timeout — try anyway

    def _process_inbox_file(self, fpath, fname):
        if not os.path.exists(fpath): return
        # Immediate toast so user always knows something landed
        _show_toast(self, f"📎  File dropped: {fname[:50]}", color="#1A2E48", duration=3000)
        q = extract_ref_from_filename(fname, self.queries)
        if q:
            self._file_to_query(fpath, fname, q)
        else:
            # Small delay lets any active grab settle before showing dialog
            self.after(500, lambda: self._ask_assign_query(fpath, fname))


    def _file_to_query(self, fpath, fname, q):
        try:
            dest_name, dest = save_attachment(self.sites_file, q, fpath)
            if dest_name:
                os.remove(fpath)
                q["log"] += " | " + stamp(self.username) + f" Auto-filed attachment: {dest_name}"
                save_all_queries(self.queries, self.excel_file)
                self._refresh_table()
                _show_toast(self,
                    f"✓  Filed under {q['ref']}\n{dest_name}",
                    color=SUCCESS, duration=4000)
        except Exception as e:
            _show_toast(self, f"⚠  Could not file {fname}:\n{e}",
                        color=DANGER, duration=5000)

    def _ask_assign_query(self, fpath, fname):
        if not os.path.exists(fpath): return
        dlg = tk.Toplevel(self); dlg.title("Assign file to query")
        dlg.geometry("520x480"); dlg.configure(bg=BG)
        dlg.resizable(False, True)
        dlg.attributes("-topmost", True)
        dlg.lift()
        dlg.focus_force()
        try: dlg.grab_set()
        except: pass  # non-fatal without modal grab


        hdr = tk.Frame(dlg, bg=NAV, padx=20, pady=14); hdr.pack(fill="x")
        tk.Label(hdr, text="📎  Assign file to query",
                 font=(FONT,12,"bold"), bg=NAV, fg="white").pack(anchor="w")
        tk.Label(hdr, text=fname, font=(FONT,9), bg=NAV, fg=TEXT2).pack(anchor="w", pady=(4,0))

        body = tk.Frame(dlg, bg=BG, padx=20, pady=16); body.pack(fill="both", expand=True)
        tk.Label(body, text="No query reference found in the filename.\nSearch for a query to assign this file to:",
                 font=(FONT,10), bg=BG, fg=TEXT, justify="left").pack(anchor="w", pady=(0,12))

        search_var = tk.StringVar()
        sf = tk.Frame(body, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        sf.pack(fill="x", pady=(0,8))
        tk.Label(sf, text="🔍", font=(FONT,10), bg=CARD, fg=MUTED).pack(side="left", padx=(8,0))
        tk.Entry(sf, textvariable=search_var, font=(FONT,10), bg=CARD, fg=TEXT,
                 relief="flat", bd=6, highlightthickness=0).pack(side="left", fill="x", expand=True)

        list_frame = tk.Frame(body, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        list_frame.pack(fill="both", expand=True)
        lb = tk.Listbox(list_frame, font=(FONT,10), bg=CARD, fg=TEXT, relief="flat",
                        selectbackground=ACCENT, selectforeground="white",
                        activestyle="none", bd=0, highlightthickness=0)
        lbsb = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview,
                              style="Modern.Vertical.TScrollbar")
        lb.configure(yscrollcommand=lbsb.set)
        lbsb.pack(side="right", fill="y"); lb.pack(fill="both", expand=True, padx=2, pady=2)

        _visible = []

        def populate(term=""):
            lb.delete(0, "end"); _visible.clear()
            term = term.lower().strip()
            for q in self.queries:
                if q["status"] == "Resolved": continue
                text = f"{q['ref']}  ·  {q['client']}  ·  {q['site']}"
                if not term or term in text.lower():
                    lb.insert("end", text)
                    _visible.append(q)

        populate()
        search_var.trace_add("write", lambda *_: populate(search_var.get()))

        footer = tk.Frame(dlg, bg=CARD, highlightthickness=1,
                          highlightbackground=BORDER, padx=16, pady=12)
        footer.pack(fill="x", side="bottom")

        def discard():
            try:
                um = os.path.join(os.path.dirname(fpath), "_UNMATCHED")
                os.makedirs(um, exist_ok=True)
                shutil.move(fpath, os.path.join(um, fname))
            except: pass
            dlg.destroy()

        def assign():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("Select a query",
                    "Please select a query from the list.", parent=dlg)
                return
            q = _visible[sel[0]]
            dlg.destroy()
            self._file_to_query(fpath, fname, q)

        make_btn(footer, "Skip — move to Unmatched", discard, "default", padx=12, pady=6).pack(side="left")
        make_btn(footer, "Assign to selected query", assign, "primary", padx=14, pady=6).pack(side="right")
        lb.bind("<Double-1>", lambda e: assign())

    def _set_tab(self,val,refresh=True):
        self.tab_var.set(val)
        for v,btn in self._tab_btns.items():
            if v==val:
                btn.configure(fg=ACCENT2,font=(FONT,9,"bold"),bg=CARD2,
                              highlightthickness=1,highlightbackground=BORDER)
            else:
                btn.configure(fg=MUTED,font=(FONT,9),bg=BG,highlightthickness=0)
        if refresh: self._refresh_table()

    def _copy_selected_rows(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return "break"
        rows = []
        for iid in selected:
            vals = self.tree.item(iid, "values")
            rows.append("\t".join(str(v) for v in vals))
        try:
            self.clipboard_clear()
            self.clipboard_append("\n".join(rows))
            self.update_idletasks()
        except:
            pass
        return "break"

    def _get_filtered(self):
        tab=self.tab_var.get(); search=self.search_var.get().lower().strip()
        fc=self.filter_client.get(); ff=self.filter_fund.get()
        ft=self.filter_type.get(); fs=self.filter_status.get()
        fa=getattr(self,"filter_assignee",None)
        fa=fa.get() if fa else "All"
        fu=getattr(self,"filter_utility",None)
        fu=fu.get() if fu else "All"
        today=today_str(); result=[]
        af=getattr(self,"_assignee_filter","")
        day_filter=getattr(self,"_calendar_day_filter","")
        for q in self.queries:
            overdue=q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today
            if tab=="action" and not overdue: continue
            if tab=="open"   and q["status"]=="Resolved": continue
            if tab=="resolved" and q["status"]!="Resolved": continue
            if fc!="All" and q["client"]!=fc: continue
            if ff!="All" and q.get("fund","")!=ff: continue
            if ft!="All" and q["type"]!=ft: continue
            if fs!="All" and q["status"]!=fs: continue
            if fu!="All" and q.get("utility","")!=fu: continue
            if fa!="All" and q.get("assigned_to","")!=fa: continue
            if af and q.get("assigned_to","")!=af: continue
            if day_filter and q.get("chase_date","")!=day_filter: continue
            if search and search not in (q["ref"]+q["client"]+q.get("fund","")+q["site"]+
                                         q["desc"]+q["type"]+q["utility"]+q["meter"]).lower(): continue
            result.append(q)

        if tab=="action":
            pri_order={"High":0,"Medium":1,"Low":2}
            result.sort(key=lambda q:(
                pri_order.get(q.get("priority","Low"),3),
                parse_iso_date(q.get("chase_date", "")) or date.max,
                q.get("ref", ""),
            ))
        return result

    def _refresh_table(self):
        self.tree.delete(*self.tree.get_children()); today=today_str()
        for q in self._get_filtered():
            overdue=q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today
            if q["status"]=="Resolved": chase_disp="—"
            elif q.get("chase_date",""):
                chase_date=parse_iso_date(q["chase_date"])
                if not chase_date:
                    chase_disp=f"Invalid date: {q['chase_date']}"
                else:
                    d=(chase_date-date.today()).days
                    chase_disp="⚠ Today" if d<=0 else f"In {d}d" if d<=3 else fmt_date(q["chase_date"])
            else: chase_disp="Not set"
            od=days_overdue(q.get("chase_date","")) if overdue else 0
            overdue_disp=f"{od}d" if od>0 else ""
            
            # Calculate query age for aging alerts
            query_age = 0
            if q["status"] != "Resolved":
                try:
                    opened_date = datetime.strptime(q["opened"], "%Y-%m-%d").date()
                    query_age = (date.today() - opened_date).days
                except:
                    pass
            
            # Determine tag with priority order
            if overdue:                          tag="overdue"
            elif q["status"]=="Resolved":        tag="resolved"
            elif query_age > 30:                 tag="stale"    # >30 days old
            elif query_age > 14:                 tag="aging"    # >14 days old
            elif q["status"]=="In Progress":     tag="in_prog"
            elif q["priority"]=="High":          tag="high"
            else:                                tag="open"
            att_count=len(list_attachments(self.sites_file,q))
            att_disp=f"📎{att_count}" if att_count else ""
            raised_disp=fmt_date(q.get("raised_date","")) if q.get("raised_date") else "—"
            self.tree.insert("","end",iid=q["id"],
                values=(q["ref"],q["client"],q.get("fund",""),q["site"],q["utility"],q["meter"],
                        q["type"],q["status"],q["priority"],overdue_disp,chase_disp,
                        raised_disp,fmt_date(q["opened"]),q.get("assigned_to",""),q.get("last_by",""),att_disp),tags=(tag,))
        self._refresh_metrics()
        n=len(self._get_filtered())
        af = getattr(self, "_assignee_filter", "")
        df = getattr(self, "_calendar_day_filter", "")
        af_txt=f"  ·  Assigned to: {af}" if af else ""
        df_txt=f"  ·  Action date: {fmt_date(df)}" if df else ""
        self.status_lbl.config(text=f"{n} of {len(self.queries)} queries shown{af_txt}{df_txt}  ·  Double-click to open  ·  {self.excel_file or 'No file configured'}")

    def _refresh_metrics(self):
        for w in self.metrics_frame.winfo_children(): w.destroy()
        today=today_str()
        chase_n=sum(1 for q in self.queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today)
        
        metrics = [
            ("Total queries",  len(self.queries),                                          ACCENT,      None),
            ("Open",           sum(1 for q in self.queries if q["status"]!="Resolved"),    ACCENT2,     "open"),
            ("Action today",    chase_n,                                                    DANGER if chase_n else MUTED,  "action"),
            ("Resolved",       sum(1 for q in self.queries if q["status"]=="Resolved"),    SUCCESS,     "resolved")
        ]
        
        for label, val, color, tab_val in metrics:
            outer=tk.Frame(self.metrics_frame,bg=CARD2,highlightthickness=1,highlightbackground=BORDER, cursor="hand2" if tab_val else "arrow")
            outer.pack(side="left",padx=(0,10))
            tk.Frame(outer,bg=color,height=3).pack(fill="x")
            inner=tk.Frame(outer,bg=CARD2,padx=16,pady=12); inner.pack(fill="both",expand=True)
            val_lbl=tk.Label(inner,text=str(val),font=(FONT,22,"bold"),bg=CARD2,fg=color)
            val_lbl.pack(anchor="w")
            lbl_widget=tk.Label(inner,text=label.upper(),font=(FONT,7,"bold"),bg=CARD2,fg=MUTED)
            lbl_widget.pack(anchor="w",pady=(2,0))
            
            # Add click handler — navigate by tab, not by status filter
            if tab_val:
                def make_click_handler(t_val, c=color, o=outer):
                    def on_click(e):
                        self._clear_filters(refresh=False)
                        self._show_page("list")
                        self._set_tab(t_val)
                    return on_click
                
                click_handler = make_click_handler(tab_val)
                for widget in [outer, inner, val_lbl, lbl_widget]:
                    widget.bind("<Button-1>", click_handler)
                    widget.bind("<Enter>", lambda e, c=color, o=outer: o.configure(highlightbackground=c))
                    widget.bind("<Leave>", lambda e, o=outer: o.configure(highlightbackground=BORDER))

    def _show_daily_banner(self):
        today=today_str()
        items=[q for q in self.queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today]
        if items:
            preview=", ".join(f"{q['ref']} ({q['client']})" for q in items[:3])
            extra=f" +{len(items)-3} more" if len(items)>3 else ""
            tone=DANGER if len(items)>=3 else WARNING
            self.banner_strip.config(bg=tone)
            self.banner_frame.config(highlightbackground=tone)
            self.banner_lbl.config(text=f"Action needed today  •  {len(items)} quer{'y' if len(items)==1 else 'ies'} require follow-up  •  {preview}{extra}")
            self.banner_frame.pack(fill="x",padx=20,pady=(12,0))
        else: self.banner_frame.pack_forget()

    def _check_notifications(self):
        """Read the _Notifications sheet and show a banner for any unread transfers."""
        if not self.excel_file or not os.path.exists(self.excel_file): return
        try:
            wb=openpyxl.load_workbook(self.excel_file)
            if "_Notifications" not in wb.sheetnames: return
            nws=wb["_Notifications"]
            unread=[]
            for row in nws.iter_rows(min_row=2,values_only=False):
                if len(row)>=2 and row[1].value=="N":
                    msg=row[0].value or ""
                    unread.append((row[0].row, msg))
            if not unread: return

            # Show a dialog listing all unread transfers
            nd=tk.Toplevel(self); nd.title("Incoming transfers")
            nd.geometry("580x400"); nd.configure(bg=BG); nd.grab_set(); nd.resizable(False,True)
            hdr_n=tk.Frame(nd,bg=NAV,padx=20,pady=14); hdr_n.pack(fill="x")
            tk.Label(hdr_n,text="📨  Incoming transfers",
                     font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
            tk.Label(hdr_n,text=f"{len(unread)} quer{'y' if len(unread)==1 else 'ies'} transferred to you",
                     font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            tk.Frame(nd,bg=ACCENT,height=2).pack(fill="x")

            inner,_=scrollable_frame(tk.Frame(nd,bg=BG)); inner_outer=nd

            for row_num,msg in unread:
                # Parse message parts
                parts={p.split(": ",1)[0].strip():p.split(": ",1)[1].strip()
                       for p in msg.replace("📨 TRANSFER from ","Sender: ").split("  |  ") if ": " in p}

                def part(*keys):
                    for k in keys:
                        if k in parts and parts[k]:
                            return parts[k]
                    return ""

                card=tk.Frame(inner,bg=CARD2,highlightthickness=1,
                              highlightbackground=ACCENT,padx=16,pady=12)
                card.pack(fill="x",pady=(0,8))
                top_row=tk.Frame(card,bg=CARD2); top_row.pack(fill="x")
                tk.Label(top_row,text=part("Ref","ref") or msg[:40],
                         font=(FONT,11,"bold"),bg=CARD2,fg=ACCENT2).pack(side="left")
                tk.Label(top_row,text=f"  from {part('Sender','sender') or 'unknown'}",
                         font=(FONT,9),bg=CARD2,fg=MUTED).pack(side="left")
                for lbl,key in [
                    ("Client",part("Client","client")),
                    ("Site",part("Site","site")),
                    ("Type",part("Type","type")),
                    ("Query",part("Query","query")),
                ]:
                    if key:
                        tk.Label(card,text=f"{lbl}: {key}",font=(FONT,9),
                                 bg=CARD2,fg=TEXT2).pack(anchor="w")

            ft_n=tk.Frame(nd,bg=CARD2,padx=16,pady=12); ft_n.pack(fill="x",side="bottom")
            tk.Label(ft_n,text="These queries are now in your tracker.",
                     font=(FONT,9),bg=CARD2,fg=MUTED).pack(side="left")

            def mark_read():
                try:
                    wb2=openpyxl.load_workbook(self.excel_file)
                    if "_Notifications" in wb2.sheetnames:
                        nws2=wb2["_Notifications"]
                        for rn,_ in unread:
                            nws2.cell(row=rn,column=2,value="Y")
                        wb2.save(self.excel_file)
                        self._excel_mtime=self._excel_mtime_now()
                except: pass
                nd.destroy()
                self._refresh_table()

            make_btn(ft_n,"Got it — mark as read",mark_read,"primary",padx=14,pady=6).pack(side="right")
        except Exception:
            pass  # notification check is non-fatal
        today=today_str()
        items=[q for q in self.queries if q["status"]!="Resolved" and q.get("chase_date","") and q["chase_date"]<=today]
        if items:
            preview=", ".join(f"{q['ref']} ({q['client']})" for q in items[:3])
            extra=f" +{len(items)-3} more" if len(items)>3 else ""
            self.banner_lbl.config(text=f"  ⚠  {len(items)} quer{'y' if len(items)==1 else 'ies'} need actioning today:  {preview}{extra}")
            self.banner_frame.pack(fill="x",padx=20,pady=(12,0))
        else: self.banner_frame.pack_forget()

    def _go_list(self,tab="all"):
        self._clear_filters(refresh=False)
        self._show_page("list")
        self._set_tab(tab)

    def _open_day_in_list(self, day_str):
        self._clear_filters(refresh=False)
        self._calendar_day_filter=day_str
        self._show_page("list")
        self._set_tab("open")

    def _action_date_block_reason(self, day_str):
        if is_weekend(day_str):
            return f"{fmt_date(day_str)} is a weekend. Please choose a working day."
        if is_bank_holiday(day_str):
            return f"{fmt_date(day_str)} is a UK bank holiday. Please choose a working day."
        return ""

    def _validate_action_date(self, day_str, parent=None):
        d=parse_iso_date(day_str)
        if not d:
            messagebox.showwarning("Invalid date", "Please enter a valid action date (YYYY-MM-DD).", parent=parent or self)
            return False
        msg=self._action_date_block_reason(day_str)
        if msg:
            messagebox.showwarning("Date not allowed", msg, parent=parent or self)
            return False
        return True

    def _day_workload(self, day_str, exclude_ids=None):
        if not parse_iso_date(day_str): return 0
        exclude_ids = set(exclude_ids or [])
        return sum(
            1 for q in self.queries
            if q.get("status") != "Resolved"
            and q.get("chase_date", "") == day_str
            and q.get("id") not in exclude_ids
        )

    def _get_high_volume_threshold(self):
        try:
            return max(1, int(self.cfg.get("high_volume_threshold", HIGH_VOLUME_DAY_THRESHOLD)))
        except:
            return HIGH_VOLUME_DAY_THRESHOLD

    def _confirm_high_volume_day(self, day_str, exclude_ids=None, add_count=1, parent=None):
        if not parse_iso_date(day_str):
            return True
        threshold = self._get_high_volume_threshold()
        base_load = self._day_workload(day_str, exclude_ids=exclude_ids)
        projected = base_load + max(0, int(add_count))
        if projected < threshold:
            return True
        pretty = fmt_date(day_str)
        return messagebox.askyesno(
            "High workload day",
            f"{pretty} already has {base_load} open quer{'y' if base_load == 1 else 'ies'}.\n"
            f"Adding this change would make {projected}.\n\n"
            "Do you still want to schedule it for this day?",
            parent=parent or self,
        )

    def _confirm_pushback_history(self, query, old_chase, new_chase, note_added=False, parent=None):
        if not is_pushback(old_chase, new_chase, query.get("last_date", "")):
            return True
        prior = pushback_count(query)
        if prior < 2 or note_added:
            return True
        return messagebox.askyesno(
            "Already pushed back",
            f"{query.get('ref','This query')} has already been pushed back {prior} times.\n"
            "No additional note was detected.\n\n"
            "Continue anyway?",
            parent=parent or self,
        )

    def _clear_filters(self,refresh=True):
        self.search_var.set("")
        self.filter_client.set("All"); self.filter_fund.set("All")
        self.filter_type.set("All"); self.filter_status.set("All")
        try: self.filter_assignee.set("All")
        except: pass
        try: self.filter_utility.set("All")
        except: pass
        self._assignee_filter=""  # also clear person filter
        self._calendar_day_filter=""
        if refresh: self._refresh_table()

    def _open_bulk_actions(self):
        """Open bulk actions dialog for selected queries."""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("No Selection", "Please select one or more queries first.", parent=self)
            return
        
        # Get the query IDs from selected items (use Treeview iid for reliability)
        selected_ids = []
        for item in selected_items:
            qid = item
            if qid:
                query = next((q for q in self.queries if q['id'] == qid), None)
                if query:
                    selected_ids.append(qid)

        if not selected_ids:
            messagebox.showerror("Error", "Could not find selected queries.", parent=self)
            return

        if len(selected_ids) != len(selected_items):
            messagebox.showwarning("Partial selection", 
                f"{len(selected_ids)} of {len(selected_items)} selected queries were found and will be updated.",
                parent=self)
        
        # Create bulk actions dialog
        dlg = tk.Toplevel(self)
        dlg.title(f"Bulk Actions ({len(selected_ids)} queries)")
        dlg.geometry("520x600")
        dlg.configure(bg=BG)
        dlg.grab_set()
        dlg.resizable(False, True)
        
        hdr = tk.Frame(dlg, bg=NAV, padx=20, pady=16)
        hdr.pack(fill="x")
        tk.Frame(dlg, bg=ACCENT, height=2).pack(fill="x")
        tk.Label(hdr, text="⚡ Bulk Actions", font=(FONT, 13, "bold"), bg=NAV, fg="white").pack(anchor="w")
        tk.Label(hdr, text=f"Apply changes to {len(selected_ids)} selected queries", font=(FONT, 9), bg=NAV, fg=TEXT2).pack(anchor="w", pady=(4, 0))
        
        body = tk.Frame(dlg, bg=BG, padx=24, pady=20)
        body.pack(fill="both", expand=True)
        
        # Action type selection
        action_var = tk.StringVar(value="assign")
        tk.Label(body, text="Select action to perform:", font=(FONT, 11, "bold"), bg=BG, fg=TEXT).pack(anchor="w", pady=(0, 12))
        
        actions_frame = tk.Frame(body, bg=CARD, highlightthickness=1, highlightbackground=BORDER, padx=12, pady=12)
        actions_frame.pack(fill="x", pady=(0, 20))
        
        for action, label, desc in [
            ("assign", "Assign to team member", "Reassign selected queries to a team member"),
            ("priority", "Change priority", "Update priority level for selected queries"),
            ("status", "Change status", "Update status for selected queries"),
            ("chase", "Set action date", "Update next action date for selected queries")
        ]:
            rb = tk.Radiobutton(actions_frame, text=label, variable=action_var, value=action, 
                              font=(FONT, 9), bg=CARD, fg=TEXT, selectcolor=CARD, activebackground=CARD)
            rb.pack(anchor="w", pady=2)
            tk.Label(actions_frame, text=f"  {desc}", font=(FONT, 7), bg=CARD, fg=MUTED).pack(anchor="w", pady=(0, 4))
        
        # Action parameters
        params_label = tk.Label(body, text="Parameters:", font=(FONT, 10, "bold"), bg=BG, fg=TEXT)
        params_label.pack(anchor="w", pady=(12, 8))
        
        params_frame = tk.Frame(body, bg=CARD2, highlightthickness=1, highlightbackground=ACCENT, padx=16, pady=16)
        params_frame.pack(fill="x", pady=(0, 20))
        
        # Assignment
        assign_frame = tk.Frame(params_frame, bg=CARD2)
        tk.Label(assign_frame, text="Assign to:", font=(FONT, 9), bg=CARD2, fg=TEXT2).pack(side="left")
        all_members = list(dict.fromkeys([self.username] + self.team_members))
        assign_var = tk.StringVar()
        make_combo(assign_frame, assign_var, ["(Unassigned)"] + all_members, readonly=False, width=20).pack(side="left", padx=(8, 0))
        assign_frame.pack(fill="x", pady=(0, 0))  # Pack initially
        
        # Priority
        priority_frame = tk.Frame(params_frame, bg=CARD2)
        tk.Label(priority_frame, text="Priority:", font=(FONT, 9), bg=CARD2, fg=TEXT2).pack(side="left")
        priority_var = tk.StringVar(value="Medium")
        make_combo(priority_frame, priority_var, PRIORITIES, readonly=True, width=15).pack(side="left", padx=(8, 0))
        
        # Status
        status_frame = tk.Frame(params_frame, bg=CARD2)
        tk.Label(status_frame, text="Status:", font=(FONT, 9), bg=CARD2, fg=TEXT2).pack(side="left")
        status_var = tk.StringVar(value="Open")
        make_combo(status_frame, status_var, STATUSES, readonly=True, width=15).pack(side="left", padx=(8, 0))
        
        # Chase date
        chase_frame = tk.Frame(params_frame, bg=CARD2)
        tk.Label(chase_frame, text="Action date:", font=(FONT, 9), bg=CARD2, fg=TEXT2).pack(side="left")
        chase_var = tk.StringVar(value=(date.today() + timedelta(days=7)).strftime("%Y-%m-%d"))
        chase_entry_frame = tk.Frame(chase_frame, bg=CARD2, highlightthickness=1, highlightbackground=BORDER)
        chase_entry_frame.pack(side="left", padx=(8, 4))
        tk.Entry(chase_entry_frame, textvariable=chase_var, font=(FONT, 10), bg=CARD2, fg=TEXT, 
                insertbackground=TEXT, relief="flat", bd=5, width=12, highlightthickness=0).pack()
        make_btn(
            chase_frame,
            "📅",
            lambda: _show_cal(
                dlg,
                chase_var,
                get_day_load=lambda d: self._day_workload(d, exclude_ids=selected_ids),
                confirm_day_selection=lambda d: self._confirm_high_volume_day(
                    d, exclude_ids=selected_ids, add_count=len(selected_ids), parent=dlg
                ),
                high_volume_threshold=self._get_high_volume_threshold(),
                date_block_reason=self._action_date_block_reason,
            ),
            "default",
            padx=6,
            pady=3,
        ).pack(side="left")
        
        def update_params_visibility(a=None, b=None, c=None):
            # Hide all parameter frames
            assign_frame.pack_forget()
            priority_frame.pack_forget()
            status_frame.pack_forget()
            chase_frame.pack_forget()
            
            # Show relevant frame based on action
            action = action_var.get()
            if action == "assign":
                assign_frame.pack(fill="x", pady=(0, 0))
            elif action == "priority":
                priority_frame.pack(fill="x", pady=(0, 0))
            elif action == "status":
                status_frame.pack(fill="x", pady=(0, 0))
            elif action == "chase":
                chase_frame.pack(fill="x", pady=(0, 0))
        
        # Bind action changes to visibility update
        for action_val in ["assign", "priority", "status", "chase"]:
            action_var.trace_add("write", update_params_visibility)
        
        # Show initial action (assign)
        self.after(100, update_params_visibility)
        
        # Buttons
        btn_frame = tk.Frame(dlg, bg=CARD2, highlightthickness=1, highlightbackground=ACCENT, padx=16, pady=14)
        btn_frame.pack(fill="x", side="bottom")
        
        def apply_bulk_action():
            action = action_var.get()
            updated_count = 0

            if action == "chase":
                new_chase = chase_var.get().strip()
                if not self._validate_action_date(new_chase, parent=dlg):
                    return
                if not self._confirm_high_volume_day(
                    new_chase, exclude_ids=selected_ids, add_count=len(selected_ids), parent=dlg
                ):
                    return
                already_pushed = [
                    q for q in self.queries
                    if q.get("id") in selected_ids and pushback_count(q) >= 2
                ]
                if already_pushed:
                    if not messagebox.askyesno(
                        "Already pushed back",
                        f"{len(already_pushed)} selected quer{'y has' if len(already_pushed)==1 else 'ies have'} "
                        "already been pushed back at least twice.\n"
                        "No bulk note can be added from this action.\n\n"
                        "Continue anyway?",
                        parent=dlg,
                    ):
                        return
            
            for qid in selected_ids:
                query = next((q for q in self.queries if q['id'] == qid), None)
                if not query:
                    continue
                
                if action == "assign":
                    new_assignee = assign_var.get()
                    if new_assignee == "(Unassigned)":
                        new_assignee = ""
                    if query.get("assigned_to", "") != new_assignee:
                        query["assigned_to"] = new_assignee
                        query["last_by"] = self.username
                        query["last_date"] = today_str()
                        query["log"] += f" | {stamp(self.username)} Assigned to {new_assignee or 'unassigned'}"
                        updated_count += 1
                        
                elif action == "priority":
                    new_priority = priority_var.get()
                    if query.get("priority", "Medium") != new_priority:
                        query["priority"] = new_priority
                        query["last_by"] = self.username
                        query["last_date"] = today_str()
                        query["log"] += f" | {stamp(self.username)} Priority changed to {new_priority}"
                        updated_count += 1
                        
                elif action == "status":
                    new_status = status_var.get()
                    if query.get("status", "Open") != new_status:
                        query["status"] = new_status
                        query["last_by"] = self.username
                        query["last_date"] = today_str()
                        if new_status == "Resolved":
                            query["resolved_date"] = today_str()
                        query["log"] += f" | {stamp(self.username)} Status changed to {new_status}"
                        updated_count += 1
                        
                elif action == "chase":
                    new_chase = chase_var.get().strip()
                    if query.get("chase_date", "") != new_chase:
                        old_chase = query.get("chase_date", "")
                        prev_last_date = query.get("last_date", "")
                        query["chase_date"] = new_chase
                        query["last_by"] = self.username
                        query["last_date"] = today_str()
                        if is_pushback(old_chase, new_chase, prev_last_date):
                            append_pushback_event(query, self.username, old_chase, new_chase)
                        query["log"] += f" | {stamp(self.username)} Action date set to {new_chase}"
                        updated_count += 1
            
            if updated_count > 0:
                self._save_queries()
                self._refresh_table()
                self._refresh_dashboard()
                self._show_daily_banner()
                messagebox.showinfo("Success", f"✓ Updated {updated_count} quer{'y' if updated_count == 1 else 'ies'} successfully.", parent=dlg)
                dlg.destroy()
            else:
                messagebox.showwarning("No changes", "No queries were updated (no changes made).", parent=dlg)
        
        make_btn(btn_frame, "Cancel", dlg.destroy, "default", padx=12, pady=6).pack(side="right", padx=(8, 0))
        make_btn(btn_frame, "Apply Changes", apply_bulk_action, "primary", padx=14, pady=6).pack(side="right")

    def _sort_by(self,col):
        asc=not self._sort_state.get(col,True)
        self._sort_state={c:True for c in self._sort_state}
        self._sort_state[col]=asc
        col_labels={"ref":"Reference","client":"Client","fund":"Fund","site":"Site","utility":"Utility",
                    "meter":"Meter","type":"Type","status":"Status","priority":"Priority",
                    "overdue":"Overdue","chase":"Next Action","raised":"Raised",
                    "opened":"Logged","assigned":"Assigned to","last_by":"Last updated by","att":"📎"}
        for c,lbl in col_labels.items():
            arrow=(" ▲" if asc else " ▼") if c==col else ""
            self.tree.heading(c,text=lbl+arrow)
        items=[(self.tree.set(iid,col),iid) for iid in self.tree.get_children()]
        items.sort(reverse=not asc)
        for idx,(_,iid) in enumerate(items): self.tree.move(iid,"",idx)

    def _save_queries(self):
        """Save queries to Excel and mark dashboard/reports as needing rebuild."""
        try:
            save_all_queries(self.queries,self.excel_file)
        except PermissionError:
            messagebox.showerror(
                "Save failed",
                "Could not save the tracker because the Excel file is open. Close it and try again.",
                parent=self,
            )
            return False
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc), parent=self)
            return False
        self._excel_mtime=self._excel_mtime_now()
        self._dash_dirty=True
        self._rpt_dirty=True
        return True

    def _reload_sites(self):
        self.clients,self.sites_by_client,self.meters,self.utilities_by_site,\
            self.funds_by_client,self.sites_by_fund=load_site_data(self.sites_file)
        self.filter_client.configure(values=["All"]+self.clients)
        self.filter_fund.configure(values=["All"]); self.filter_fund.set("All")
        messagebox.showinfo("Reloaded",f"{len(self.clients)} clients · {sum(len(v) for v in self.sites_by_client.values())} sites loaded.")

    def _open_settings(self):
        def on_complete(cfg):
            self.cfg=cfg; self.username=cfg.get("username","Unknown")
            self.excel_file=cfg.get("excel_file",""); self.sites_file=cfg.get("sites_file","")
            raw=cfg.get("team_members",[])
            self.team_members=raw if raw else [self.username]
            if self.username not in self.team_members:
                self.team_members=[self.username]+self.team_members
            self.linked_trackers=cfg.get("linked_trackers",[])
            self.queries=load_queries(self.excel_file); self._reload_sites()
            self._excel_mtime=self._excel_mtime_now()
            self._refresh_table(); self._show_daily_banner()
            self._refresh_dashboard()
        SetupWizard(self,existing_cfg=self.cfg,on_complete=on_complete)

    def _labeled_combo(self,parent,label,var,values,readonly=False):
        f=tk.Frame(parent,bg=BG); f.pack(fill="x",pady=4)
        tk.Label(f,text=label,font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
        cb=make_combo(f,var,values,readonly=readonly,width=34)
        cb.pack(side="left",fill="x",expand=True); return cb

    def _labeled_entry(self,parent,label,var):
        f=tk.Frame(parent,bg=BG); f.pack(fill="x",pady=4)
        tk.Label(f,text=label,font=(FONT,9),bg=BG,fg=TEXT2,width=22,anchor="w").pack(side="left")
        card=tk.Frame(f,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); card.pack(side="left",fill="x",expand=True)
        tk.Entry(card,textvariable=var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,highlightthickness=0).pack(fill="x")
        return card

    def _bind_autocomplete(self,cb,get_full_list,on_select=None):
        """Autocomplete that filters the dropdown list while typing without hijacking focus."""
        def on_keyrelease(e):
            if e.keysym in ("Return","Tab","Escape","Down","Up","Left","Right"): return
            typed=cb.get().strip().lower(); full=get_full_list()
            if typed:
                # Prioritise prefix matches so typing the first letter jumps quickly.
                starts=[v for v in full if v.lower().startswith(typed)]
                contains=[v for v in full if typed in v.lower() and not v.lower().startswith(typed)]
                filtered=starts+contains
            else:
                filtered=full
            cb.configure(values=filtered)
            # Only open dropdown if there are matches and user typed something
            if typed and filtered:
                try:
                    cb.tk.call("ttk::combobox::Post",cb)
                    # Restore cursor position after posting
                    cb.icursor("end")
                except: pass
        def on_focusout(e):
            # Restore full list so dropdown isn't permanently filtered
            cb.configure(values=get_full_list())
            if cb.get() in get_full_list() and on_select: on_select()
        def on_selected(e):
            cb.configure(values=get_full_list())
            if on_select: on_select()
        cb.bind("<KeyRelease>",on_keyrelease)
        cb.bind("<<ComboboxSelected>>",on_selected)
        cb.bind("<FocusOut>",on_focusout)

    def _open_add_dialog(self,preset_type=None,copy_from=None,edit_query=None):
        # edit_query = existing query dict to edit in place
        # copy_from  = existing query dict to duplicate
        is_edit = edit_query is not None
        cf = edit_query if is_edit else (copy_from or {})

        dlg=tk.Toplevel(self)
        dlg.title(f"Edit {cf['ref']}" if is_edit else "New Query")
        dlg.geometry("640x860"); dlg.configure(bg=BG); dlg.grab_set(); dlg.resizable(False,True)
        hdr=tk.Frame(dlg,bg=NAV,padx=24,pady=16); hdr.pack(fill="x")
        tk.Label(hdr,text=f"✎  Edit {cf['ref']}" if is_edit else "＋  Log new query",
                 font=(FONT,13,"bold"),bg=NAV,fg="white").pack(anchor="w")
        if is_edit:
            tk.Label(hdr,text="Changes will overwrite this query.",font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
        elif copy_from:
            tk.Label(hdr,text=f"Copied from {copy_from['ref']}",font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))

        outer=tk.Frame(dlg,bg=BG); outer.pack(fill="both",expand=True)
        form,_=scrollable_frame(outer)

        # cf is already set above from edit_query or copy_from
        client_var=tk.StringVar(value=cf.get("client","")); fund_var=tk.StringVar(value=cf.get("fund",""))
        site_var=tk.StringVar(value=cf.get("site","")); utility_var=tk.StringVar(value=cf.get("utility",""))
        meter_var=tk.StringVar(value=cf.get("meter","") if (is_edit or copy_from) else "")
        type_var=tk.StringVar(value=preset_type or cf.get("type",""))
        status_var=tk.StringVar(value=cf.get("status","Open") if is_edit else "Open")
        priority_var=tk.StringVar(value=cf.get("priority","Medium"))
        ref_var=tk.StringVar(value=cf.get("ref","") if is_edit else "")
        # Chase date: use existing for edits, otherwise default to +7 days
        _default_chase=(date.today()+timedelta(days=7)).strftime("%Y-%m-%d")
        chase_var=tk.StringVar(value=cf.get("chase_date",_default_chase) if is_edit else _default_chase)
        prop_var=tk.StringVar(value=cf.get("prop_code","")); address_var=tk.StringVar(value=cf.get("address",""))
        spid_var=tk.StringVar(value=cf.get("spid","")); serial_var=tk.StringVar(value=cf.get("serial",""))
        contact_var=tk.StringVar(value=cf.get("contact",""))
        raised_var=tk.StringVar(value=cf.get("raised_date",""))
        is_type_changed_query = bool(re.search(r"Query created from type change", cf.get("log", "")))

        tk.Label(form,text="QUERY DETAILS",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,8))

        def open_new_site_dialog():
            """Dedicated dialog to add a new client/site — saves to sites.xlsx and pre-fills the form."""
            ns=tk.Toplevel(dlg); ns.title("Add new site")
            ns.geometry("520x640"); ns.configure(bg=BG); ns.grab_set(); ns.resizable(False,True)
            ns.minsize(520,500)
            hdr_ns=tk.Frame(ns,bg=NAV,padx=20,pady=14); hdr_ns.pack(fill="x")
            tk.Label(hdr_ns,text="＋  Add new site",font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
            tk.Label(hdr_ns,text="Fill in the details — this will be saved to your site list.",
                     font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            tk.Frame(ns,bg=ACCENT,height=2).pack(fill="x")

            body_ns=tk.Frame(ns,bg=BG,padx=24,pady=16); body_ns.pack(fill="both",expand=True)

            ns_client=tk.StringVar(value=client_var.get())
            ns_fund=tk.StringVar(value=fund_var.get())
            ns_site=tk.StringVar(value=site_var.get())
            ns_prop=tk.StringVar(value=prop_var.get())
            ns_addr=tk.StringVar(value=address_var.get())
            ns_util=tk.StringVar(value=utility_var.get())
            ns_spid=tk.StringVar(value=spid_var.get())
            ns_serial=tk.StringVar(value=serial_var.get())
            ns_contact=tk.StringVar(value=contact_var.get())

            def ns_row(lbl,var,opts=None,required=False):
                f=tk.Frame(body_ns,bg=BG); f.pack(fill="x",pady=4)
                star=" *" if required else ""
                tk.Label(f,text=lbl+star,font=(FONT,9),bg=BG,fg=MUTED,width=20,anchor="w").pack(side="left")
                if opts is not None:
                    cb2=make_combo(f,var,opts,readonly=False,width=28)
                    cb2.pack(side="left",fill="x",expand=True)
                    return cb2
                else:
                    fc=tk.Frame(f,bg=CARD,highlightthickness=1,highlightbackground=BORDER)
                    fc.pack(side="left",fill="x",expand=True)
                    tk.Entry(fc,textvariable=var,font=(FONT,10),bg=CARD,fg=TEXT,
                             insertbackground=TEXT,relief="flat",bd=6,highlightthickness=0).pack(fill="x")

            tk.Label(body_ns,text="REQUIRED",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,4))
            ns_row("Client",ns_client,required=True)
            ns_row("Site name",ns_site,required=True)
            tk.Label(body_ns,text="OPTIONAL",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(10,4))
            ns_row("Fund",ns_fund)
            ns_row("Property code",ns_prop)
            ns_row("Address",ns_addr)
            ns_row("Utility type",ns_util,opts=UTILITY_OPTIONS)
            ns_row("Supply point ID",ns_spid)
            ns_row("Meter serial",ns_serial)
            ns_row("Managing agent contact",ns_contact)

            def save_new_site():
                client_val=ns_client.get().strip()
                site_val=ns_site.get().strip()
                if not client_val:
                    messagebox.showwarning("Required","Please enter a client name.",parent=ns); return
                if not site_val:
                    messagebox.showwarning("Required","Please enter a site name.",parent=ns); return
                if not self.sites_file or not os.path.exists(self.sites_file):
                    messagebox.showerror("No site file",
                        "Cannot find your sites.xlsx.\nCheck the path in Settings.",parent=ns); return
                try:
                    swb=openpyxl.load_workbook(self.sites_file)
                    sws=swb["Sites"] if "Sites" in swb.sheetnames else swb.active
                    sws.append([
                        client_val, ns_fund.get().strip(), ns_prop.get().strip(),
                        site_val,   ns_addr.get().strip(), ns_util.get().strip(),
                        ns_spid.get().strip(), ns_serial.get().strip(), ns_contact.get().strip()
                    ])
                    swb.save(self.sites_file)
                except PermissionError:
                    messagebox.showerror("File in use",
                        "sites.xlsx is open in Excel — close it and try again.",parent=ns); return
                except Exception as e:
                    messagebox.showerror("Error",str(e),parent=ns); return

                # Reload site data into app
                self.clients,self.sites_by_client,self.meters,self.utilities_by_site,\
                    self.funds_by_client,self.sites_by_fund=load_site_data(self.sites_file)
                # Update client dropdown
                client_cb.configure(values=self.clients)
                # Pre-fill the query form
                client_var.set(client_val)
                site_cb.configure(values=self.sites_by_client.get(client_val,[]))
                site_var.set(site_val)
                fund_cb.configure(values=self.funds_by_client.get(client_val,[]))
                if ns_fund.get().strip(): fund_var.set(ns_fund.get().strip())
                if ns_addr.get().strip(): address_var.set(ns_addr.get().strip())
                if ns_prop.get().strip(): prop_var.set(ns_prop.get().strip())
                if ns_spid.get().strip(): spid_var.set(ns_spid.get().strip())
                if ns_serial.get().strip(): serial_var.set(ns_serial.get().strip())
                if ns_contact.get().strip(): contact_var.set(ns_contact.get().strip())
                if ns_util.get().strip():
                    utility_var.set(ns_util.get().strip())
                    utility_cb.configure(values=[ns_util.get().strip()])
                ns.destroy()
                messagebox.showinfo("Site added",
                    f"'{site_val}' added to your site list and pre-filled in the form.",parent=dlg)

            # Footer packed after save_new_site is defined — fixes UnboundLocalError
            ft_ns=tk.Frame(ns,bg=CARD2,padx=16,pady=12); ft_ns.pack(fill="x",side="bottom")
            tk.Frame(ns,bg=BORDER,height=1).pack(fill="x",side="bottom")
            make_btn(ft_ns,"Cancel",ns.destroy,"default",padx=12,pady=6).pack(side="right",padx=(6,0))
            make_btn(ft_ns,"Save & apply",save_new_site,"primary",padx=14,pady=6).pack(side="right")
        ns_btn_row=tk.Frame(form,bg=BG); ns_btn_row.pack(fill="x",pady=(0,6))
        make_btn(ns_btn_row,"＋ New site / client",open_new_site_dialog,"active",padx=12,pady=5).pack(side="left")
        tk.Label(ns_btn_row,text="Add a new client or site to your site list",
                 font=(FONT,8),bg=BG,fg=MUTED).pack(side="left",padx=(10,0))

        # Client dropdown
        client_row=tk.Frame(form,bg=BG); client_row.pack(fill="x",pady=4)
        tk.Label(client_row,text="Client *",font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
        client_cb=make_combo(client_row,client_var,self.clients,readonly=False,width=34)
        client_cb.pack(side="left",fill="x",expand=True)

        site_cb  =self._labeled_combo(form,"Site *",site_var,[],readonly=False)
        fund_cb  =self._labeled_combo(form,"Fund",fund_var,[],readonly=False)

        utility_cb=self._labeled_combo(form,"Utility type *",utility_var,[],readonly=True)
        meter_cb =self._labeled_combo(form,"Specific meter",meter_var,[])
        type_cb  =self._labeled_combo(form,"Query type *",type_var,QUERY_TYPES,readonly=True)
        self._labeled_combo(form,"Status",status_var,STATUSES,readonly=True)
        self._labeled_combo(form,"Priority",priority_var,PRIORITIES,readonly=True)
        self._labeled_entry(form,"Reference (auto-gen)",ref_var)

        # ── Assigned to ───────────────────────────────────────────────────────
        assignee_var=tk.StringVar(value=cf.get("assigned_to",""))
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        assignee_row=tk.Frame(form,bg=BG); assignee_row.pack(fill="x",pady=4)
        tk.Label(assignee_row,text="Assigned to",font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
        assignee_cb=make_combo(assignee_row,assignee_var,["(Unassigned)"]+all_members,readonly=False,width=34)
        assignee_cb.pack(side="left",fill="x",expand=True)

        # ── Query raised date ──────────────────────────────────────────────────
        if not (is_edit and is_type_changed_query):
            raised_frame=tk.Frame(form,bg=BG); raised_frame.pack(fill="x",pady=4)
            tk.Label(raised_frame,text="Query raised by client",font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
            rc=tk.Frame(raised_frame,bg=CARD,highlightthickness=1,highlightbackground=BORDER); rc.pack(side="left",fill="x",expand=True)
            tk.Entry(rc,textvariable=raised_var,font=(FONT,10),bg=CARD,fg=TEXT,relief="flat",bd=6,
                     highlightthickness=0).pack(fill="x",side="left",expand=True)
            make_btn(raised_frame,"📅",lambda:_show_cal(dlg,raised_var),"default",padx=8,pady=4).pack(side="left",padx=(4,0))
            tk.Label(raised_frame,text="When the client emailed / raised this",
                     font=(FONT,8),bg=BG,fg=MUTED).pack(side="left",padx=(8,0))

        chase_frame=tk.Frame(form,bg=BG); chase_frame.pack(fill="x",pady=4)
        tk.Label(chase_frame,text="Action date",font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
        cc=tk.Frame(chase_frame,bg=CARD,highlightthickness=1,highlightbackground=BORDER); cc.pack(side="left",fill="x",expand=True)
        tk.Entry(cc,textvariable=chase_var,font=(FONT,10),bg=CARD,fg=TEXT,relief="flat",bd=6,highlightthickness=0).pack(fill="x",side="left",expand=True)
        weekend_lbl=tk.Label(cc,text="⚠ Non-working day",font=(FONT,8),bg=CARD,fg=WARNING)
        def check_weekend(*_):
            if self._action_date_block_reason(chase_var.get()): weekend_lbl.pack(side="right",padx=6)
            else: weekend_lbl.pack_forget()
        chase_var.trace_add("write",check_weekend)
        for d,lbl in [(7,"+7d"),(14,"+14d")]:
            make_btn(chase_frame,lbl,lambda days=d:chase_var.set((date.today()+timedelta(days=days)).strftime("%Y-%m-%d")),
                     "default",padx=8,pady=4).pack(side="left",padx=(6,0))
        make_btn(
            chase_frame,
            "📅",
            lambda: _show_cal(
                dlg,
                chase_var,
                get_day_load=lambda d: self._day_workload(d, exclude_ids=[edit_query.get("id")] if is_edit else None),
                confirm_day_selection=lambda d: self._confirm_high_volume_day(
                    d,
                    exclude_ids=[edit_query.get("id")] if is_edit else None,
                    add_count=1,
                    parent=dlg,
                ),
                high_volume_threshold=self._get_high_volume_threshold(),
                date_block_reason=self._action_date_block_reason,
            ),
            "default",
            padx=8,
            pady=4,
        ).pack(side="left",padx=(4,0))

        tk.Label(form,text="Description *",font=(FONT,9),bg=BG,fg=MUTED).pack(anchor="w",pady=(8,4))
        dc=tk.Frame(form,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); dc.pack(fill="x")
        desc_text=tk.Text(dc,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=8,height=4,wrap="word",highlightthickness=0)
        desc_text.pack(fill="x")
        if cf.get("desc"): desc_text.insert("1.0",cf["desc"])
        attach_text_spellcheck(desc_text)

        desc_text.bind("<MouseWheel>", lambda e: "break")

        # Drag-to-resize handle
        resize_bar=tk.Frame(form,bg=BORDER,height=6,cursor="sb_v_double_arrow")
        resize_bar.pack(fill="x",pady=(0,2))
        _drag={"start_y":0,"start_h":4}
        def _resize_start(e):
            _drag["start_y"]=e.y_root
            _drag["start_h"]=int(desc_text.cget("height"))

        def _resize_drag(e):
            total_dy = e.y_root - _drag["start_y"]
            # Make resizing responsive: 1 row per ~8px movement.
            delta = int(round(total_dy / 8.0))
            new_h = max(3, min(24, _drag["start_h"] + delta))
            desc_text.configure(height=new_h)

        resize_bar.bind("<ButtonPress-1>",_resize_start)
        resize_bar.bind("<B1-Motion>",_resize_drag)
        tk.Label(form,text="↕ drag to resize",font=(FONT,7),bg=BG,fg=MUTED).pack(anchor="e",pady=(0,4))

        divider(form)
        tk.Label(form,text="SITE DETAILS",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,8))

        prop_frame=tk.Frame(form,bg=BG)
        pf=tk.Frame(prop_frame,bg=BG); pf.pack(fill="x")
        tk.Label(pf,text="Property code",font=(FONT,9),bg=BG,fg=MUTED,width=22,anchor="w").pack(side="left")
        pc=tk.Frame(pf,bg=CARD,highlightthickness=1,highlightbackground=BORDER); pc.pack(side="left",fill="x",expand=True)
        tk.Entry(pc,textvariable=prop_var,font=(FONT,10),bg=CARD,fg=TEXT,relief="flat",bd=6,highlightthickness=0).pack(fill="x")

        self._labeled_entry(form,"Site address",address_var)
        self._labeled_entry(form,"Supply point ID",spid_var)
        self._labeled_entry(form,"Meter serial",serial_var)
        self._labeled_entry(form,"Managing agent contact",contact_var)

        nm_frame=tk.Frame(form,bg=BG); nm_frame.pack(fill="x",pady=(10,0))

        def get_site_list(): return self.sites_by_client.get(client_var.get(),[])
        def get_meter_list():
            rows=[r for r in self.meters.get((client_var.get(),site_var.get()),[]) if r["utility"]==utility_var.get()]
            return [f"{r['serial']} ({r['spid']})" if r["serial"] and r["spid"] else r["serial"] or r["spid"] for r in rows]

        nm_btn=None
        related_site_lbl=tk.Label(form,text="",font=(FONT,8),bg=BG,fg=WARNING,justify="left",wraplength=560)
        related_site_lbl.pack(fill="x",pady=(0,4))
        warned_site=[""]

        def on_client(*_):
            c=client_var.get()
            if c not in self.clients:
                # New client — allow free typing everywhere
                site_cb.configure(values=[])
                fund_cb.configure(state="normal")
                return
            # Populate site list immediately — this is what was missing
            sites=self.sites_by_client.get(c,[])
            site_cb.configure(values=sites)
            # Clear dependent fields
            for v in [site_var,fund_var,utility_var,meter_var,
                      prop_var,address_var,spid_var,serial_var,contact_var]:
                v.set("")
            utility_cb.configure(values=[])
            meter_cb.configure(values=[])
            # Populate fund list for this client
            funds=self.funds_by_client.get(c,[])
            fund_cb.configure(values=funds,state="normal")
            if len(funds)==1: fund_var.set(funds[0])

        def on_fund(*_):
            c=client_var.get(); f=fund_var.get()
            # Filter site list by fund if one is selected
            if f:
                sites=self.sites_by_fund.get((c,f),[])
            else:
                sites=self.sites_by_client.get(c,[])
            site_cb.configure(values=sites)
            # Only reset site if it no longer fits the fund filter
            if site_var.get() and site_var.get() not in sites:
                for v in [site_var,utility_var,meter_var,
                          prop_var,address_var,spid_var,serial_var,contact_var]:
                    v.set("")
                utility_cb.configure(values=[])

        def on_site(*_):
            c=client_var.get(); s=site_var.get(); key=(c,s)
            is_new_site = s and s not in self.sites_by_client.get(c,[])
            if is_new_site:
                utility_cb.configure(values=UTILITY_OPTIONS,state="normal")
                prop_frame.pack(fill="x",pady=4)
                if nm_btn is not None: nm_btn.pack_forget()
                return
            if not s:
                if nm_btn is not None: nm_btn.pack_forget()
                return
            rows=self.meters.get(key,[])
            # Fill fund from site data
            if rows and rows[0].get("fund"):
                fund_var.set(rows[0].get("fund",""))
                fund_cb.configure(values=self.funds_by_client.get(c,[]),state="normal")
            utils=self.utilities_by_site.get(key,[])
            utility_cb.configure(values=utils,state="readonly" if utils else "normal")
            utility_var.set(""); meter_var.set(""); meter_cb.configure(values=[])
            if rows:
                address_var.set(rows[0]["address"]); contact_var.set(rows[0]["contact"])
                prop_var.set(rows[0]["prop_code"])
                if rows[0]["prop_code"]: prop_frame.pack(fill="x",pady=4)
                else: prop_frame.pack_forget()
            open_same_site=[q for q in self.queries if q["client"]==c and q["site"]==s
                and q["status"]!="Resolved" and q.get("id")!=cf.get("id")]
            if open_same_site:
                preview="; ".join(f"{x['ref']} ({x['type']})" for x in open_same_site[:4])
                extra="" if len(open_same_site)<=4 else f" +{len(open_same_site)-4} more"
                related_site_lbl.config(text=f"Open queries for this site: {preview}{extra}")
                site_key=f"{c}|{s}"
                if not copy_from and warned_site[0]!=site_key:
                    warned_site[0]=site_key
                    messagebox.showwarning(
                        "Open queries already exist",
                        f"This site already has {len(open_same_site)} open quer{'y' if len(open_same_site)==1 else 'ies'}.\n\n"
                        f"{preview}{extra}\n\n"
                        "Please check these first to avoid duplicate logging.",
                        parent=dlg,
                    )
            else:
                related_site_lbl.config(text="")
            if nm_btn is not None: nm_btn.pack(side="left")

        def on_utility(*_):
            c=client_var.get(); s=site_var.get(); u=utility_var.get(); key=(c,s)
            rows=[r for r in self.meters.get(key,[]) if r["utility"]==u]
            opts=["— Whole site (no specific meter) —"]+get_meter_list()
            meter_cb.configure(values=opts)
            if len(opts)==2: meter_var.set(opts[1]); on_meter()
            else:
                meter_var.set(opts[0])
                spid_var.set(", ".join(r["spid"] for r in rows if r["spid"]))
                serial_var.set(", ".join(r["serial"] for r in rows if r["serial"]))

        def on_meter(*_):
            m=meter_var.get()
            if not m or m.startswith("— Whole site"):
                c=client_var.get(); s=site_var.get(); u=utility_var.get()
                rows=[r for r in self.meters.get((c,s),[]) if r["utility"]==u]
                spid_var.set(", ".join(r["spid"] for r in rows if r["spid"]))
                serial_var.set(", ".join(r["serial"] for r in rows if r["serial"]))
                meter_var.set(""); return
            for r in self.meters.get((client_var.get(),site_var.get()),[]):
                if r["utility"]!=utility_var.get(): continue
                lbl=f"{r['serial']} ({r['spid']})" if r["serial"] and r["spid"] else r["serial"] or r["spid"]
                if lbl==m: spid_var.set(r["spid"]); serial_var.set(r["serial"]); break

        def on_type(*_):
            if type_var.get() and not ref_var.get():
                ref_var.set(next_ref(self.queries,type_var.get()))

        # Client: autocomplete + cascade via trace
        utility_var.trace_add("write",on_utility)
        type_var.trace_add("write",on_type)
        fund_cb.bind("<<ComboboxSelected>>",lambda e:on_fund())
        fund_cb.bind("<FocusOut>",lambda e:on_fund())
        utility_cb.bind("<<ComboboxSelected>>",lambda _:on_utility())

        def _on_client_var(*_):
            c=client_var.get()
            if c in self.clients:
                on_client()
        client_var.trace_add("write",_on_client_var)

        # Pass on_client as on_select so _bind_autocomplete calls it on selection AND focusout
        self._bind_autocomplete(client_cb,lambda:self.clients,on_select=on_client)

        # Site: refresh values on focus/click, fire on_site on selection
        def _refresh_site_values(e=None):
            sites=get_site_list()
            site_cb.configure(values=sites)
        site_cb.bind("<FocusIn>",_refresh_site_values)
        site_cb.bind("<Button-1>",_refresh_site_values)
        site_cb.bind("<<ComboboxSelected>>",lambda e:on_site())
        self._bind_autocomplete(site_cb,get_site_list,on_select=on_site)

        self._bind_autocomplete(meter_cb,get_meter_list,on_select=on_meter)

        if copy_from:
            on_client()   # populates site list
            on_site()     # populates utility list and fills fund
            if copy_from.get("utility"): utility_var.set(copy_from["utility"]); on_utility()

        def add_new_meter():
            c=client_var.get(); s=site_var.get()
            if not c or not s: messagebox.showwarning("Select a site first","Please select a client and site first.",parent=dlg); return
            key=(c,s); rows=self.meters.get(key,[]); existing=rows[0] if rows else {}
            m_dlg=tk.Toplevel(dlg); m_dlg.title("Add new meter"); m_dlg.geometry("500x480")
            m_dlg.configure(bg=BG); m_dlg.grab_set(); m_dlg.resizable(False,True)
            hdr2=tk.Frame(m_dlg,bg=NAV,padx=20,pady=14); hdr2.pack(fill="x")
            tk.Label(hdr2,text="＋  New utility / meter",font=(FONT,12,"bold"),bg=NAV,fg="white").pack(anchor="w")
            tk.Label(hdr2,text=f"{c}  ·  {s}",font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))

            nu_var=tk.StringVar(); ns_var=tk.StringVar(); nr_var=tk.StringVar()
            def save_meter():
                if not nu_var.get().strip(): messagebox.showwarning("Required","Please enter a utility type.",parent=m_dlg); return
                if not self.sites_file or not os.path.exists(self.sites_file):
                    messagebox.showerror("No site file","Cannot find your sites.xlsx. Check Settings.",parent=m_dlg); return
                try:
                    wb=openpyxl.load_workbook(self.sites_file)
                    wb["Sites"].append([c,existing.get("fund",""),existing.get("prop_code",""),s,
                        existing.get("address",""),nu_var.get().strip(),ns_var.get().strip(),
                        nr_var.get().strip(),existing.get("contact","")])
                    wb.save(self.sites_file)
                except PermissionError:
                    messagebox.showerror("File in use","Sites.xlsx is open in Excel. Close it and try again.",parent=m_dlg); return
                except Exception as e:
                    messagebox.showerror("Error",str(e),parent=m_dlg); return
                self.clients,self.sites_by_client,self.meters,self.utilities_by_site,\
                    self.funds_by_client,self.sites_by_fund=load_site_data(self.sites_file)
                self.filter_client.configure(values=["All"]+self.clients)
                utility_cb.configure(values=self.utilities_by_site.get(key,[]))
                utility_var.set(nu_var.get().strip())
                spid_var.set(ns_var.get().strip()); serial_var.set(nr_var.get().strip())
                messagebox.showinfo("Added",f"New {nu_var.get()} meter added to {s}.",parent=m_dlg)
                m_dlg.destroy()
            ft=tk.Frame(m_dlg,bg=CARD,highlightthickness=1,highlightbackground=BORDER,padx=20,pady=12)
            ft.pack(fill="x",side="bottom")
            make_btn(ft,"Cancel",m_dlg.destroy,"default",padx=12,pady=6).pack(side="right",padx=(8,0))
            make_btn(ft,"Add to site list",save_meter,"primary",padx=14,pady=6).pack(side="right")

            body=tk.Frame(m_dlg,bg=BG,padx=24,pady=16); body.pack(fill="both",expand=True)
            tk.Label(body,text="Site address, contact and property code will be copied from the existing site.",
                     font=(FONT,9),bg=BG,fg=MUTED,wraplength=440).pack(anchor="w",pady=(0,12))
            def ml_combo(lbl,var,vals):
                f=tk.Frame(body,bg=BG); f.pack(fill="x",pady=4)
                tk.Label(f,text=lbl,font=(FONT,9),bg=BG,fg=MUTED,width=18,anchor="w").pack(side="left")
                cb=make_combo(f,var,vals,readonly=False,width=28); cb.pack(side="left",fill="x",expand=True)
            def ml_entry(lbl,var):
                f=tk.Frame(body,bg=BG); f.pack(fill="x",pady=4)
                tk.Label(f,text=lbl,font=(FONT,9),bg=BG,fg=MUTED,width=18,anchor="w").pack(side="left")
                ca=tk.Frame(f,bg=CARD,highlightthickness=1,highlightbackground=BORDER); ca.pack(side="left",fill="x",expand=True)
                tk.Entry(ca,textvariable=var,font=(FONT,10),bg=CARD,fg=TEXT,relief="flat",bd=6,highlightthickness=0).pack(fill="x")
            ml_combo("Utility type *",nu_var,UTILITY_OPTIONS)
            ml_entry("Supply point ID",ns_var); ml_entry("Meter serial",nr_var)
            divider(body)
            for lbl,val in [("Address",existing.get("address","")),("Contact",existing.get("contact",""))]:
                if val:
                    r=tk.Frame(body,bg=BG); r.pack(fill="x",pady=1)
                    tk.Label(r,text=lbl,font=(FONT,8),bg=BG,fg=MUTED,width=12,anchor="w").pack(side="left")
                    tk.Label(r,text=val,font=(FONT,8),bg=BG,fg=TEXT2).pack(side="left")

        nm_btn=make_btn(nm_frame,"＋  Add new utility / meter for this site",add_new_meter,"default",padx=14,pady=6)

        footer=tk.Frame(dlg,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=20,pady=14); footer.pack(fill="x",side="bottom")
        make_btn(footer,"Cancel",dlg.destroy,"default",padx=14,pady=7).pack(side="right",padx=(8,0))

        def save():
            if not client_var.get().strip() or not type_var.get() or not desc_text.get("1.0","end").strip():
                messagebox.showwarning("Required","Client, type and description are required.",parent=dlg); return
            if not self._validate_action_date(chase_var.get().strip(), parent=dlg):
                return
            if not self._confirm_high_volume_day(
                chase_var.get().strip(),
                exclude_ids=[edit_query.get("id")] if is_edit else None,
                add_count=1,
                parent=dlg,
            ):
                return
            client=client_var.get().strip()
            site=site_var.get().strip()
            assignee=assignee_var.get().strip()
            if assignee=="(Unassigned)": assignee=""
            raised=raised_var.get().strip()

            if is_edit:
                old_chase=edit_query.get("chase_date","").strip()
                new_chase=chase_var.get().strip()
                prev_last_date=edit_query.get("last_date", "")
                old_desc=(edit_query.get("desc", "") or "").strip()
                new_desc=desc_text.get("1.0","end").strip()
                note_added=(new_desc != old_desc and bool(new_desc))
                live_query = next((qx for qx in self.queries if qx.get("id") == edit_query.get("id")), None)
                if not live_query:
                    messagebox.showwarning("Query changed", "This query could not be found in the latest data. Please reopen it and try again.", parent=dlg)
                    return
                if not self._confirm_pushback_history(live_query, old_chase, new_chase, note_added=note_added, parent=dlg):
                    return
                # Update existing query in place — preserve id, ref, opened, original log
                live_query.update({
                    "client":client,"fund":fund_var.get().strip(),"site":site,
                    "utility":utility_var.get(),"meter":meter_var.get(),"type":type_var.get(),
                    "status":status_var.get(),"priority":priority_var.get(),
                    "desc":new_desc,
                    "chase_date":new_chase,
                    "address":address_var.get(),"spid":spid_var.get(),
                    "serial":serial_var.get(),"contact":contact_var.get(),
                    "prop_code":prop_var.get(),"assigned_to":assignee,
                    "raised_date":live_query.get("raised_date","") if is_type_changed_query else raised,
                    "last_by":self.username,"last_date":today_str(),
                })
                if is_pushback(old_chase, new_chase, prev_last_date):
                    append_pushback_event(live_query, self.username, old_chase, new_chase)
                live_query["log"]+= f" | {stamp(self.username)} Query details edited"
                if not self._save_queries():
                    return
                self._refresh_table(); self._show_daily_banner(); dlg.destroy()
                return

            # ── New query ──────────────────────────────────────────────────────
            is_new_site = (client not in self.clients) or (site and site not in self.sites_by_client.get(client,[]))
            if is_new_site and self.sites_file and os.path.exists(self.sites_file):
                try:
                    wb=openpyxl.load_workbook(self.sites_file)
                    ws_s=wb["Sites"] if "Sites" in wb.sheetnames else wb.active
                    ws_s.append([
                        client, fund_var.get().strip(), prop_var.get().strip(),
                        site, address_var.get().strip(),
                        utility_var.get().strip(), spid_var.get().strip(),
                        serial_var.get().strip(), contact_var.get().strip()
                    ])
                    wb.save(self.sites_file)
                    self.clients,self.sites_by_client,self.meters,self.utilities_by_site,\
                        self.funds_by_client,self.sites_by_fund=load_site_data(self.sites_file)
                    self.filter_client.configure(values=["All"]+self.clients)
                except Exception as e:
                    messagebox.showwarning("Site list not updated",f"Query saved but could not add to sites.xlsx:\n{e}",parent=dlg)
            log_entry=f"{stamp(self.username)} Query logged"
            if raised and raised!=today_str():
                log_entry+=f" — client raised on {fmt_date(raised)}"
            q={"id":str(int(datetime.now().timestamp()*1000)),
               "ref":ref_var.get() or next_ref(self.queries,type_var.get()),
               "client":client,"fund":fund_var.get().strip(),"site":site,
               "utility":utility_var.get(),"meter":meter_var.get(),"type":type_var.get(),
               "status":status_var.get(),"priority":priority_var.get(),
               "desc":desc_text.get("1.0","end").strip(),"opened":today_str(),
               "chase_date":chase_var.get().strip(),"resolved_date":"",
               "log":log_entry,
               "address":address_var.get(),"spid":spid_var.get(),"serial":serial_var.get(),
               "contact":contact_var.get(),"prop_code":prop_var.get(),
               "last_by":self.username,"last_date":today_str(),
               "assigned_to":assignee,"raised_date":raised}
            self.queries.insert(0,q)
            if not self._save_queries():
                return
            self._refresh_table(); self._show_daily_banner(); dlg.destroy()

        make_btn(footer,"Save changes" if is_edit else "Log query",save,"primary",padx=16,pady=7).pack(side="right")

    def _open_detail(self,event):
        sel=self.tree.focus()
        if not sel: return
        q=next((x for x in self.queries if x["id"]==sel),None)
        if q: self._open_detail_query(q)

    def _open_detail_query(self,q):
        if q["id"] in self.recent_ids: self.recent_ids.remove(q["id"])
        self.recent_ids.insert(0,q["id"])
        self.recent_ids=self.recent_ids[:5]

        dlg=tk.Toplevel(self); dlg.title(f"{q['ref']}")
        dlg.geometry("700x860"); dlg.configure(bg=BG); dlg.grab_set(); dlg.resizable(False,True)

        sc,scbg=S_COLORS.get(q["status"],(TEXT2,CARD2))
        hdr=tk.Frame(dlg,bg=NAV,padx=24,pady=18); hdr.pack(fill="x")
        tk.Frame(dlg,bg=ACCENT,height=2).pack(fill="x")
        hr=tk.Frame(hdr,bg=NAV); hr.pack(fill="x")
        tk.Label(hr,text=q["ref"],font=(FONT,15,"bold"),bg=NAV,fg=TEXT).pack(side="left")
        pill=tk.Label(hr,text=f"  {q['status']}  ",font=(FONT,8,"bold"),bg=scbg,fg=sc,padx=6,pady=3)
        pill.pack(side="left",padx=12)
        pc,pbg=P_COLORS.get(q["priority"],("#94A3B8","#1E2D45"))
        tk.Label(hr,text=f"  {q['priority']}  ",font=(FONT,8),bg=pbg,fg=pc,padx=4,pady=3).pack(side="left")
        edit_btn=tk.Label(hr,text="✎ Edit query",font=(FONT,9),bg=NAV2,fg=TEXT2,
                          padx=10,pady=4,cursor="hand2")
        edit_btn.pack(side="right")
        def open_edit():
            dlg.destroy()
            self._open_add_dialog(edit_query=q)
        edit_btn.bind("<Button-1>",lambda e:open_edit())
        tk.Label(hdr,text=q["type"],font=(FONT,10),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(6,0))

        outer=tk.Frame(dlg,bg=BG); outer.pack(fill="both",expand=True)
        body,_=scrollable_frame(outer)
        note_added_since_open=[False]

        def get_live_query(require_exists=False):
            live=next((qx for qx in self.queries if qx.get("id")==q.get("id")), None)
            if live is None and require_exists:
                messagebox.showwarning(
                    "Query changed",
                    "This query is no longer available in the latest data. Please reopen it and try again.",
                    parent=dlg,
                )
            return live

        def info_card(items):
            card=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=12)
            card.pack(fill="x",pady=(0,8))
            for lbl,val in items:
                if not val: continue
                row=tk.Frame(card,bg=CARD2); row.pack(fill="x",pady=3)
                tk.Label(row,text=lbl,font=(FONT,9),bg=CARD2,fg=TEXT2,width=18,anchor="w").pack(side="left")
                val_lbl=tk.Label(row,text=val,font=(FONT,10),bg=CARD2,fg=TEXT,wraplength=340,justify="left",cursor="hand2")
                val_lbl.pack(side="left")
                # Add right-click copy for values
                def make_copy_handler(v):
                    def on_copy(e):
                        dlg.clipboard_clear()
                        dlg.clipboard_append(v)
                        dlg.update_idletasks()
                    return on_copy
                val_lbl.bind("<Button-3>", make_copy_handler(val))

        # Build raised/logged info string with intake time if raised is set
        raised_info=fmt_date(q.get("raised_date",""))
        logged_info=fmt_date(q["opened"])
        if q.get("raised_date") and q["raised_date"]!=q.get("opened",""):
            try:
                intake=sla_intake_working_days(q.get("raised_date", ""), q.get("opened", ""))
                sla_status="✓ MET" if intake<=1 else "✗ BREACHED"
                logged_info+=f"  |  SLA: {intake}d intake — {sla_status}"
            except: pass

        info_card([("Client",q["client"]),("Fund",q.get("fund","")),("Site",q["site"]),
                   ("Utility",q.get("utility","")),("Meter",q.get("meter","")),
                   ("Property code",q.get("prop_code","")),("Address",q.get("address","")),
                   ("Supply point ID",q.get("spid","")),("Meter serial",q.get("serial","")),
                   ("MA contact",q.get("contact","")),
                   ("Raised by client",raised_info),("Logged",logged_info),
                   ("Last updated by",q.get("last_by","")+" on "+fmt_date(q.get("last_date","")) if q.get("last_by") else "")])

        tk.Label(body,text="DESCRIPTION",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(4,4))
        dc=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); dc.pack(fill="x",pady=(0,8))
        desc_scroll = tk.Scrollbar(dc)
        desc_scroll.pack(side="right", fill="y")
        desc_box=tk.Text(dc,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=10,height=3,wrap="word",highlightthickness=0, yscrollcommand=desc_scroll.set)
        desc_box.insert("1.0",q["desc"])
        desc_box.pack(fill="x")
        desc_scroll.config(command=desc_box.yview)
        # Editable in detail view and still spell-checked
        attach_text_spellcheck(desc_box)
        desc_box.bind("<Control-a>", lambda e: (desc_box.tag_add("sel","1.0","end"), "break")[1])

        # Drag-to-resize handle (same behavior as New Query)
        resize_bar=tk.Frame(body,bg=BORDER,height=6,cursor="sb_v_double_arrow")
        resize_bar.pack(fill="x",pady=(0,2))
        _drag_desc={"start_y":0,"start_h":3}

        def _detail_resize_start(e):
            _drag_desc["start_y"]=e.y_root
            _drag_desc["start_h"]=int(desc_box.cget("height"))

        def _detail_resize_drag(e):
            total_dy=e.y_root-_drag_desc["start_y"]
            delta=int(round(total_dy/8.0))
            new_h=max(3,min(24,_drag_desc["start_h"]+delta))
            desc_box.configure(height=new_h)

        resize_bar.bind("<ButtonPress-1>",_detail_resize_start)
        resize_bar.bind("<B1-Motion>",_detail_resize_drag)
        tk.Label(body,text="↕ drag to resize",font=(FONT,7),bg=BG,fg=MUTED).pack(anchor="e",pady=(0,4))

        tk.Label(body,text="UPDATE",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(4,4))
        ec=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=12); ec.pack(fill="x",pady=(0,8))
        status_var=tk.StringVar(value=q["status"]); priority_var=tk.StringVar(value=q["priority"])
        chase_var=tk.StringVar(value=q.get("chase_date",""))
        assignee_var_d=tk.StringVar(value=q.get("assigned_to",""))
        erow=tk.Frame(ec,bg=CARD2); erow.pack(fill="x")
        for lbl,var,vals,w in [("Status",status_var,STATUSES,14),("Priority",priority_var,PRIORITIES,10)]:
            tk.Label(erow,text=lbl,font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,4))
            make_combo(erow,var,vals,readonly=True,width=w).pack(side="left",padx=(0,14))
        tk.Label(erow,text="Action date",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,4))
        cf2=tk.Frame(erow,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); cf2.pack(side="left")
        tk.Entry(cf2,textvariable=chase_var,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=5,width=12,highlightthickness=0).pack()
        make_btn(
            erow,
            "📅",
            lambda: _show_cal(
                dlg,
                chase_var,
                get_day_load=lambda d: self._day_workload(d, exclude_ids=[q.get("id")]),
                confirm_day_selection=lambda d: self._confirm_high_volume_day(
                    d, exclude_ids=[q.get("id")], add_count=1, parent=dlg
                ),
                high_volume_threshold=self._get_high_volume_threshold(),
                date_block_reason=self._action_date_block_reason,
            ),
            "default",
            padx=6,
            pady=3,
        ).pack(side="left",padx=(4,0))
        for days,lbl in [(7,"+7d"),(14,"+14d")]:
            def snooze(d=days,cv=chase_var):
                base=date.today()
                try: base=datetime.strptime(cv.get(),"%Y-%m-%d").date()
                except: pass
                cv.set((base+timedelta(days=d)).strftime("%Y-%m-%d"))
            make_btn(erow,lbl,snooze,"default",padx=8,pady=3).pack(side="left",padx=(6,0))

        wknd_lbl=tk.Label(erow,text="⚠ Non-working day",font=(FONT,8),bg=CARD2,fg=WARNING)
        def chk_wknd(*_):
            if self._action_date_block_reason(chase_var.get()): wknd_lbl.pack(side="left",padx=6)
            else: wknd_lbl.pack_forget()
        chase_var.trace_add("write",chk_wknd); chk_wknd()

        # Assignee row
        arow=tk.Frame(ec,bg=CARD2); arow.pack(fill="x",pady=(8,0))
        tk.Label(arow,text="Assigned to",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,8))
        all_members=list(dict.fromkeys([self.username]+self.team_members))
        make_combo(arow,assignee_var_d,["(Unassigned)"]+all_members,readonly=False,width=22).pack(side="left")

        # Raised date row — when the client originally raised the query
        raised_var_d=tk.StringVar(value=q.get("raised_date",""))
        is_type_changed_query = bool(re.search(r"Query created from type change", q.get("log", "")))
        if not is_type_changed_query:
            rrow=tk.Frame(ec,bg=CARD2); rrow.pack(fill="x",pady=(8,0))
            tk.Label(rrow,text="Query raised by client",font=(FONT,9),bg=CARD2,fg=TEXT2).pack(side="left",padx=(0,8))
            rf2=tk.Frame(rrow,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); rf2.pack(side="left")
            tk.Entry(rf2,textvariable=raised_var_d,font=(FONT,10),bg=CARD2,fg=TEXT,
                     insertbackground=TEXT,relief="flat",bd=5,width=12,highlightthickness=0).pack()
            make_btn(rrow,"📅",lambda:_show_cal(dlg,raised_var_d),"default",padx=6,pady=3).pack(side="left",padx=(4,0))
            # Show intake time if raised date is set and different from opened
            if q.get("raised_date") and q["raised_date"]!=q.get("opened",""):
                try:
                    intake=sla_intake_working_days(q.get("raised_date", ""), q.get("opened", ""))
                    sla_met=intake<=1
                    intake_col=SUCCESS if sla_met else DANGER
                    sla_txt=f"  {intake}d — SLA {"MET ✓" if sla_met else "BREACHED ✗"}"
                    tk.Label(rrow,text=sla_txt,font=(FONT,8,"bold"),
                             bg=CARD2,fg=intake_col).pack(side="left",padx=(8,0))
                except: pass

        # Action buttons: Copy Summary and Related Queries
        action_frame = tk.Frame(body, bg=BG)
        action_frame.pack(fill="x", pady=(4, 8))
        
        def copy_query_summary():
            summary_lines = [
                f"Reference: {q['ref']}",
                f"Client: {q['client']}",
                f"Site: {q['site']}",
                f"Type: {q['type']}",
                f"Status: {q['status']}",
                f"Priority: {q['priority']}",
                f"Description: {q['desc']}",
            ]
            if q.get('log'):
                log_entries = public_log_entries(q.get('log', ''))
                if log_entries:
                    summary_lines.append("Activity Log:")
                    for entry in log_entries[:10]:
                        summary_lines.append(f"  • {entry}")
            summary = "\n".join(line for line in summary_lines if line)
            dlg.clipboard_clear()
            dlg.clipboard_append(summary)
            dlg.update_idletasks()
            messagebox.showinfo("Copied", "Query summary copied to clipboard.", parent=dlg)
        
        make_btn(action_frame, "📋 Copy Summary", copy_query_summary, "default", padx=12, pady=6).pack(side="left")
        
        related_count = sum(1 for rq in self.queries if rq["site"] == q["site"] and rq["id"] != q["id"])
        if related_count > 0:
            def open_related_window():
                self._open_related_queries_window(q)
            make_btn(action_frame, f"🔗 View Related Queries ({related_count})", open_related_window, "default", padx=12, pady=6).pack(side="left", padx=(6, 0))

        tk.Label(body,text="ACTIVITY LOG",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(4,4))
        lc=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); lc.pack(fill="x",pady=(0,8))
        log_scroll = tk.Scrollbar(lc)
        log_scroll.pack(side="right", fill="y")
        log_box=tk.Text(lc,font=(FONT,9),bg=CARD2,fg=TEXT2,relief="flat",bd=10,height=6,wrap="word",highlightthickness=0, yscrollcommand=log_scroll.set)
        log_box.pack(fill="x")
        log_scroll.config(command=log_box.yview)
        # Bind copy and select-all
        log_box.bind("<Control-c>", lambda e: (dlg.clipboard_clear(), dlg.clipboard_append(log_box.get("1.0","end-1c")), "break")[2])
        log_box.bind("<Control-a>", lambda e: (log_box.tag_add("sel","1.0","end"), "break")[1])
        log_box.bind("<Key>", lambda e: "break")

        def refresh_log():
            live_q=get_live_query() or q
            log_box.delete("1.0","end")
            for entry in public_log_entries(live_q.get("log", "")):
                log_box.insert("end","• "+entry+"\n")
            log_box.see("end")
            log_box.yview_moveto(1.0)

        refresh_log()

        tk.Label(body,text="ADD NOTE",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(0,4))
        nc=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=12,pady=8); nc.pack(fill="x",pady=(0,10))
        nr=tk.Frame(nc,bg=CARD2); nr.pack(fill="x")
        nef=tk.Frame(nr,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); nef.pack(fill="x",expand=True)
        note_box=tk.Text(nef,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=6,height=3,wrap="word",highlightthickness=0)
        note_box.pack(fill="x")
        attach_text_spellcheck(note_box)
        tk.Label(nc,text="Adds a timestamped entry to the activity log.",font=(FONT,8),bg=CARD2,fg=MUTED).pack(anchor="w",pady=(6,0))
        note_actions=tk.Frame(nc,bg=CARD2)
        note_actions.pack(fill="x",pady=(8,0))
        def add_note():
            live_q=get_live_query(True)
            if not live_q:
                return
            n=note_box.get("1.0","end").strip()
            if not n: return
            live_q["log"]+=" | "+stamp(self.username)+" "+n
            live_q["last_by"]=self.username
            live_q["last_date"]=today_str()
            if not self._save_queries():
                return
            note_box.delete("1.0","end")
            note_added_since_open[0]=True
            refresh_log()
            self._refresh_table()
        note_box.bind("<Control-Return>",lambda e:(add_note(), "break")[1])
        add_note_btn=make_btn(note_actions,"Add note",add_note,"primary",padx=18,pady=7)
        add_note_btn.pack(anchor="w")

        divider(body)
        att_hdr=tk.Frame(body,bg=BG); att_hdr.pack(fill="x",pady=(0,6))
        tk.Label(att_hdr,text="ATTACHMENTS",font=(FONT,8,"bold"),bg=BG,fg=MUTED).pack(side="left")

        att_folder=get_attachment_folder(self.sites_file,q)
        def open_att_folder():
            f=get_attachment_folder(self.sites_file,q)
            if f: open_folder(f)
        if att_folder:
            make_btn(att_hdr,"Open folder",open_att_folder,"default",padx=10,pady=3).pack(side="right")

        att_list_frame=tk.Frame(body,bg=CARD2,highlightthickness=1,highlightbackground=BORDER)
        att_list_frame.pack(fill="x",pady=(0,8))

        def refresh_attachments():
            for w in att_list_frame.winfo_children(): w.destroy()
            files=list_attachments(self.sites_file,q)
            if not files:
                tk.Label(att_list_frame,text="No attachments yet.",font=(FONT,9),
                         bg=CARD2,fg=MUTED,pady=10).pack(anchor="w",padx=14)
            else:
                for fname,fpath in files:
                    row=tk.Frame(att_list_frame,bg=CARD2); row.pack(fill="x",padx=10,pady=3)
                    ext=os.path.splitext(fname)[1].lower()
                    icon={"pdf":"📄",".xlsx":"📊",".xls":"📊",".docx":"📝",".doc":"📝",
                          ".png":"🖼",".jpg":"🖼",".jpeg":"🖼",".msg":"✉",".eml":"✉"}.get(ext,"📎")
                    lbl=tk.Label(row,text=f"{icon}  {fname}",font=(FONT,9),bg=CARD2,fg=ACCENT2,
                                 cursor="hand2",anchor="w")
                    lbl.pack(side="left",fill="x",expand=True)
                    lbl.bind("<Button-1>",lambda e,p=fpath:open_file(p))
                    lbl.bind("<Enter>",lambda e,l=lbl:l.configure(fg=ACCENT2))
                    lbl.bind("<Leave>",lambda e,l=lbl:l.configure(fg=ACCENT))
                    try:
                        sz=os.path.getsize(fpath)
                        sz_str=f"{sz//1024}KB" if sz>1024 else f"{sz}B"
                    except: sz_str=""
                    tk.Label(row,text=sz_str,font=(FONT,8),bg=CARD2,fg=MUTED,width=8).pack(side="right",padx=(0,4))
                    def rm(fp=fpath,fn=fname):
                        if messagebox.askyesno("Remove attachment",
                            f"Remove '{fn}' from the folder?\n\nThis will delete the file permanently.",
                            parent=dlg):
                            try: os.remove(fp)
                            except Exception as e: messagebox.showerror("Error",str(e),parent=dlg); return
                            q["log"]+=" | "+stamp(self.username)+f" Removed attachment: {fn}"
                            refresh_attachments(); refresh_log()
                    make_btn(row,"✕",rm,"danger",padx=6,pady=2).pack(side="right",padx=(4,0))

        refresh_attachments()

        att_actions=tk.Frame(body,bg=BG); att_actions.pack(fill="x",pady=(0,8))

        def open_drop_folder():
            try:
                if self.sites_file and (self.sites_file.startswith("http://") or self.sites_file.startswith("https://")):
                    messagebox.showerror("SharePoint not supported", "Sites file cannot be a SharePoint URL.\n\nDownload the file locally or map the folder as a local drive, then update Settings.", parent=dlg)
                    return
                inbox=get_drop_inbox(self.sites_file)
                # Ensure folder exists before opening
                os.makedirs(inbox, exist_ok=True)
                if inbox and os.path.exists(inbox):
                    # Use full absolute path for Windows Explorer
                    abs_path = os.path.abspath(inbox)
                    open_folder(abs_path)
                    messagebox.showinfo("Opened", f"Drop folder:\n\n{abs_path}", parent=dlg)
                else:
                    messagebox.showerror("Folder not found", f"Cannot locate:\n\n{inbox}\n\nCheck Sites file path in Settings.", parent=dlg)
            except Exception as e:
                messagebox.showerror("Error opening folder", f"Cannot open folder:\n\n{str(e)}", parent=dlg)

        make_btn(att_actions,"Open _DROP_HERE folder",open_drop_folder,"default",padx=12,pady=5).pack(side="left")

        def browse_attach():
            paths=filedialog.askopenfilenames(
                title="Select files to attach", parent=dlg,
                filetypes=[("All files","*.*"),("Emails","*.msg *.eml"),
                           ("PDFs","*.pdf"),("Excel","*.xlsx *.xls"),
                           ("Images","*.png *.jpg *.jpeg")])
            if not paths: return
            added=[]
            for p in paths:
                fname,dest=save_attachment(self.sites_file,q,p)
                if fname: added.append(fname)
            if added:
                note=f"Added attachment{'s' if len(added)>1 else ''}: {', '.join(added)}"
                q["log"]+=" | "+stamp(self.username)+" "+note
                refresh_attachments(); refresh_log()

        make_btn(att_actions,"Browse & attach files",browse_attach,"default",padx=12,pady=5).pack(side="left",padx=(10,0))

        def draft_email():
            subject=f"Re: {q['ref']} – {q['type']} – {q['site']}"
            body_txt=(f"Dear {q.get('contact','') or 'Team'},\n\n"
                      f"I am writing to follow up on the above query.\n\n"
                      f"Reference: {q['ref']}\nSite: {q['site']}\nUtility: {q.get('utility','')}\n"
                      f"Query type: {q['type']}\n\nDetails: {q['desc']}\n\n"
                      f"Could you please provide an update at your earliest convenience?\n\n"
                      f"Kind regards,\n{self.username}")
            em_dlg=tk.Toplevel(dlg); em_dlg.title("Email draft"); em_dlg.geometry("580x480")
            em_dlg.configure(bg=BG); em_dlg.grab_set()
            hdr3=tk.Frame(em_dlg,bg=CARD,padx=20,pady=14); hdr3.pack(fill="x")
            tk.Frame(em_dlg,bg=ACCENT,height=2).pack(fill="x")
            tk.Label(hdr3,text="✉  Email draft",font=(FONT,12,"bold"),bg=NAV,fg="white").pack(anchor="w")
            tk.Label(hdr3,text="Copy and paste into your email client.",font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            eb=tk.Frame(em_dlg,bg=BG,padx=20,pady=16); eb.pack(fill="both",expand=True)
            tk.Label(eb,text="Subject",font=(FONT,9,"bold"),bg=BG,fg=MUTED).pack(anchor="w")
            sc2=tk.Frame(eb,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); sc2.pack(fill="x",pady=(4,12))
            subj_e=tk.Entry(sc2,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=8,highlightthickness=0); subj_e.pack(fill="x")
            subj_e.insert(0,subject)
            tk.Label(eb,text="Body",font=(FONT,9,"bold"),bg=BG,fg=MUTED).pack(anchor="w")
            bc2=tk.Frame(eb,bg=CARD2,highlightthickness=1,highlightbackground=BORDER); bc2.pack(fill="both",expand=True,pady=(4,0))
            body_t=tk.Text(bc2,font=(FONT,10),bg=CARD2,fg=TEXT,insertbackground=TEXT,relief="flat",bd=8,wrap="word",highlightthickness=0)
            body_t.pack(fill="both",expand=True); body_t.insert("1.0",body_txt)
            ef2=tk.Frame(em_dlg,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=12); ef2.pack(fill="x",side="bottom")
            def copy_all():
                em_dlg.clipboard_clear()
                em_dlg.clipboard_append(f"Subject: {subj_e.get()}\n\n{body_t.get('1.0','end').strip()}")
                messagebox.showinfo("Copied","Email copied to clipboard.",parent=em_dlg)
            make_btn(ef2,"Copy to clipboard",copy_all,"primary",padx=14,pady=6).pack(side="right")
            make_btn(ef2,"Close",em_dlg.destroy,"default",padx=12,pady=6).pack(side="right",padx=(0,8))

        footer=tk.Frame(dlg,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=20,pady=14); footer.pack(fill="x",side="bottom")
        tk.Frame(dlg,bg=BORDER,height=1).pack(fill="x",side="bottom")

        def delete_query():
            live_q=get_live_query(True)
            if not live_q:
                return
            if messagebox.askyesno("Delete",f"Delete {live_q['ref']}? This cannot be undone.",parent=dlg):
                self.queries=[x for x in self.queries if x["id"]!=live_q["id"]]
                if not self._save_queries():
                    return
                self._refresh_table(); self._show_daily_banner(); dlg.destroy()

        def change_type():
            """Close this query and spawn a new linked one of a different type."""
            live_q=get_live_query(True)
            if not live_q:
                return
            if live_q["status"]=="Resolved":
                messagebox.showwarning("Already resolved",
                    "This query is already resolved. Duplicate it instead if you need a new one.",parent=dlg)
                return
            ct_dlg=tk.Toplevel(dlg); ct_dlg.title("Change query type")
            ct_dlg.geometry("480x340"); ct_dlg.configure(bg=BG); ct_dlg.grab_set(); ct_dlg.resizable(False,False)
            hdr_ct=tk.Frame(ct_dlg,bg=NAV,padx=20,pady=14); hdr_ct.pack(fill="x")
            tk.Label(hdr_ct,text="⇄  Change query type",font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
            tk.Label(hdr_ct,text=f"{live_q['ref']}  ·  {live_q['client']}  ·  {live_q['site']}",
                     font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            tk.Frame(ct_dlg,bg=ACCENT,height=2).pack(fill="x")

            body_ct=tk.Frame(ct_dlg,bg=BG,padx=24,pady=20); body_ct.pack(fill="both",expand=True)
            tk.Label(body_ct,text="This will:",font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(0,6))
            for bullet in [
                f"• Mark {live_q['ref']} as Resolved with a note explaining the type change",
                "• Create a new linked query of the type you choose below",
                "• Copy all site details, description and attachments to the new query",
            ]:
                tk.Label(body_ct,text=bullet,font=(FONT,9),bg=BG,fg=TEXT2,
                         justify="left").pack(anchor="w",pady=1)

            tk.Label(body_ct,text="New query type",font=(FONT,9,"bold"),bg=BG,fg=MUTED).pack(anchor="w",pady=(16,4))
            new_type_var=tk.StringVar()
            other_types=[t for t in QUERY_TYPES if t!=live_q["type"]]
            type_cb=make_combo(body_ct,new_type_var,other_types,readonly=True,width=34)
            type_cb.pack(anchor="w")

            ft_ct=tk.Frame(ct_dlg,bg=CARD2,padx=16,pady=12); ft_ct.pack(fill="x",side="bottom")
            make_btn(ft_ct,"Cancel",ct_dlg.destroy,"default",padx=12,pady=6).pack(side="right",padx=(8,0))

            def confirm_change():
                new_type=new_type_var.get().strip()
                if not new_type:
                    messagebox.showwarning("Select a type","Please select the new query type.",parent=ct_dlg)
                    return
                current_q=get_live_query(True)
                if not current_q:
                    return
                # 1. Close the original query with a log note
                note=(f"Query type changed from '{current_q['type']}' to '{new_type}' by {self.username}. "
                      f"This query closed and continued as new reference below.")
                current_q["status"]="Resolved"
                current_q["resolved_date"]=today_str()
                current_q["log"]+=f" | {stamp(self.username)} {note}"
                current_q["last_by"]=self.username; current_q["last_date"]=today_str()

                # 2. Create new linked query
                new_ref=next_ref(self.queries, new_type)
                new_q={
                    "id":   str(int(datetime.now().timestamp()*1000)),
                    "ref":  new_ref,
                    "client":   current_q["client"],   "fund":  current_q.get("fund",""),
                    "site":     current_q["site"],      "utility": current_q.get("utility",""),
                    "meter":    current_q.get("meter",""), "type": new_type,
                    "status":   "Open",         "priority": current_q["priority"],
                    "desc":     current_q["desc"],
                    "opened":   today_str(),    "chase_date": current_q.get("chase_date",""),
                    "resolved_date": "",
                    "log":  (f"{stamp(self.username)} Query created from type change. "
                             f"Continued from {current_q['ref']} ({current_q['type']})."),
                    "address":  current_q.get("address",""),  "spid":    current_q.get("spid",""),
                    "serial":   current_q.get("serial",""),   "contact": current_q.get("contact",""),
                    "prop_code":current_q.get("prop_code",""),"last_by": self.username,
                    "last_date":today_str(),          "assigned_to": current_q.get("assigned_to",""),
                    "raised_date": "",
                }
                self.queries.insert(0, new_q)
                if not self._save_queries():
                    return
                self._refresh_table(); self._show_daily_banner()
                ct_dlg.destroy(); dlg.destroy()
                # Open the new query immediately
                self._open_detail_query(new_q)

            make_btn(ft_ct,"Change type & create new query",confirm_change,"primary",padx=14,pady=6).pack(side="right")

        def save_changes():
            live_q=get_live_query(True)
            if not live_q:
                return
            if live_q["status"]=="Resolved" and status_var.get()!="Resolved":
                if not messagebox.askyesno("Re-open query",f"{live_q['ref']} is already resolved. Re-open it?",parent=dlg): return
            if not self._validate_action_date(chase_var.get().strip(), parent=dlg):
                return
            old_chase=live_q.get("chase_date","").strip()
            prev_last_date=live_q.get("last_date", "")
            new_chase=chase_var.get().strip()
            if not self._confirm_high_volume_day(new_chase, exclude_ids=[live_q.get("id")], add_count=1, parent=dlg):
                return
            old_desc=(live_q.get("desc", "") or "").strip()
            new_desc=desc_box.get("1.0","end").strip()
            pending_note=note_box.get("1.0","end").strip()
            note_added=note_added_since_open[0] or bool(pending_note) or (new_desc != old_desc and bool(new_desc))
            if not self._confirm_pushback_history(live_q, old_chase, new_chase, note_added=note_added, parent=dlg):
                return
            live_q["status"]=status_var.get(); live_q["priority"]=priority_var.get()
            live_q["chase_date"]=new_chase; live_q["desc"]=new_desc
            live_q["last_by"]=self.username; live_q["last_date"]=today_str()
            assignee=assignee_var_d.get().strip()
            live_q["assigned_to"]="" if assignee=="(Unassigned)" else assignee
            if not is_type_changed_query:
                live_q["raised_date"]=raised_var_d.get().strip()
            if pending_note:
                live_q["log"] += f" | {stamp(self.username)} {pending_note}"
            if is_pushback(old_chase, new_chase, prev_last_date):
                append_pushback_event(live_q, self.username, old_chase, new_chase)
            if live_q["status"]=="Resolved" and not live_q.get("resolved_date"):
                live_q["resolved_date"]=today_str(); live_q["log"]+=f" | {stamp(self.username)} Marked as resolved"
            elif live_q["status"]!="Resolved":
                live_q["resolved_date"]=""
            if not self._save_queries():
                return
            messagebox.showinfo("Saved", f"Changes to {live_q['ref']} have been saved successfully.", parent=dlg)
            self._excel_mtime=self._excel_mtime_now()  # update our snapshot
            self._refresh_table(); self._show_daily_banner(); dlg.destroy()

        def transfer_query():
            """Copy this query to another linked tracker, mark it transferred here."""
            trackers=getattr(self,"linked_trackers",[])
            if not trackers:
                messagebox.showinfo("No linked trackers",
                    "No linked trackers are configured.\n\n"
                    "Go to Settings → Linked Trackers to add them.",parent=dlg)
                return

            tr_dlg=tk.Toplevel(dlg); tr_dlg.title("Transfer query")
            tr_dlg.geometry("520x420"); tr_dlg.configure(bg=BG); tr_dlg.grab_set(); tr_dlg.resizable(False,False)
            hdr_tr=tk.Frame(tr_dlg,bg=NAV,padx=20,pady=14); hdr_tr.pack(fill="x")
            tk.Label(hdr_tr,text="↗  Transfer query",font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
            tk.Label(hdr_tr,text=f"{q['ref']}  ·  {q['client']}  ·  {q['site']}",
                     font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            tk.Frame(tr_dlg,bg=ACCENT,height=2).pack(fill="x")

            body_tr=tk.Frame(tr_dlg,bg=BG,padx=24,pady=20); body_tr.pack(fill="both",expand=True)
            tk.Label(body_tr,text="Select destination tracker:",
                     font=(FONT,10,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(0,12))

            # Tracker list
            for t in trackers:
                tname=t.get("name","Unnamed"); tfile=t.get("excel_file","")
                exists=os.path.exists(tfile) if tfile else False
                card=tk.Frame(body_tr,bg=CARD2,highlightthickness=1,
                              highlightbackground=BORDER,padx=14,pady=10,cursor="hand2" if exists else "arrow")
                card.pack(fill="x",pady=(0,8))
                tk.Label(card,text=tname,font=(FONT,10,"bold"),bg=CARD2,
                         fg=TEXT if exists else MUTED).pack(anchor="w")
                status_txt=tfile if tfile else "No file configured"
                tk.Label(card,text=status_txt,font=(FONT,8),bg=CARD2,fg=TEXT2 if exists else DANGER).pack(anchor="w")
                if not exists:
                    tk.Label(card,text="⚠ File not found",font=(FONT,8),bg=CARD2,fg=DANGER).pack(anchor="w")
                    continue

                def do_transfer(tfile=tfile,tname=tname,tracker_cfg=t):
                    # Guard: don't let someone transfer to their own tracker
                    if os.path.abspath(tfile)==os.path.abspath(self.excel_file):
                        messagebox.showerror("Invalid destination",
                            "That is your own tracker — please choose a different one.",parent=tr_dlg)
                        return

                    # ── Custom confirm dialog with keep-open choice ───────────
                    confirm_dlg=tk.Toplevel(tr_dlg)
                    confirm_dlg.title("Confirm transfer")
                    confirm_dlg.geometry("480x320")
                    confirm_dlg.configure(bg=BG); confirm_dlg.grab_set(); confirm_dlg.resizable(False,False)
                    ch=tk.Frame(confirm_dlg,bg=NAV,padx=20,pady=12); ch.pack(fill="x")
                    tk.Label(ch,text=f"Transfer {q['ref']} to {tname}?",
                             font=(FONT,11,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
                    tk.Frame(confirm_dlg,bg=ACCENT,height=2).pack(fill="x")
                    cb=tk.Frame(confirm_dlg,bg=BG,padx=20,pady=16); cb.pack(fill="both",expand=True)
                    for bullet in [
                        "• A copy will be added to their tracker as an open query",
                        "• The site will be added to their site list if not already there",
                        "• They will see a notification when they open the app",
                    ]:
                        tk.Label(cb,text=bullet,font=(FONT,9),bg=BG,fg=TEXT2,anchor="w").pack(anchor="w")
                    tk.Frame(cb,bg=BORDER,height=1).pack(fill="x",pady=12)
                    tk.Label(cb,text="What should happen to your copy?",
                             font=(FONT,9,"bold"),bg=BG,fg=TEXT).pack(anchor="w",pady=(0,8))
                    resolve_var=tk.BooleanVar(value=True)
                    keep_row=tk.Frame(cb,bg=BG); keep_row.pack(fill="x")

                    def make_opt(parent,text,sub,val):
                        f=tk.Frame(parent,bg=CARD2,highlightthickness=1,
                                   highlightbackground=BORDER,padx=12,pady=8,cursor="hand2")
                        f.pack(side="left",padx=(0,10),fill="x",expand=True)
                        tk.Label(f,text=text,font=(FONT,9,"bold"),bg=CARD2,fg=TEXT).pack(anchor="w")
                        tk.Label(f,text=sub,font=(FONT,8),bg=CARD2,fg=TEXT2,wraplength=160).pack(anchor="w")
                        def select(v=val,fr=f):
                            resolve_var.set(v)
                            for sibling in parent.winfo_children():
                                sibling.configure(highlightbackground=BORDER)
                            fr.configure(highlightbackground=ACCENT)
                        f.bind("<Button-1>",lambda e:select())
                        for w in f.winfo_children(): w.bind("<Button-1>",lambda e:select())
                        if resolve_var.get()==val: f.configure(highlightbackground=ACCENT)
                        return f

                    make_opt(keep_row,"Mark as Resolved","Query closed on your end",True)
                    make_opt(keep_row,"Keep Open","Query stays active on both trackers",False)

                    transfer_choice=[None]
                    def do_confirm(choice,transfer_choice=transfer_choice):
                        transfer_choice[0]=choice; confirm_dlg.destroy()
                    cf=tk.Frame(confirm_dlg,bg=CARD2,padx=16,pady=10); cf.pack(fill="x",side="bottom")
                    make_btn(cf,"Cancel",lambda:do_confirm(None),"default",padx=12,pady=5).pack(side="right",padx=(6,0))
                    make_btn(cf,"Transfer →",lambda:do_confirm(resolve_var.get()),"primary",padx=14,pady=5).pack(side="right")
                    confirm_dlg.wait_window()
                    if transfer_choice[0] is None: return
                    resolve_original=transfer_choice[0]

                    try:
                        import copy

                        # ── 1. Build the new query for the target ─────────────
                        new_q=copy.deepcopy(q)
                        new_q["id"]=str(int(datetime.now().timestamp()*1000)+1)
                        new_q["status"]="Open"
                        new_q["resolved_date"]=""
                        new_q["assigned_to"]=""
                        new_q["opened"]=today_str()
                        new_q["chase_date"]=(date.today()+timedelta(days=7)).strftime("%Y-%m-%d")
                        new_q["last_by"]=self.username
                        new_q["last_date"]=today_str()
                        new_q["log"]=(
                            f"{stamp(self.username)} Transferred in from {self.username} "
                            f"(original ref {q['ref']}, opened {fmt_date(q['opened'])}). "
                            f"Site: {q['site']}. Type: {q['type']}."
                        )

                        # ── 2. Write query row directly into target Excel ─────
                        # Load or create the target workbook
                        if os.path.exists(tfile):
                            try: target_wb=openpyxl.load_workbook(tfile)
                            except PermissionError:
                                messagebox.showerror("File in use",
                                    f"{tname} is open in Excel — ask them to close it and try again.",
                                    parent=tr_dlg); return
                        else:
                            target_wb=openpyxl.Workbook()
                            if target_wb.active.title=="Sheet":
                                target_wb.active.title="Queries"

                        if "Queries" not in target_wb.sheetnames:
                            target_wb.create_sheet("Queries")
                        tws=target_wb["Queries"]

                        # Write header row if sheet is empty
                        if tws.max_row<1 or tws.cell(row=1,column=1).value is None:
                            thin_s=Side(style="thin",color="CCCCCC")
                            bdr_s=Border(left=thin_s,right=thin_s,top=thin_s,bottom=thin_s)
                            hfil_s=PatternFill("solid",fgColor="0F1B2D")
                            hfnt_s=Font(bold=True,color="FFFFFF",name=FONT,size=10)
                            col_widths=[10,12,22,22,24,18,22,18,14,10,44,12,12,14,50,30,20,20,28,14,18,14,18]
                            for ci,(col,w) in enumerate(zip(COLS,col_widths),1):
                                hc=tws.cell(row=1,column=ci,value=col)
                                hc.font=hfnt_s; hc.fill=hfil_s; hc.border=bdr_s
                                tws.column_dimensions[get_column_letter(ci)].width=w

                        # Insert new row at position 2 (top of data, below header)
                        tws.insert_rows(2)
                        thin_r=Side(style="thin",color="CCCCCC")
                        bdr_r=Border(left=thin_r,right=thin_r,top=thin_r,bottom=thin_r)
                        row_vals=[
                            new_q["id"],new_q["ref"],new_q["client"],new_q.get("fund",""),
                            new_q["site"],new_q.get("utility",""),new_q.get("meter",""),
                            new_q["type"],new_q["status"],new_q["priority"],new_q["desc"],
                            new_q["opened"],new_q.get("chase_date",""),new_q.get("resolved_date",""),
                            new_q["log"],new_q.get("address",""),new_q.get("spid",""),
                            new_q.get("serial",""),new_q.get("contact",""),new_q.get("prop_code",""),
                            new_q["last_by"],new_q["last_date"],new_q.get("assigned_to","")
                        ]
                        for ci,v in enumerate(row_vals,1):
                            cell=tws.cell(row=2,column=ci,value=v)
                            cell.font=Font(name=FONT,size=10)
                            cell.alignment=Alignment(vertical="top",wrap_text=(ci in (11,15)))
                            cell.border=bdr_r

                        # ── 3. Add site to target sites.xlsx if configured ────
                        target_sites=tracker_cfg.get("sites_file","")
                        if target_sites and os.path.exists(target_sites):
                            try:
                                swb=openpyxl.load_workbook(target_sites)
                                if "Sites" in swb.sheetnames:
                                    sws=swb["Sites"]
                                    existing=set()
                                    for srow in sws.iter_rows(min_row=2,values_only=True):
                                        if srow[0] and srow[3]:
                                            existing.add((str(srow[0]).strip(),str(srow[3]).strip()))
                                    if (q["client"].strip(),q["site"].strip()) not in existing:
                                        sws.append([
                                            q["client"],q.get("fund",""),q.get("prop_code",""),
                                            q["site"],q.get("address",""),
                                            q.get("utility",""),q.get("spid",""),
                                            q.get("serial",""),q.get("contact","")
                                        ])
                                        swb.save(target_sites)
                            except Exception:
                                pass  # site write failure is non-fatal

                        # ── 4. Write notification into target Excel ───────────
                        query_preview=(q.get("desc","") or "").strip().replace("\n"," ")
                        if len(query_preview)>120:
                            query_preview=query_preview[:117].rstrip()+"..."
                        notif_text=(
                            f"📨 TRANSFER from {self.username}  |  "
                            f"Ref: {q['ref']}  |  Client: {q['client']}  |  Site: {q['site']}  |  "
                            f"Type: {q['type']}  |  Query: {query_preview or '(no description)'}  |  Date: {today_str()}"
                        )
                        if "_Notifications" not in target_wb.sheetnames:
                            nws=target_wb.create_sheet("_Notifications")
                            nws["A1"]="Message"; nws["B1"]="Read"
                            nws["A1"].font=Font(bold=True,name=FONT)
                            nws["B1"].font=Font(bold=True,name=FONT)
                        else:
                            nws=target_wb["_Notifications"]
                        next_row=nws.max_row+1
                        nws.cell(row=next_row,column=1,value=notif_text)
                        nws.cell(row=next_row,column=2,value="N")

                        # Save target workbook
                        target_wb.save(tfile)

                        # ── 5. Handle original query based on user's choice ───
                        if resolve_original:
                            q["status"]="Resolved"
                            q["resolved_date"]=today_str()
                            q["log"]+=f" | {stamp(self.username)} Transferred to '{tname}' and resolved here. Ref {q['ref']} now open in their tracker."
                        else:
                            q["log"]+=f" | {stamp(self.username)} Transferred a copy to '{tname}'. This query remains open here."
                        q["last_by"]=self.username; q["last_date"]=today_str()
                        self._save_queries()
                        self._refresh_table(); self._show_daily_banner()
                        tr_dlg.destroy(); dlg.destroy()
                        status_note="Marked as Resolved in your tracker" if resolve_original else "Still open in your tracker"
                        messagebox.showinfo("Transferred ✓",
                            f"{q['ref']} transferred to {tname}.\n\n"
                            f"• Added to their query list as open\n"
                            f"• Site added to their site list if new\n"
                            f"• They'll see a notification when they open the app\n"
                            f"• {status_note}",
                            parent=self)

                    except PermissionError:
                        messagebox.showerror("File in use",
                            f"{tname} is open in Excel — ask them to close it and try again.",
                            parent=tr_dlg)
                    except Exception as e:
                        messagebox.showerror("Transfer failed",
                            f"Something went wrong:\n\n{e}",parent=tr_dlg)

                card.bind("<Button-1>",lambda e,f=do_transfer:f())
                for w in card.winfo_children():
                    w.bind("<Button-1>",lambda e,f=do_transfer:f())
                card.bind("<Enter>",lambda e,c=card:c.configure(highlightbackground=ACCENT))
                card.bind("<Leave>",lambda e,c=card:c.configure(highlightbackground=BORDER))

            ft_tr=tk.Frame(tr_dlg,bg=CARD2,padx=16,pady=12); ft_tr.pack(fill="x",side="bottom")
            make_btn(ft_tr,"Cancel",tr_dlg.destroy,"default",padx=12,pady=6).pack(side="right")

        def show_timeline():
            """Show the full chain of queries linked by type changes."""
            # Find all queries in this chain by tracing log references
            chain=[]
            visited=set()
            def find_chain(ref):
                if ref in visited: return
                visited.add(ref)
                for qx in self.queries:
                    if qx["ref"]==ref:
                        chain.append(qx)
                        # Look for the "Continued from X" pattern in the log
                        for entry in public_log_entries(qx.get("log", "")):
                            m=re.search(r"Continued from ([A-Z]{2}-\d+)",entry)
                            if m: find_chain(m.group(1))
                        # Also look for queries that continued FROM this one
                        for qy in self.queries:
                            for entry in public_log_entries(qy.get("log", "")):
                                if f"Continued from {ref}" in entry:
                                    find_chain(qy["ref"])
                        break
            find_chain(q["ref"])
            # Sort by opened date
            chain.sort(key=lambda x:x.get("opened",""))
            if len(chain)<=1:
                messagebox.showinfo("No timeline",
                    f"{q['ref']} has not been linked to any other queries via a type change.",parent=dlg)
                return

            tl=tk.Toplevel(dlg); tl.title(f"Query timeline — {q['ref']}")
            tl.geometry("620x480"); tl.configure(bg=BG); tl.grab_set(); tl.resizable(False,True)
            hdr_tl=tk.Frame(tl,bg=NAV,padx=20,pady=14); hdr_tl.pack(fill="x")
            tk.Label(hdr_tl,text="⏱  Query timeline",font=(FONT,12,"bold"),bg=NAV,fg=TEXT).pack(anchor="w")
            tk.Label(hdr_tl,text=f"{q['client']}  ·  {q['site']}  ·  {len(chain)} linked queries",
                     font=(FONT,9),bg=NAV,fg=TEXT2).pack(anchor="w",pady=(4,0))
            tk.Frame(tl,bg=ACCENT,height=2).pack(fill="x")

            scroll_outer=tk.Frame(tl,bg=BG); scroll_outer.pack(fill="both",expand=True)
            inner,_=scrollable_frame(scroll_outer)

            for i,qx in enumerate(chain):
                sc,spill=S_COLORS.get(qx["status"],("#94A3B8","#1E2D45"))
                # Connector line between cards
                if i>0:
                    conn=tk.Frame(inner,bg=BG); conn.pack(fill="x",pady=0)
                    tk.Label(conn,text="     │",font=(FONT,10),bg=BG,fg=BORDER).pack(anchor="w")
                    tk.Label(conn,text="     ↓  type changed",font=(FONT,8),bg=BG,fg=MUTED).pack(anchor="w")
                    tk.Label(conn,text="     │",font=(FONT,10),bg=BG,fg=BORDER).pack(anchor="w")

                card=tk.Frame(inner,bg=CARD2,highlightthickness=1,
                              highlightbackground=ACCENT if qx["ref"]==q["ref"] else BORDER,
                              padx=16,pady=12); card.pack(fill="x",pady=(0,0))
                top=tk.Frame(card,bg=CARD2); top.pack(fill="x")
                tk.Label(top,text=qx["ref"],font=(FONT,11,"bold"),bg=CARD2,
                         fg=ACCENT2 if qx["ref"]==q["ref"] else TEXT).pack(side="left")
                pill=tk.Frame(top,bg=spill); pill.pack(side="left",padx=8)
                tk.Label(pill,text=f"  {qx['status']}  ",font=(FONT,8,"bold"),
                         bg=spill,fg=sc,pady=2).pack()
                if qx["ref"]==q["ref"]:
                    tk.Label(top,text="  ← current",font=(FONT,8),bg=CARD2,fg=MUTED).pack(side="left")
                tk.Label(card,text=qx["type"],font=(FONT,10),bg=CARD2,fg=TEXT2).pack(anchor="w",pady=(2,0))
                info=tk.Frame(card,bg=CARD2); info.pack(fill="x",pady=(6,0))
                tk.Label(info,text=f"Opened {fmt_date(qx['opened'])}",
                         font=(FONT,8),bg=CARD2,fg=MUTED).pack(side="left")
                if qx.get("resolved_date"):
                    tk.Label(info,text=f"  ·  Resolved {fmt_date(qx['resolved_date'])}",
                             font=(FONT,8),bg=CARD2,fg=SUCCESS).pack(side="left")
                # Click to open
                def _open(e,qx=qx):
                    tl.destroy(); dlg.destroy(); self._open_detail_query(qx)
                card.bind("<Button-1>",_open); card.bind("<Enter>",lambda e,c=card:c.configure(highlightbackground=ACCENT2))
                card.bind("<Leave>",lambda e,c=card,qx=qx:c.configure(
                    highlightbackground=ACCENT if qx["ref"]==q["ref"] else BORDER))
                for w in list(card.winfo_children())+[pill]+list(pill.winfo_children()):
                    try: w.bind("<Button-1>",_open)
                    except: pass

            ft_tl=tk.Frame(tl,bg=CARD2,padx=16,pady=12); ft_tl.pack(fill="x",side="bottom")
            make_btn(ft_tl,"Close",tl.destroy,"default",padx=12,pady=6).pack(side="right")

        # Determine if this query is part of a type-change chain
        is_chained=(
            any(f"Continued from {q['ref']}" in qx["log"] for qx in self.queries)
            or bool(re.search(r"Continued from [A-Z]{2}-\d+",q.get("log","")))
        )

        footer=tk.Frame(dlg,bg=CARD2,highlightthickness=1,highlightbackground=BORDER,padx=16,pady=10)
        footer.pack(fill="x",side="bottom")
        tk.Frame(dlg,bg=BORDER,height=1).pack(fill="x",side="bottom")

        # Row 1 — primary actions
        frow1=tk.Frame(footer,bg=CARD2); frow1.pack(fill="x",pady=(0,6))
        make_btn(frow1,"Save changes",save_changes,"success",padx=14,pady=5).pack(side="right")
        make_btn(frow1,"Cancel",dlg.destroy,"default",padx=12,pady=5).pack(side="right",padx=(0,6))

        # Row 2 — secondary actions
        frow2=tk.Frame(footer,bg=CARD2); frow2.pack(fill="x")
        make_btn(frow2,"Delete",delete_query,"danger",padx=10,pady=5).pack(side="left")
        make_btn(frow2,"✉ Email",draft_email,"default",padx=10,pady=5).pack(side="left",padx=(6,0))
        make_btn(frow2,"⧉ Duplicate",lambda:self._dup_query(q,dlg),"default",padx=10,pady=5).pack(side="left",padx=(6,0))
        make_btn(frow2,"⇄ Change type",change_type,"default",padx=10,pady=5).pack(side="left",padx=(6,0))
        if getattr(self,"linked_trackers",[]):
            make_btn(frow2,"↗ Transfer",transfer_query,"default",padx=10,pady=5).pack(side="left",padx=(6,0))
        if is_chained or re.search(r"Continued from [A-Z]{2}-\d+",q.get("log","")):
            make_btn(frow2,"⏱ Timeline",show_timeline,"active",padx=10,pady=5).pack(side="left",padx=(6,0))

    def _open_related_queries_window(self, q):
        """Open a separate window showing all related queries for the same site."""
        related_dlg = tk.Toplevel(self)
        related_dlg.title(f"Related Queries - {q['site']} ({q['client']})")
        related_dlg.geometry("800x600")
        related_dlg.configure(bg=BG)
        related_dlg.grab_set()
        related_dlg.resizable(True, True)
        
        # Header
        hdr = tk.Frame(related_dlg, bg=NAV, padx=20, pady=16)
        hdr.pack(fill="x")
        tk.Frame(related_dlg, bg=ACCENT, height=2).pack(fill="x")
        tk.Label(hdr, text=f"🔗 Related Queries for {q['site']}", font=(FONT, 13, "bold"), bg=NAV, fg="white").pack(anchor="w")
        tk.Label(hdr, text=f"Client: {q['client']} • Current query: {q['ref']}", font=(FONT, 9), bg=NAV, fg=TEXT2).pack(anchor="w", pady=(4, 0))
        
        # Main content
        content = tk.Frame(related_dlg, bg=BG, padx=20, pady=16)
        content.pack(fill="both", expand=True)
        
        # Get related queries for the same site
        related_queries = [rq for rq in self.queries if rq["site"] == q["site"] and rq["id"] != q["id"]]
        
        if not related_queries:
            tk.Label(content, text="No other queries for this site.", font=(FONT, 10), bg=BG, fg=MUTED).pack(anchor="w", pady=20)
        else:
            # Sort: open queries first (by priority: High, Medium, Low), then resolved (most recent first)
            def sort_key(rq):
                if rq["status"] != "Resolved":
                    pri_order = {"High": 0, "Medium": 1, "Low": 2}
                    return (0, pri_order.get(rq.get("priority", "Low"), 3), rq.get("opened", ""))
                else:
                    return (1, rq.get("resolved_date", rq.get("last_date", "")), rq.get("opened", ""))
            
            related_queries.sort(key=sort_key)
            
            # Summary stats
            open_count = sum(1 for rq in related_queries if rq["status"] != "Resolved")
            resolved_count = len(related_queries) - open_count
            
            stats_frame = tk.Frame(content, bg=BG)
            stats_frame.pack(fill="x", pady=(0, 16))
            tk.Label(stats_frame, text=f"Total: {len(related_queries)} • Open: {open_count} • Resolved: {resolved_count}", 
                    font=(FONT, 9, "bold"), bg=BG, fg=TEXT).pack(anchor="w")
            
            # Scrollable list
            list_frame = tk.Frame(content, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
            list_frame.pack(fill="both", expand=True)
            
            canvas = tk.Canvas(list_frame, bg=CARD, highlightthickness=0)
            scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg=CARD)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack canvas and scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Add related queries
            for rq in related_queries:
                row = tk.Frame(scrollable_frame, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
                row.pack(fill="x", padx=10, pady=5)
                
                # Status indicator
                sc, sbg = S_COLORS.get(rq["status"], (TEXT2, CARD2))
                status_pill = tk.Label(row, text=f"  {rq['status']}  ", font=(FONT, 7, "bold"), bg=sbg, fg=sc, padx=4, pady=2)
                status_pill.pack(side="left")
                
                # Priority indicator
                pc, pbg = P_COLORS.get(rq.get("priority", "Low"), ("#94A3B8", "#1E2D45"))
                pri_pill = tk.Label(row, text=f"  {rq.get('priority', 'Low')}  ", font=(FONT, 7), bg=pbg, fg=pc, padx=3, pady=2)
                pri_pill.pack(side="left", padx=(4, 0))
                
                # Query info
                info_frame = tk.Frame(row, bg=CARD)
                info_frame.pack(side="left", fill="x", expand=True, padx=(8, 0))
                
                ref_label = tk.Label(info_frame, text=rq["ref"], font=(FONT, 10, "bold"), bg=CARD, fg=ACCENT2, cursor="hand2")
                ref_label.pack(anchor="w")
                
                details = f"{rq['type']} • {rq['utility']} • {fmt_date(rq.get('opened', ''))}"
                if rq.get("assigned_to"):
                    details += f" • Assigned to: {rq['assigned_to']}"
                type_label = tk.Label(info_frame, text=details, font=(FONT, 8), bg=CARD, fg=TEXT2)
                type_label.pack(anchor="w")
                
                # Action date if exists
                if rq.get("chase_date") and rq["status"] != "Resolved":
                    chase_text = f"Action: {fmt_date(rq['chase_date'])}"
                    if rq["chase_date"] <= today_str():
                        chase_text += " ⚠ OVERDUE"
                    chase_label = tk.Label(info_frame, text=chase_text, font=(FONT, 8), bg=CARD, fg=DANGER if "OVERDUE" in chase_text else TEXT2)
                    chase_label.pack(anchor="w")
                
                # Make the entire row clickable
                def open_related_detail(rqx=rq):
                    related_dlg.destroy()
                    self._open_detail_query(rqx)
                
                for widget in [row, ref_label, type_label]:
                    widget.bind("<Button-1>", lambda e, rq=rq: open_related_detail(rq))
                    widget.bind("<Enter>", lambda e, r=row: r.configure(highlightbackground=ACCENT))
                    widget.bind("<Leave>", lambda e, r=row: r.configure(highlightbackground=BORDER))
        
        # Close button
        btn_frame = tk.Frame(related_dlg, bg=CARD2, highlightthickness=1, highlightbackground=BORDER, padx=16, pady=12)
        btn_frame.pack(fill="x", side="bottom")
        make_btn(btn_frame, "Close", related_dlg.destroy, "default", padx=12, pady=6).pack(side="right")

    def _dup_query(self,q,parent_dlg):
        parent_dlg.destroy()
        self._open_add_dialog(copy_from=q)

    def _apply_escalation_rules(self):
        """Apply automated escalation rules to queries. Returns number of escalated queries."""
        rules = self.cfg.get("escalation_rules", [])
        if not rules:
            return 0
        
        today = date.today()
        escalated_count = 0
        
        for query in self.queries:
            if query["status"] == "Resolved":
                continue
                
            # Calculate query age
            try:
                opened_date = datetime.strptime(query["opened"], "%Y-%m-%d").date()
                age_days = (today - opened_date).days
            except:
                continue
            
            current_priority = query.get("priority", "Low")
            
            # Check each rule
            for rule in rules:
                if (rule["priority"] == current_priority and 
                    age_days >= rule["days_old"] and 
                    rule["new_priority"] != current_priority):
                    
                    # Apply escalation
                    old_priority = current_priority
                    query["priority"] = rule["new_priority"]
                    query["last_by"] = "Auto-escalation"
                    query["last_date"] = today_str()
                    query["log"] += f" | {stamp('Auto-escalation')} Priority escalated from {old_priority} to {rule['new_priority']} (age: {age_days}d)"
                    escalated_count += 1
                    break  # Only apply first matching rule
        
        if escalated_count > 0:
            self._save_queries()
        
        return escalated_count

if __name__=="__main__":
    app=QueryTrackerApp()
    app.mainloop()
    app.mainloop()
